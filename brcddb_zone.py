"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Zone utilities.

**Public Methods**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| add_zone_worksheet    | Adds a worksheet to the workbook created with create_zone_workbook()                      |
+-----------------------+-------------------------------------------------------------------------------------------+
| alias_compare         | Compares two aliases                                                                      |
+-----------------------+-------------------------------------------------------------------------------------------+
| create_zone_workbook  | Creates an Excel workbook with the Instructions sheet.                                    |
+-----------------------+-------------------------------------------------------------------------------------------+
| is_alias_match        | Simple True/False comparison between two aliases                                          |
+-----------------------+-------------------------------------------------------------------------------------------+
| zone_compare          | Compares two zones                                                                        |
+-----------------------+-------------------------------------------------------------------------------------------+
| is_zone_match         | Simple True/False comparison between two zones                                            |
+-----------------------+-------------------------------------------------------------------------------------------+
| cfg_compare           | Compares two zone configurations. Validates membership only. Does not compare             |
|                       | individual members                                                                        |
+-----------------------+-------------------------------------------------------------------------------------------+
| is_cfg_match          | Simple True/False comparison between two configurations                                   |
+-----------------------+-------------------------------------------------------------------------------------------+
| zone_type             | Returns the zone type (User defined peer, Target driven peer, or Standard) in plain text  |
+-----------------------+-------------------------------------------------------------------------------------------+
| eff_zoned_to_wwn      | Finds all WWNs in the effective zone that are zoned to wwn.                               |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 26 Dec 2024   | Added create_zone_workbook() and add_zone_worksheet()                                 |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 01 Mar 2025   | Updated "Instructions" sheet. Added support for 'Comments' in add_zone_worksheet()    |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024, 2025 Consoli Solutions, LLC'
__date__ = '01 Mar 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.3'

import collections
import openpyxl.utils.cell as xl
from openpyxl.worksheet.datavalidation import DataValidation
import brcdapi.gen_util as gen_util
import brcdapi.log as brcdapi_log
import brcdapi.util as brcdapi_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.excel_util as excel_util
import brcddb.brcddb_common as brcddb_common

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_hdr1_font = excel_fonts.font_type('hdr_1')
_hdr2_font = excel_fonts.font_type('hdr_2')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_align_wrap_r = excel_fonts.align_type('wrap_right')
_border_thin = excel_fonts.border_type('thin')

_valid_zone_items_d = dict(
    Zone_Object=collections.OrderedDict(comment=True, zone_cfg=True, peer_zone=True, zone=True, alias=True),
    Action=collections.OrderedDict(add_mem=True, remove_mem=True, copy=True, create=True, delete=True, purge=True,
                                   full_purge=True, ignore=True, rename=True, replace=True),
    Match=collections.OrderedDict(exact=True, wild=True, regex_m=True, regex_s=True)
)

# Zone_Object
_buf = ','.join([str(_key) for _key in _valid_zone_items_d['Zone_Object'].keys()])
_zone_object_dv = DataValidation(type='list', formula1='"' + _buf + '"', allow_blank=True)
_zone_object_dv.errorTitle = 'Invalid Entry'
_zone_object_dv.error = 'Valid values are: ' + _buf.replace(',', ', ')
_zone_object_dv.showErrorMessage = True

# Action
_buf = ','.join([str(_key) for _key in _valid_zone_items_d['Action'].keys()])
_zone_action_dv = DataValidation(type='list', formula1='"' + _buf + '"', allow_blank=True)
_zone_action_dv.errorTitle = 'Invalid Entry'
_zone_action_dv.error = 'Valid values are: ' + _buf.replace(',', ', ')
_zone_action_dv.showErrorMessage = True

# Match
_buf = ','.join([str(_key) for _key in _valid_zone_items_d['Match'].keys()])
_zone_match_dv = DataValidation(type='list', formula1='"' + _buf + '"', allow_blank=True)
_zone_match_dv.errorTitle = 'Invalid Entry'
_zone_match_dv.error = 'Valid values are: ' + _buf.replace(',', ', ')
_zone_match_dv.showErrorMessage = True
_zone_dv_l = (_zone_object_dv, _zone_action_dv, _zone_match_dv)
"""_zone_worksheet_hdr is as follows:
+-------+-------------------------------------------------------+
| Key   | Value                                                 |
+=======+=======================================================+
| col   | Column width                                          |
+-------+-------------------------------------------------------+
| dv    | Data validation object to be attached to each cell    |
+-------+-------------------------------------------------------+
"""
_zone_worksheet_hdr = collections.OrderedDict({
    'Zone_Object': dict(col=14, dv=_zone_object_dv),
    'Action': dict(col=12, dv=_zone_action_dv),
    'Name': dict(col=30),
    'Match': dict(col=12, dv=_zone_match_dv),
    'Member': dict(col=30),
    'Principal Member': dict(col=30),
    'Comments': dict(col=110),
})
_zone_instructions_col_l = (14, 70)
# _zonecfg_instructions is used in create_zone_workbook(). Although not used when this was written, 'border' and 'link'
# are also added to the cell
_zonecfg_instructions = [
    [dict(t='Intended for use with zone_config.py', font=_hdr1_font, span=2)],
    list(),
    [dict(t='Rules, Tips, and Guidelines', font=_bold_font, span=2)],
    list(),
    [
        dict(t='1', align=_align_wrap_c),
        dict(t='Typical use is to copy the "zone" sheet and modify it to meet your needs. Do not modify the headers: '
               '“Type”, “Action”, “Name”, “Member”, and “Principal Member”. All other columns are ignored so that you '
               'can add columns if needed.'),
    ],
    [
        dict(t='2', align=_align_wrap_c),
        dict(t='If “Type”, “Action”, or “Name” is empty, the previous entry is assumed. See "sample" sheet for an '
               'example.'),
    ],
    [
        dict(t='3', align=_align_wrap_c),
        dict(t='If “Zone_Object” is "comment", the row is ignored. This means you may merge the remaining cells in the '
               'row. Since comments are completely ignored, it has no effect on previous stored entries.'),
    ],
    [
        dict(t='4', align=_align_wrap_c),
        dict(t='Hidden rows are not read.'),
    ],
    [
        dict(t='5', align=_align_wrap_c),
        dict(t='Only one member per cell'),
    ],
    [
        dict(t='6', align=_align_wrap_c),
        dict(t='The zoning database is read from the switch or input file before taking any actions. This means that '
               'members and zone objects do not have to be in this workbook as long as they are already in the zone '
               'database.'),
    ],
    [
        dict(t='7', align=_align_wrap_c),
        dict(t='Use report.py for detailed zone analysis.'),
    ],
    [
        dict(t='8', align=_align_wrap_c),
        dict(t='The "purge" and "full_purge" actions are intended to help decommission server clusters and storage '
               'arrays, but can also be used for zone cleanup.'),
    ],
    [
        dict(t='9', align=_align_wrap_c),
        dict(t='For detailed help:\npy zone_config.py -eh'),
    ],
    list(),
    [dict(t='Actions', font=_bold_font, span=2)],
    list(),
    [
        dict(t='add_mem'),
        dict(t='Adds a member to an alias, zone, or zone configuration membership list.'),
    ],
    [
        dict(t='copy'),
        dict(t='Copies an alias, zone, or zone configuration'),
    ],
    [
        dict(t='create'),
        dict(t='Creates an alias, zone, or zone configuration.'),
    ],
    [
        dict(t='delete'),
        dict(t='Deletes an alias, zone, or zone configuration.'),
    ],
    [
        dict(t='purge'),
        dict(t='Supported for alias and zone actions only. The value in the "Member" cell is used to define the alias '
               'or zone for the purge action to act on. This is the equivalent to the CLI zoneobjectexpunge command.'),
    ],
    [
        dict(t='alias', align=_align_wrap_r),
        dict(t='Removes the alias, d,i, or WWN in the "Member" cell from any zone it is used in. Typically, a '
               'full_purge action is used to remove no longer needed aliases to avoid runt zones.'),
    ],
    [
        dict(t='zone', align=_align_wrap_r),
        dict(t='Removes the zone in the "Member" cell from any zone configuration it is used in.'),
    ],
    [
        dict(t='full_purge'),
        dict(t='For zone objects only. Purges all members, then purges the zone object.The value in the "Member" cell '
               'is used to define the alias, zone, or zone configuration for the full_purge action to act on.'),
    ],
    [
        dict(t='alias', align=_align_wrap_r),
        dict(t='Executes a purge action on the alias. A full_purge is then executed on any zone affected by the alias '
               'purge that has no remaining members. See  Ignore  for additional details.'),
    ],
    [
        dict(t='zone', align=_align_wrap_r),
        dict(t='After executing a purge action on the zone, any affected zone configuration with no remaining members '
               'is deleted. Any alias not used in any other zone is also deleted.'),
    ],
    [
        dict(t='zone_cgf', align=_align_wrap_r),
        dict(t='Deletes the zone configuration. A full_purge action is executed on any affected zone that is not used '
               'in any other zone configuration.'),
    ],
    [
        dict(t='ignore'),
        dict(t='Typically not used. Run "zone_config.py -eh" for details.'),
    ],
    [
        dict(t='remove_mem'),
        dict(t='Removes members from the object specified in the "Name" cell. The same rules regarding "Members" and '
               '"Principal Members" in add_mem apply.'),
    ],
    list(),
    [dict(t='Match', font=_bold_font, span=2)],
    list(),
    [
        dict(),
        dict(t='If the cell is empty, an exact match is performed')
    ],
    [
        dict(t='exact'),
        dict(t='Perform an exact match.'),
    ],
    [
        dict(t='wild'),
        dict(t='The cell in the "Name" column may contain * and ? For wild card matching. Applies to delete, '
               'remove_mem, purge, full_purge, and ignore actions only.'),
    ],
    [
        dict(t='regex_m'),
        dict(t='The cell in the "Name" column may contain ReGex matching text. Applies to delete, remove_mem, purge, '
               'full_purge, and ignore actions only.'),
    ],
    [
        dict(t='regex_s'),
        dict(t='The cell in the "Name" column may contain ReGex searching text. Applies to delete, remove_mem, purge, '
               'full_purge, and ignore actions only.'),
    ],
]


def add_zone_worksheet(wb, sheet_name, content_l, sheet_i=0):
    """Adds a zone configuration worksheet

    content_l is a list of dictionaries. Each dictionary is:

    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Key               | Value Type        | Value Description                                                     |
    +===================+===================+=======================================================================+
    | Zone_Object       | str, None         | Value for "Zone_Object" column. When the value is "comment", all      |
    |                   |                   | remaining columns are merged. The contents are the value in           |
    |                   |                   | "Comments". All other key/value pairs are ignored.                    |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Action            | str, None         | Value for "Action" column                                             |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Name              | str, None         | Value for "Name" column                                               |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Match             | str, None         | Value for "Match" column                                              |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Member            | str, list, None   | Value(s) for "Member" column                                          |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Principal Member  | str, list, None   | Value(s) for "Principal Member" column                                |
    +-------------------+-------------------+-----------------------------------------------------------------------+
    | Comments          | str, None         | Value(s) for "Comments" column                                        |
    +-------------------+-------------------+-----------------------------------------------------------------------+

    :param wb: Excel workbook
    :type wb: Workbook object
    :param sheet_name: Name of the worksheet
    :type sheet_name: str
    :param content_l: Content of the worksheet. See function description for details
    :type content_l: dict
    :param sheet_i: Index where sheet is to be inserted
    :rtype: None
    """
    global _zone_dv_l, _zone_worksheet_hdr, _std_font, _hdr2_font, _align_wrap, _border_thin, _valid_zone_items_d

    # Create the worksheet and do some basic setup
    sheet = wb.create_sheet(index=sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    for dv in _zone_dv_l:
        sheet.add_data_validation(dv)
    row = col = 1
    for key, value_d in _zone_worksheet_hdr.items():
        sheet.column_dimensions[xl.get_column_letter(col)].width = value_d['col']
        excel_util.cell_update(sheet, row, col, key, font=_hdr2_font, align=_align_wrap, border=_border_thin)
        col += 1
    row += 1

    # Add the content
    for zone_d in content_l:
        next_row, col = row, 1
        for key, value_d in _zone_worksheet_hdr.items():
            zone_item = zone_d.get(key)

            # Check the data validation
            if zone_item is not None:
                valid_d = _valid_zone_items_d.get(key)
                if isinstance(valid_d, dict):
                    if not valid_d.get(zone_item, False):
                        brcdapi_log.exception('Data validation: ' + str(zone_item) + ' is not valid for ' + str(key),
                                              echo=True)
                        break

            # Is it a comment?
            if key == 'Zone_Object' and zone_item == 'comment':
                excel_util.cell_update(sheet, row, col, zone_item)
                excel_util.cell_update(sheet, row, col+1, '\n'.join(gen_util.convert_to_list(zone_d.get('Comments'))))
                x = len(_zone_worksheet_hdr)
                sheet.merge_cells(start_row=row, start_column=col+1, end_row=row, end_column=len(_zone_worksheet_hdr))
                break

            # Add the zone item to the worksheet
            if isinstance(zone_item, (list, tuple)):
                mem_row = row
                for mem in zone_item:
                    excel_util.cell_update(sheet, mem_row, col, mem)
                    mem_row += 1
                next_row = max(next_row, mem_row)
            else:
                excel_util.cell_update(sheet, row, col, zone_item)
            col += 1
        row = next_row + 1

    # Format all the cells + 12 more rows, but at least 150 rows. WARNING: At least one row is needed because the
    # data validations added to the sheet must be present in at least once cell. Otherwise, Excel cannot open the file
    # without an error.
    for next_row in range(2, max(150, row + 12)):
        col = 1
        for value_d in _zone_worksheet_hdr.values():
            excel_util.cell_update(
                sheet,
                next_row,
                col,
                None,
                font=_std_font,
                align=_align_wrap,
                border=_border_thin,
                dv=value_d.get('dv')
            )
            col += 1

    return


def alias_compare(base_alias_obj, comp_alias_obj):
    """Compares two aliases
    
    :param base_alias_obj: Alias object to compare against
    :type base_alias_obj: brcddb.classes.zone.AliasObj
    :param comp_alias_obj: New alias object
    :type comp_alias_obj: brcddb.classes.zone.AliasObj
    :return add_members: Members is comp_alias_obj that do not exist in base_alias_obj
    :rtype add_members: list
    :return del_members: Members is base_alias_obj that do not exist in comp_alias_obj
    :rtype del_members: list
    """
    bl, cl = base_alias_obj.r_members(), comp_alias_obj.r_members()
    return [mem for mem in cl if mem not in bl], [mem for mem in bl if mem not in cl]


def create_zone_workbook():
    """Creates an Excel workbook with the Instructions sheet.

    :return: Excel workbook
    :rtype: Workbook object
    """
    global _zone_instructions_col_l, _zonecfg_instructions, _std_font, _align_wrap

    # Create the workbook and set up the Instructions sheet
    wb = excel_util.new_report()
    sheet = wb.create_sheet(index=0, title='Instructions')
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    for col in range(0, len(_zone_instructions_col_l)):
        sheet.column_dimensions[xl.get_column_letter(col+1)].width = _zone_instructions_col_l[col]

    # Add the content to the Instructions sheet
    row = 1
    for row_l in _zonecfg_instructions:
        col = 1
        for col_d in row_l:
            excel_util.cell_update(
                sheet,
                row,
                col,
                col_d.get('t'),
                font=col_d.get('font', _std_font),
                align=col_d.get('align', _align_wrap),
                border=col_d.get('border'),
                link=col_d.get('link')
            )
            col_span = col_d.get('span', 1)
            if col_span > 1:
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + col_span-1)
            col += col_span
        row += 1

    return wb


def is_alias_match(base_alias_obj, comp_alias_obj):
    """Simple True/False comparison between two aliases

    :param base_alias_obj: Alias object to compare against
    :type base_alias_obj: brcddb.classes.zone.AliasObj
    :param comp_alias_obj: New alias object
    :type comp_alias_obj: brcddb.classes.zone.AliasObj
    :return: True if the aliases are the same, otherwise False
    :rtype: bool
    """
    bl, cl = alias_compare(base_alias_obj, comp_alias_obj)
    return True if len(bl) + len(cl) == 0 else False


def zone_compare(base_zone_obj, comp_zone_obj):
    """Compares two zones

    :param base_zone_obj: Zone object to compare against
    :type base_zone_obj: brcddb.classes.zone.ZoneObj
    :param comp_zone_obj: New zone object
    :type comp_zone_obj: brcddb.classes.zone.ZoneObj
    :return type_flag: True if the zone type is the same, False if not
    :rtype type_flag: bool
    :return add_members: Members is comp_zone_obj that do not exist in base_zone_obj
    :rtype add_members: list
    :return del_members: Members is base_zone_obj that do not exist in comp_zone_obj
    :rtype del_members: list
    :return add_pmembers: Principal members is comp_zone_obj that do not exist in base_zone_obj
    :rtype add_pmembers: list
    :return del_pmembers: Principal members is base_zone_obj that do not exist in comp_zone_obj
    :rtype del_pmembers: list
    """
    bl, cl = base_zone_obj.r_members(), comp_zone_obj.r_members()
    pbl, pcl = base_zone_obj.r_pmembers(), comp_zone_obj.r_pmembers()
    return True if base_zone_obj.r_type() != comp_zone_obj.r_type() else False, \
        [mem for mem in cl if mem not in bl], \
        [mem for mem in bl if mem not in cl], \
        [mem for mem in pcl if mem not in pbl], \
        [mem for mem in pbl if mem not in pcl]
            

def is_zone_match(base_zone_obj, comp_zone_obj):
    """Simple True/False comparison between two zones

    :param base_zone_obj: Zone object to compare against
    :type base_zone_obj: brcddb.classes.zone.ZoneObj
    :param comp_zone_obj: New alias object
    :type comp_zone_obj: brcddb.classes.zone.ZoneObj
    :return: True if the zones are the same, otherwise False
    :rtype: bool
    """
    t, bl, cl, pbl, pcl = zone_compare(base_zone_obj, comp_zone_obj)
    return True if t and len(bl) + len(cl) + len(pbl) + len(pcl) == 0 else False


def cfg_compare(base_cfg_obj, comp_cfg_obj):
    """Compares two zone configurations. Validates membership only. Does not compare individual members

    :param base_cfg_obj: Alias object to compare against
    :type base_cfg_obj: brcddb.classes.zone.AliasObj
    :param comp_cfg_obj: New cfg object
    :type comp_cfg_obj: brcddb.classes.zone.AliasObj
    :return add_members: Members is comp_cfg_obj that do not exist in base_cfg_obj
    :rtype add_members: list
    :return del_members: Members is base_cfg_obj that do not exist in comp_cfg_obj
    :rtype del_members: list
    """
    bl, cl = base_cfg_obj.r_members(), comp_cfg_obj.r_members()
    return [mem for mem in cl if mem not in bl], [mem for mem in bl if mem not in cl]


def is_cfg_match(base_cfg_obj, comp_cfg_obj):
    """Simple True/False comparison between two configurations

    :param base_cfg_obj: Alias object to compare against
    :type base_cfg_obj: brcddb.classes.zone.AliasObj
    :param comp_cfg_obj: New cfg object
    :type comp_cfg_obj: brcddb.classes.zone.AliasObj
    :return: True if the configurations are the same, otherwise False
    :rtype: bool
    """
    bl, cl = cfg_compare(base_cfg_obj, comp_cfg_obj)
    return True if len(bl) + len(cl) == 0 else False


def zone_type(zone_obj, num_flag=False):
    """Returns the zone type (User defined peer, Target driven peer, or Standard) in plain text

    :param zone_obj: Zone Object
    :type zone_obj: brcddb_classes.ZoneObj
    :param num_flag: If True, append (type) where type is the numerical zone type returned from the API
    :type num_flag: bool
    :return: Zone type
    :rtype: str
    """
    z_type = zone_obj.r_type()
    try:
        buf = brcddb_common.zone_conversion_tbl['zone-type'][z_type]
        return buf + '(' + str(z_type) + ')' if num_flag else buf
    except KeyError:
        return 'Unknown (' + str(z_type) + ')'


def eff_zoned_to_wwn(fab_obj, wwn, target=False, initiator=False, all_types=False):
    """Finds all WWNs in the effective zone that are zoned to wwn.

    WARNING: if all_types == True, the FC4 type in the login data is not checked and therefore, all WWNs will be
    returned. When filtering on "target" or "initiator", there must be something logged in and the login data from the
    name server must have been retrieved in order to check the login type.

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param wwn: WWN to look for
    :type wwn: str
    :param target: If True, include all online targets
    :type target: bool
    :param initiator: If True, include any online port that is not a target
    :type initiator: bool
    :param all_types: If True, include all ports, online or offline.
    :type all_types: bool
    :return: Dictionary - Key: WWN of device zoned to wwn. Value is the list of zone names common to both WWNs.
    :rtype: dict
    """
    rd, zones_for_wwn_d = dict(), dict()

    for zone in fab_obj.r_eff_zones_for_wwn(wwn):  # Used to find common zones with members
        zones_for_wwn_d.update({zone: True})

    for zone_obj in fab_obj.r_eff_zone_objects_for_wwn(wwn):
        if zone_obj.r_type() == brcddb_common.ZONE_STANDARD_ZONE:
            mem_l = zone_obj.r_members()
        else:
            mem_l = zone_obj.r_members() if wwn in zone_obj.r_pmembers() else zone_obj.r_pmembers()
        for mem in mem_l:
            if mem == wwn or mem in rd:
                continue
            if all_types:
                rd.update({mem: [m for m in fab_obj.r_eff_zones_for_wwn(mem) if bool(zones_for_wwn_d.get(m))]})
                continue
            login_obj = fab_obj.r_login_obj(mem)
            if login_obj is None:
                continue
            fc4 = login_obj.r_get(brcdapi_util.bns_fc4_features)
            if fc4 is None:
                continue
            if target and 'target' in fc4.lower():
                rd.update({mem: [m for m in fab_obj.r_eff_zones_for_wwn(mem) if bool(zones_for_wwn_d.get(m))]})
            elif initiator and 'initiator' in fc4.lower():
                # Used "elif" because if both target & initiator was specified, it will already be in rd
                rd.update({mem: [m for m in fab_obj.r_eff_zones_for_wwn(mem) if bool(zones_for_wwn_d.get(m))]})

    return rd
