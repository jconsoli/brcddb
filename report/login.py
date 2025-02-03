"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Includes methods to create device login page

**Public Methods & Data**

    +-----------------------+---------------------------------------------------------------------------------------+
    | Method                | Description                                                                           |
    +=======================+=======================================================================================+
    | login_page            | Creates a login detail worksheet for the Excel report.                                |
    +-----------------------+---------------------------------------------------------------------------------------+

**Version Control**

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 4.0.0     | 04 Aug 2023   | Re-Launch                                                                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                       |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 4.0.2     | 03 Feb 2025   | Added link to fabric page.                                                        |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""

__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024, 2025 Consoli Solutions, LLC'
__date__ = '03 Feb 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.2'

import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.excel_util as excel_util
import brcddb.brcddb_common as brcddb_common
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.util.util as brcddb_util
import brcddb.brcddb_switch as brcddb_switch
import brcddb.report.utils as report_utils
import brcddb.app_data.report_tables as brcddb_rt

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')

_login_flags = dict(yes='\u221A', no='', unknown='?')

##################################################################
#
#        Case statements in login_page()
#
###################################################################


def _login_flag(login_obj, k):
    """Common flag method"""
    global _login_flags

    val = login_obj.r_get(k)
    if val is not None:
        try:
            return _login_flags[val]
        except KeyError:
            brcdapi_log.exception('Unknown login flag: ' + str(val) + ' for ' + str(k) + ' ' + login_obj.r_obj_key(),
                                  echo=True)
    return ''


def _l_switch_name_case(login_obj):
    return brcddb_switch.best_switch_name(login_obj.r_switch_obj(), False)


def _l_switch_name_and_wwn_case(login_obj):
    return brcddb_switch.best_switch_name(login_obj.r_switch_obj(), True)


def _l_switch_wwn_case(login_obj):
    switch_obj = login_obj.r_switch_obj()
    return '' if switch_obj is None else switch_obj.r_obj_key()


def _l_port_number_case(login_obj):
    port_obj = login_obj.r_port_obj()
    return '' if port_obj is None else port_obj.r_obj_key()


def _l_alias_case(login_obj):
    return '\n'.join(login_obj.r_fabric_obj().r_alias_for_wwn(login_obj.r_obj_key()))


def _l_zones_def_case(login_obj):
    # You can have an alias that includes a WWN + the same WWN as a WWN in the zone in which case this would return
    # the same zone twice. As a practical matter, I've never seen it done and nothing breaks in the code so this is
    # good enough
    return '\n'.join(login_obj.r_fabric_obj().r_zones_for_wwn(login_obj.r_obj_key()))


def _l_zones_eff_case(login_obj):
    return '\n'.join(login_obj.r_fabric_obj().r_eff_zones_for_wwn(login_obj.r_obj_key()))


def _l_fdmi_node_case(login_obj, k):
    # Remember, it's the base WWN we want for the node (hba), not the login WWN
    try:
        wwn = login_obj.r_port_obj().r_get(brcdapi_util.fc_neighbor_wwn)[0]
        buf = login_obj.r_fabric_obj().r_fdmi_node_obj(wwn).r_get(k)
        return '' if buf is None else buf
    except AttributeError:  # This happens when there is no FDMI information
        return ''


def _l_fdmi_port_case(login_obj, k):
    try:
        buf = login_obj.r_fabric_obj().r_fdmi_port_obj(login_obj.r_obj_key()).r_get(k)
    except AttributeError:
        buf = ''  # We get here when r_fdmi_port_obj() returns None meaning there is no FDMI port information

    return '' if buf is None else buf


def _l_port_name_case(login_obj):
    return login_obj.r_obj_key()


def _l_comment_case(login_obj):
    a_list = report_utils.combined_login_alert_objects(login_obj)
    return '\n'.join([obj.fmt_msg() for obj in a_list]) if len(a_list) > 0 else ''


def _l_share_area_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_share_area)


def _l_frame_redirection_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_redirection)


def _l_partial_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_partial)


def _l_lsan_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_lsan)


def _l_cascaded_ag_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_cascade_ag)


def _l_connected_through_ag_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_connect_ag)


def _l_real_device_behind_ag_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_dev_behind_ag)


def _l_fcoe_device_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_fcoe_dev)


def _l_slow_drain_device_quarantine_case(login_obj):
    return _login_flag(login_obj, brcdapi_util.bns_sddq)


_login_case = {
    '_FABRIC_NAME': report_utils.fabric_name_case,
    '_LOGIN_WWN': _l_port_name_case,
    '_FABRIC_NAME_AND_WWN': report_utils.fabric_name_or_wwn_case,
    '_FABRIC_WWN': report_utils.fabric_wwn_case,
    '_LOGIN_COMMENTS': _l_comment_case,
    '_PORT_NUMBER': _l_port_number_case,
    '_SWITCH_NAME': _l_switch_name_case,
    '_SWITCH_NAME_AND_WWN': _l_switch_name_and_wwn_case,
    '_SWITCH_WWN': _l_switch_wwn_case,
    '_ALIAS': _l_alias_case,
    '_ZONES_DEF': _l_zones_def_case,
    '_ZONES_EFF': _l_zones_eff_case,
    # 'brocade-name-server/fibrechannel-name-server/port-name': _l_port_name_case,
    brcdapi_util.bns_share_area: _l_share_area_case,
    brcdapi_util.bns_redirection: _l_frame_redirection_case,
    brcdapi_util.bns_partial: _l_partial_case,
    brcdapi_util.bns_lsan: _l_lsan_case,
    brcdapi_util.bns_cascade_ag: _l_cascaded_ag_case,
    brcdapi_util.bns_connect_ag: _l_connected_through_ag_case,
    brcdapi_util.bns_dev_behind_ag: _l_real_device_behind_ag_case,
    brcdapi_util.bns_fcoe_dev: _l_fcoe_device_case,
    brcdapi_util.bns_sddq: _l_slow_drain_device_quarantine_case,
}

_fdmi_case = dict(
    _FDMI_NODE=_l_fdmi_node_case,
    _FDMI_PORT=_l_fdmi_port_case,
)


def login_page(wb, tc, sheet_name, sheet_i, sheet_title, l_list, in_display=None, in_login_display_tbl=None, s=True):
    """Creates a login detail worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param l_list: List of login objects (LoginObj) to display
    :type l_list: list, tuple
    :param in_display: List of parameters to display. If None, use brcddb.app_data.report_tables.Login.login_tbl
    :type in_display: list, tuple, None
    :param in_login_display_tbl: Display control table. If None, use brcddb.report.report_tables.login_display_tbl
    :type in_login_display_tbl: dict, None
    :param s: If True, sorts the logins by port-id (port address where the login was found)
    :rtype: None
    """
    global _login_case, _fdmi_case, _align_wrap_vc, _align_wrap_c, _align_wrap, _border_thin, _bold_font, _std_font
    global _link_font, _hdr1_font

    # Validate the user input
    err_msg = list()
    if l_list is None:
        err_msg.append('l_list was not defined.')
    elif not isinstance(l_list, (list, tuple)):
        err_msg.append('l_list was type ' + str(type(l_list)) + '. Must be a list or tuple.')
    if len(err_msg) > 0:
        err_msg.append('Failed to create login_page().')
        brcdapi_log.exception(err_msg, echo=True)
        return
    login_display_tbl = brcddb_rt.Login.login_display_tbl if in_login_display_tbl is None else in_login_display_tbl
    display = brcddb_rt.Login.login_tbl if in_display is None else in_display

    # Create the worksheet with a title.
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)

    # Add a link to the fabric
    row, col = row + 1, 1
    fab_obj = l_list[0].r_fabric_obj() if len(l_list) > 0 else None
    if fab_obj is not None:
        excel_util.cell_update(
            sheet,
            row,
            col,
            'Fabric: ' + brcddb_fabric.best_fab_name(fab_obj),
            font=_link_font,
            border=_border_thin,
            link=fab_obj.r_get('report_app/hyperlink/fab')
        )
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # Add the headers, set the column widths, and freeze the header row.
    row += 1
    sheet.freeze_panes = sheet['A4']
    row, col = row + 1, 1
    for k in display:
        buf, align = k, _align_wrap
        if k in login_display_tbl and 'dc' in login_display_tbl[k] and login_display_tbl[k]['dc']:
            continue
        if k in login_display_tbl:
            if 'c' in login_display_tbl[k]:
                sheet.column_dimensions[xl.get_column_letter(col)].width = login_display_tbl[k]['c']
            if 'v' in login_display_tbl[k] and login_display_tbl[k]['v']:
                align = _align_wrap_vc
            buf = login_display_tbl[k]['d'] if 'd' in login_display_tbl[k] else k
        excel_util.cell_update(sheet, row, col, buf, font=_bold_font, align=align, border=_border_thin)
        col += 1

    # Add a row for each login
    row += 1
    if s:
        wl = gen_util.sort_obj_num(l_list, 'port-id', r=False, h=True)  # Filters out anything not found
        wl.extend(list(set(l_list) - set(wl)))
    else:
        wl = l_list

    port_obj_l, login_l = list(), list()
    for login_obj in [obj for obj in wl if obj is not None]:
        # The name server is on a per-switch basis so port_obj can be None if some switches weren't polled
        port_obj = login_obj.r_port_obj()
        if port_obj is None:
            login_l.append(login_obj)
        else:
            port_obj_l.append(port_obj)
    port_obj_l = brcddb_util.sort_ports(port_obj_l)
    for port_obj in port_obj_l:
        login_l.extend(port_obj.r_login_objects())

    for login_obj in login_l:
        if login_obj is None:
            continue
        col = 1

        # Report link
        brcddb_util.add_to_obj(login_obj,
                               'report_app/hyperlink/log',
                               '#' + sheet_name + '!' + xl.get_column_letter(col) + str(row))

        # Add it to the report
        font = report_utils.font_type(report_utils.combined_login_alert_objects(login_obj))
        for k in display:
            align = _align_wrap
            if k in login_display_tbl:
                if 'dc' in login_display_tbl[k] and login_display_tbl[k]['dc']:
                    continue
                if 'm' in login_display_tbl[k] and login_display_tbl[k]['m']:
                    align = _align_wrap_c
            k_list = k.split('.')
            if len(k_list) > 1:
                try:
                    buf = _fdmi_case[k_list[0]](login_obj, k_list[1])
                except BaseException as e:
                    buf = ''
                    e_buf = str(type(e)) + ': ' + str(e)
                    brcdapi_log.exception(['Unknown key: ' + k_list[0], e_buf], echo=True)
            elif k in _login_case:
                buf = _login_case[k](login_obj)
            elif k in login_display_tbl:
                v = login_obj.r_get(k)
                if v is None:
                    v1 = ''
                else:
                    try:
                        v1 = brcddb_common.login_conversion_tbl[k][v]
                    except KeyError:
                        v1 = v
                if isinstance(v1, bool):
                    buf = '\u221A' if v1 else ''
                else:
                    buf = v1
            else:
                font = _std_font
                buf = '' if login_obj.r_get(k) is None else str(login_obj.r_get(k))
            excel_util.cell_update(sheet, row, col, buf, font=font, align=align, border=_border_thin)
            col += 1
        row += 1
