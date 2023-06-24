# Copyright 2019, 2020, 2021, 2022, 2023 Jack Consoli.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
:mod:`report.zone` - Creates a zoning page to be added to an Excel Workbook

Public Methods & Data::

    +-----------------------+---------------------------------------------------------------------------------------+
    | Method                | Description                                                                           |
    +=======================+=======================================================================================+
    | zone_page             | Creates a zone detail worksheet for the Excel report                                  |
    +-----------------------+---------------------------------------------------------------------------------------+
    | alias_page            | Creates a port detail worksheet for the Excel report.                                 |
    +-----------------------+---------------------------------------------------------------------------------------+
    | target_zone_page      | Creates a target zone detail worksheet for the Excel report.                          |
    +-----------------------+---------------------------------------------------------------------------------------+
    | non_target_zone_page  | Creates a non-target zone detail worksheet for the Excel report.                      |
    +-----------------------+---------------------------------------------------------------------------------------+

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 1.x.x     | 03 Jul 2019   | Experimental                                                                      |
    | 2.x.x     |               |                                                                                   |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.0     | 19 Jul 2020   | Initial Launch                                                                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.1     | 02 Aug 2020   | PEP8 Clean up                                                                     |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.2     | 22 Aug 2020   | Fixed orientation of check marks.                                                 |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.3     | 29 Sep 2020   | Added alerts associated with the port for the zone member                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.4     | 26 Jan 2021   | Miscellaneous cleanup. No functional changes                                      |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.5     | 13 Feb 2021   | Removed the shebang line                                                          |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.6     | 13 Mar 2021   | Added login speed to the zone report.                                             |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.7     | 17 Jul 2021   | Added target_zone_page()                                                          |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.8     | 07 Aug 2021   | Minor display issues cleaned up.                                                  |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.9     | 14 Nov 2021   | Added non_target_zone_page() and defaults.                                        |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.0     | 31 Dec 2021   | Replaced bare except with explicit exceptions                                     |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.1     | 28 Apr 2022   | Added report links to zone and alias objects                                      |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.2     | 22 Jun 2022   | Increased column width for 'Member'                                               |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.3     | 04 Sep 2022   | Fixed missing zones on zone by target page.                                       |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.4     | 01 Jan 2023   | Fixed bad formatting of alert messages associated with logins.                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.5     | 11 Feb 2023   | Fixed missing alias and WWN info on zoning pages. Added common zones to zone by   |
    |           |               | target and zone by initiator pages.                                               |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.6     | 26 Mar 2023   | Added missing 'ha' member in _target_zone_hdr for "Target"                        |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.7     | 27 May 2023   | Cleaned up zone by target and group zone pages.                                   |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.8     | 29 May 2023   | Fixed wrong switch and port for login in zone group page.                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.1.9     | 04 Jun 2023   | Use URI references in brcdapi.util                                                |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.2.0     | 10 Jun 2023   | Added ungrouped ports to group zone page                                          |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.2.1     | 24 Jun 2023   | Fixed zone group logged into data and added number of logins column to zone_group |                      |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2019, 2020, 2021, 2022, 2023 Jack Consoli'
__date__ = '24 Jun 2023'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack.consoli@broadcom.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '3.2.1'

import collections
import openpyxl.utils.cell as xl
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcddb.brcddb_switch as brcddb_switch
import brcddb.brcddb_zone as brcddb_zone
import brcddb.report.fonts as report_fonts
import brcddb.brcddb_login as brcddb_login
import brcddb.app_data.alert_tables as al
import brcddb.util.search as brcddb_search
import brcddb.util.util as brcddb_util
import brcddb.report.utils as report_utils

UNGROUPED_TARGET = 'ungrouped_target_d'
UNGROUPED_INITIATOR = 'ungrouped_initiator_d'

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_italic_font = excel_fonts.font_type('italic')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_align_wrap_r = excel_fonts.align_type('wrap_right')
_border_thin = excel_fonts.border_type('thin')
_lightblue_fill = excel_fonts.fill_type('lightblue')

##################################################################
#
#              Generic case methods
#
##################################################################


def _comment_case(obj, p0=None, p1=None, p2=None):
    return '\n'.join([a.fmt_msg() for a in obj.r_alert_objects() if not a.is_flag()])


def _name_case(obj, p0=None, p1=None, p2=None):
    return obj.r_obj_key()


def _null_case(obj, p0=None, p1=None, p2=None):
    return ''


##################################################################
#
#              Case methods for hdr in zone_page()
#
##################################################################


def _zone_effective_case(obj):
    return '\u221A' if obj.r_is_effective() else ''


def _zone_peer_case(obj):
    return '\u221A' if obj.r_is_peer() else ''


def _zone_target_case(obj):
    return '\u221A' if obj.r_is_target_driven() else ''


def _zone_cfg_case(obj):
    return '\n'.join(obj.r_zone_configurations())


def _zone_member_case(obj):
    return len(obj.r_members()) + len(obj.r_pmembers())


def _zone_member_wwn_case(obj):
    members = list(obj.r_members())
    members.extend(list(obj.r_pmembers()))
    i, fab_obj = 0, obj.r_fabric_obj()
    for mem in members:
        i += 1 if fab_obj.r_alias_obj(mem) is None else len(fab_obj.r_alias_obj(mem).r_members())
    return i


def _mem_member_case(obj, mem, wwn, port_obj, obj_l=None):
    return mem


def _mem_comment_case(obj, mem, wwn, port_obj, obj_l=None):
    # Get any comments for this WWN in the zone object
    alert_l = [a.fmt_msg() for a in obj.r_alert_objects() if a.is_flag() and a.p0() == wwn]
    # If there is a login, add any comments associated with the login that are zone specific
    login_obj = obj.r_fabric_obj().r_login_obj(wwn)
    if obj.r_fabric_obj() is not None and login_obj is not None:
        alert_l.extend([a.fmt_msg() for a in login_obj.r_alert_objects() if a.is_flag() and a.p0() == mem])
    if port_obj is not None:  # Add the alerts associated with the port for this member
        alert_l.extend([a.fmt_msg() for a in port_obj.r_alert_objects() if a.is_error() or a.is_warn()])
    return '\n'.join(alert_l)


def _mem_principal_case(obj, mem, wwn, port_obj, obj_l=None):
    return '\u221A' if mem in obj.r_pmembers() else ''


def _mem_member_wwn_case(obj, mem, wwn, port_obj, obj_l=None):
    return '' if wwn is None else '' if ',' in wwn else wwn


def _mem_switch_case(obj, mem, wwn, port_obj, obj_l=None):
    return '' if port_obj is None else brcddb_switch.best_switch_name(port_obj.r_switch_obj())


def _mem_port_case(obj, mem, wwn, port_obj, obj_l=None):
    return '' if port_obj is None else port_obj.r_obj_key()


def _mem_speed_case(obj, mem, wwn, port_obj, obj_l=None):
    speed = None if port_obj is None else port_obj.r_get('fibrechannel/speed')
    return '' if speed is None else speed/1000000000


def _mem_desc_case(obj, mem, wwn, port_obj, obj_l=None):
    try:
        t_wwn = port_obj.r_get('fibrechannel/neighbor/wwn')[0] if ',' in wwn else wwn
    except (AttributeError, TypeError, IndexError):
        # If the corresponding zone member wasn't found, port_obj will be None which results in an AttributeError.
        # Nothing logged in retuns None (so a TypeError) with current FOS but if this ever gets fixed and returns an
        # empty list, nothing logged in will be an IndexError so adding IndexError is future proofing. I think the only
        # time this happens is with a d,i zone when nothing is logged in at the corresponding port
        return ''

    return brcddb_login.login_best_port_desc(obj.r_fabric_obj().r_login_obj(t_wwn))


def _alias_case(obj, mem, wwn, port_obj, obj_l=None):
    try:
        return obj.r_fabric_obj().r_alias_for_wwn(wwn)[0]
    except (AttributeError, IndexError):
        pass
    return ''


def _zone_list_case(obj, mem, wwn, port_obj, obj_l=None):
    return ', '.join(gen_util.convert_to_list(obj_l))


def _group_case(obj, mem, wwn, port_obj, obj_l=None):
    return '\u221A'


def _group_comment_case(obj, mem, wwn, port_obj, obj_l=None):
    l_port_obj = None if obj is None else obj.r_port_obj()
    return '' if l_port_obj is None else report_utils.combined_alerts(l_port_obj, wwn)


def _group_desc_case(obj, mem, wwn, port_obj, obj_l=None):
    return brcddb_login.login_best_node_desc(obj)


def _login_port_case(obj, mem, wwn, port_obj, obj_l=None):
    l_port_obj = None if obj is None else obj.r_port_obj()
    return '' if l_port_obj is None else l_port_obj.r_obj_key()


def _zoned_to_port_case(obj, mem, wwn, port_obj, obj_l=None):
    fab_obj = None if port_obj is None else port_obj.r_fabric_obj()
    return 0 if fab_obj is None else len(brcddb_zone.eff_zoned_to_wwn(fab_obj, wwn, all_types=True))


def _login_speed_case(obj, mem, wwn, port_obj, obj_l=None):
    l_port_obj = None if obj is None else obj.r_port_obj()
    return '' if l_port_obj is None else l_port_obj.r_get('_search/speed')


def _login_switch_case(obj, mem, wwn, port_obj, obj_l=None):
    return '' if obj is None else brcddb_switch.best_switch_name(obj.r_switch_obj())


"""
_zone_hdr & _zone_group_hdr_d: Key is the column header. Value is a dict as follows:
+-------+-----------+-----------------------------------------------------------------------------------------------+
| Key   | Type      | Description                                                                                   |
+=======+===========+===============================================================================================+
| c     | int       | Column width                                                                                  |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| z     | method    | Case method to fill in the cell for basic zone information                                    |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| m     | method    | Case method to fill the cell for member information                                           |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| v     | bool      | True - display centered in column. Headers vertical, Otherwise, use default wrap alignment    |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| g     | bool      | When True, fill in the cell for the group port.                                               |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| zg    | bool      | Same as 'g' except this is used for what is zoned to the group port.                          | 
+-------+-----------+-----------------------------------------------------------------------------------------------+
"""
_zone_hdr = collections.OrderedDict({
    'Comments': dict(c=30, z=_comment_case, m=_mem_comment_case),
    'Zone': dict(c=22, z=_name_case, m=_null_case),
    'Effective': dict(c=5, v=True, z=_zone_effective_case, m=_null_case),
    'Peer': dict(c=5, v=True, z=_zone_peer_case, m=_null_case),
    'Target Driven': dict(c=5, v=True, z=_zone_target_case, m=_null_case),
    'Principal': dict(c=5, v=True, z=_null_case, m=_mem_principal_case),
    'Configurations': dict(c=22, z=_zone_cfg_case, m=_null_case),
    'Member': dict(c=32, z=_zone_member_case, m=_mem_member_case),
    'Member WWN': dict(c=22, z=_zone_member_wwn_case, m=_mem_member_wwn_case),
    'Switch': dict(c=22, z=_null_case, m=_mem_switch_case),
    'Port': dict(c=7, z=_null_case, m=_mem_port_case),
    'Speed Gbps': dict(c=7, z=_null_case, m=_mem_speed_case),
    'Description': dict(c=50, z=_null_case, m=_mem_desc_case)
})
_zone_group_hdr_d = collections.OrderedDict({  # Same format as _zone_hdr
    'Group': dict(c=9, z=_null_case, m=_group_case, g=True, v=True),
    'Comments': dict(c=30, z=_null_case, m=_group_comment_case, g=True, zg=True),
    'Switch': dict(c=30, z=_null_case, m=_login_switch_case, g=True, zg=True),
    'Port': dict(c=8, z=_null_case, m=_login_port_case, g=True, zg=True),
    'Zoned To': dict(c=8, z=_null_case, m=_zoned_to_port_case, g=True, zg=False),
    'Speed Gbps': dict(c=8, z=_null_case, m=_login_speed_case, g=True, zg=True),
    'WWN': dict(c=23, z=_null_case, m=_mem_member_wwn_case, g=True, zg=True),
    'Alias': dict(c=34, z=_null_case, m=_alias_case, g=True, zg=True),
    'Description': dict(c=40, z=_null_case, m=_group_desc_case, g=True, zg=True),
    'Common Zone': dict(c=30, z=_null_case, m=_zone_list_case, zg=True),
})


def zone_page(fab_obj, tc, wb, sheet_name, sheet_i, sheet_title):
    """Creates a zone detail worksheet for the Excel report.

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param wb: Workbook object
    :type wb: class
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0.
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :rtype: None
    """
    global _zone_hdr, _link_font, _border_thin, _align_wrap, _hdr1_font, _std_font, _bold_font, _hdr2_font
    global _align_wrap_vc, _align_wrap_c, _lightblue_fill

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    row, col = row + 2, 1
    sheet.freeze_panes = sheet['A4']
    for k in _zone_hdr:
        sheet.column_dimensions[xl.get_column_letter(col)].width = _zone_hdr[k]['c']
        alignment = _align_wrap_vc if 'v' in _zone_hdr[k] and _zone_hdr[k]['v'] else _align_wrap
        excel_util.cell_update(sheet, row, col, k, font=_hdr2_font, align=alignment, border=_border_thin)
        col += 1

    # Fill out all the zoning information
    row, col = row + 1, 1
    zone_key_l = fab_obj.r_zone_keys()
    zone_key_l.sort()  # Just to make it easier to read the report
    for zone_obj in [fab_obj.r_zone_obj(k) for k in zone_key_l]:

        # Report link
        brcddb_util.add_to_obj(zone_obj,
                               'report_app/hyperlink/za',
                               '#' + sheet_name + '!' + xl.get_column_letter(col) + str(row))

        # The zone information
        for k in _zone_hdr:
            font = report_utils.font_type(zone_obj.r_alert_objects()) if k == 'Comments' else _bold_font
            alignment = _align_wrap_c if 'v' in _zone_hdr[k] and _zone_hdr[k]['v'] else _align_wrap
            excel_util.cell_update(sheet, row, col, _zone_hdr[k]['z'](zone_obj), font=font, align=alignment,
                                     border=_border_thin, fill=_lightblue_fill)
            col += 1

        # The member information
        row, col = row + 1, 1
        mem_list = list(zone_obj.r_pmembers())
        mem_list.extend(list(zone_obj.r_members()))
        for mem in mem_list:
            wwn_list = list()  # WWNs or d,i
            if fab_obj.r_alias_obj(mem) is None:
                wwn_list.append(mem)
            else:
                wwn_list.extend(fab_obj.r_alias_obj(mem).r_members())
            for wwn in wwn_list:
                if ',' in wwn:  # It's a d,i zone
                    port_obj = fab_obj.r_port_object_for_di(wwn)
                else:
                    port_obj = fab_obj.r_port_obj(wwn)
                if port_obj is None:
                    # brcddb_fabric.zone_analysis() already checked to see it's in another fabric. Although it would be
                    # much faster to check for the associated alert, if that zoning check were disabled, we would miss
                    # it. Time is not of the essence so spin through the loop once again.
                    for fobj in fab_obj.r_project_obj().r_fabric_objects():
                        if fab_obj.r_obj_key() != fobj.r_obj_key():
                            port_obj = fobj.r_port_obj(wwn)
                            if port_obj is not None:
                                break
                for k in _zone_hdr:
                    # Display font based on alerts associated with the zone only, not members, for a zone object
                    if k == 'Comments':
                        alerts = [a for a in zone_obj.r_alert_objects() if a.is_flag() and a.p0() == wwn]
                        font = report_utils.font_type(alerts)
                    else:
                        font = _std_font
                    alignment = _align_wrap_c if 'v' in _zone_hdr[k] and _zone_hdr[k]['v'] else _align_wrap
                    excel_util.cell_update(sheet, row, col, _zone_hdr[k]['m'](zone_obj, mem, wwn, port_obj),
                                             font=font, align=alignment, border=_border_thin)
                    col += 1
                row, col = row + 1, 1


def _alias_node_desc_case(obj, mem):
    try:
        return brcddb_login.login_best_node_desc(obj.r_fabric_obj().r_login_obj(mem))
    except AttributeError:
        return ''


def _alias_mem_member_case(obj, mem):
    return mem


def _alias_mem_switch_case(obj, mem):
    try:
        return brcddb_switch.best_switch_name(obj.r_fabric_obj().r_login_obj(mem).r_switch_obj(), False)
    except AttributeError:
        return ''


def _alias_mem_port_case(obj, mem):
    try:
        return obj.r_fabric_obj().r_login_obj(mem).r_port_obj().r_obj_key()
    except AttributeError:
        return ''


def _alias_mem_desc_case(obj, mem):
    try:
        return brcddb_login.login_best_port_desc(obj.r_fabric_obj().r_login_obj(mem))
    except AttributeError:
        return ''


def _alias_used_in_zone_case(obj, mem):
    return ', '.join([zobj.r_obj_key() for zobj in obj.r_zone_objects()])


def _alias_member_case(obj, mem):
    return obj.r_members()[0] if len(obj.r_members()) > 0 else ''


# Key is the column header. Value is a dict as follows:
#   'c' Column width
#   'v' Case method to fill in the cell for alias and first member
#   'm' Case method to fill the cell for remaining member information
_alias_hdr = collections.OrderedDict({
    'Comments': dict(c=40, v=_comment_case, m=_null_case),
    'Alias': dict(c=32, v=_name_case, m=_null_case),
    'Used in Zone': dict(c=32, v=_alias_used_in_zone_case, m=_null_case),
    'Member': dict(c=32, v=_alias_member_case, m=_alias_mem_member_case),
    'Switch': dict(c=22, v=_alias_mem_switch_case, m=_alias_mem_switch_case),
    'Port': dict(c=7, v=_alias_mem_port_case, m=_alias_mem_port_case),
    'Description': dict(c=50, v=_alias_node_desc_case, m=_alias_mem_desc_case)
})


def alias_page(fab_obj, tc, wb, sheet_name, sheet_i, sheet_title):
    """Creates a port detail worksheet for the Excel report.

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param wb: Workbook object
    :type wb: class
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0.
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :rtype: None
    """
    global _alias_hdr, _hdr1_font, _std_font, _link_font, _hdr2_font, _align_wrap, _border_thin

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font, border=_border_thin)
    row, col = row+2, 1
    sheet.freeze_panes = sheet['A4']
    for k in _alias_hdr:
        sheet.column_dimensions[xl.get_column_letter(col)].width = _alias_hdr[k]['c']
        excel_util.cell_update(sheet, row, col, k, font=_hdr2_font, align=_align_wrap, border=_border_thin)
        col += 1

    # Fill out the alias information
    for alias_obj in fab_obj.r_alias_objects():
        mem_list = alias_obj.r_members().copy()
        mem = mem_list.pop(0) if len(mem_list) > 0 else None
        row, col = row+1, 1

        # Report link
        brcddb_util.add_to_obj(alias_obj,
                               'report_app/hyperlink/ali',
                               '#' + sheet_name + '!' + xl.get_column_letter(col) + str(row))

        for k in _alias_hdr:
            font = report_utils.font_type(alias_obj.r_alert_objects()) if k == 'Comments' else _std_font
            excel_util.cell_update(sheet, row, col,  _alias_hdr[k]['v'](alias_obj, mem), font=font, align=_align_wrap,
                                     border=_border_thin)
            col += 1

        # Usually just one member per alias, but just in case...
        while len(mem_list) > 1:
            mem = mem_list.pop(0)
            row, col = row+1, 1
            for k in _alias_hdr:
                excel_util.cell_update(sheet, row, col, _alias_hdr[k]['m'](alias_obj, mem), font=_std_font,
                                         align=_align_wrap, border=_border_thin)
                col += 1


##################################################################
#
# Case methods for hdr in target_zone_page() & non_target_zone_page()
#
##################################################################
_t_skip = [_i for _i in range(al.ALERT_NUM.LOGIN_BASE, al.ALERT_NUM.ZONE_BASE)
           if _i != al.ALERT_NUM.LOGIN_MIXED_SPEED_T and _i != al.ALERT_NUM.LOGIN_FASTER_S]


def _filter_alerts(a_obj_l):
    rl = list()
    for obj in a_obj_l:
        if obj is not None:
            rl.extend([a for a in obj.r_alert_objects() if a.alert_num() not in _t_skip and not a.is_flag()])
    return rl


def _null_target_case(wwn, obj, obj2, zone_l, zone_d):
    """Returns an empty string

    :param wwn: WWN of login in progress
    :type wwn: str
    :param obj:
    :type obj: None, AliasObj, ZoneObj, ZoneCfgObj
    :param obj2:
    :type obj2: None, AliasObj, ZoneObj, ZoneCfgObj
    :param zone_l: List of zones byt name
    :type zone_l: list
    :param zone_d: Dictionary returned from _get_zoned_to()
    :type zone_d: dict
    """
    return ''


def _tzone_comment_case(wwn, obj, obj2, zone_l, zone_d):
    return '\n'.join([a_obj.fmt_msg() for a_obj in _filter_alerts([obj, obj2])])


def _tzone_name_case(wwn, obj, obj2, zone_l, zone_d):
    return wwn if obj is None else brcddb_login.best_login_name(obj.r_fabric_obj(), obj.r_obj_key(), True)


def _tzone_zone_case(wwn, obj, obj2, zone_l, zone_d):
    return ', '.join(zone_l)


def _tzone_switch_case(wwn, obj, obj2, zone_l, zone_d):
    return wwn if obj is None else brcddb_switch.best_switch_name(obj.r_switch_obj())


def _tzone_port_case(wwn, obj, obj2, zone_l, zone_d):
    return '' if obj2 is None else obj2.r_obj_key()


def _tzone_speed_case(wwn, obj, obj2, zone_l, zone_d):
    port_obj = None if obj is None else obj.r_port_obj()
    speed = None if port_obj is None else port_obj.r_get('fibrechannel/speed')
    return '' if speed is None else speed/1000000000


def _tzone_desc_case(wwn, obj, obj2, zone_l, zone_d):
    return brcddb_login.login_best_node_desc(obj)


def _tzone_t_count(wwn, obj, obj2, zone_l, zone_d):
    return len(zone_d['target'])


def _tzone_s_count(wwn, obj, obj2, zone_l, zone_d):
    return len(zone_d['non_target'])


# Key is the column header. Value is a dict as follows:
#   'c' Column width
#   'ha'    Case method to call when filling in the cell for the sub-header
#   'a'     Case method to call when filling in the cell for the item associated with the sub-header
_target_zone_hdr = collections.OrderedDict({
    'Comments': dict(c=30, ha=_tzone_comment_case, a=_tzone_comment_case),
    'Target': dict(c=30, ha=_tzone_name_case, a=_null_target_case),
    'Non-Target': dict(c=30, ha=_tzone_s_count, a=_tzone_name_case),
    'Zoned Target': dict(c=30, ha=_tzone_t_count, a=_tzone_name_case),
    'Common Zone': dict(c=30, ha=_tzone_zone_case, a=_tzone_zone_case),
    'Switch': dict(c=22, ha=_tzone_switch_case, a=_tzone_switch_case),
    'Port': dict(c=22, ha=_tzone_port_case, a=_tzone_port_case),
    'Speed Gbps': dict(c=7, ha=_tzone_speed_case, a=_tzone_speed_case),
    'Description': dict(c=50, ha=_tzone_desc_case, a=_tzone_desc_case)
})

_server_zone_hdr = collections.OrderedDict({
    'Comments': dict(c=30, ha=_tzone_comment_case, a=_tzone_comment_case),
    'Initiator': dict(c=30, ha=_tzone_name_case, a=_null_target_case),
    'Target': dict(c=30, ha=_tzone_t_count, a=_tzone_name_case),
    'Zoned Server': dict(c=30, ha=_tzone_s_count, a=_tzone_name_case),
    'Common Zone': dict(c=30, ha=_tzone_zone_case, a=_tzone_zone_case),
    'Switch': dict(c=22, ha=_tzone_switch_case, a=_tzone_switch_case),
    'Port': dict(c=22, ha=_tzone_port_case, a=_tzone_port_case),
    'Speed Gbps': dict(c=7, ha=_tzone_speed_case, a=_tzone_speed_case),
    'Description': dict(c=50, ha=_tzone_desc_case, a=_tzone_desc_case)
})


def _common_zone_page(fab_obj, tc, wb, sheet_name, sheet_i, sheet_title, hdr):
    """Creates a target zone detail worksheet for the Excel report.

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param wb: Workbook object
    :type wb: class
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0.
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param hdr: Either _target_zone_hdr or _server_zone_hdr
    :return sheet: Sheet (page) added to the workbook (wb)
    :rtype sheet: class
    :return row: Next available row on the worksheet
    :rtype row: int
    """
    global _hdr2_font, _hdr1_font, _link_font, _align_wrap, _border_thin

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, align=_align_wrap, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    row, col = row+2, 1
    sheet.freeze_panes = sheet['A4']
    for k, d in hdr.items():
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['c']
        excel_util.cell_update(sheet, row, col, k, font=_hdr2_font, align=_align_wrap, border=_border_thin)
        col += 1
    row += 1

    return sheet, row


def _get_zoned_to(fab_obj, wwn):
    """Get the targets and all else in the effective zone configuration zoned to a WWN.

    Return dictionary:

    +---------------+-------------------------------------------+
    | Key           | Value                                     |
    +===============+===========================================|
    | target        | list of target WWNs zoned to this wwn     |
    +---------------+-------------------------------------------+
    | non_target    | list of non-target WWNs zoned to this wwn |
    +---------------+-------------------------------------------+

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param wwn: WWN to find what's zoned to it
    :type wwn: str, None
    :return: Dictionary as defined in the method header
    :rtype: dict
    """
    rd = dict(target=brcddb_zone.eff_zoned_to_wwn(fab_obj, wwn, target=True), non_target=dict())
    if not gen_util.is_wwn(wwn):
        return rd
    for k, zone_l in brcddb_zone.eff_zoned_to_wwn(fab_obj, wwn, all_types=True).items():
        if k not in rd['target']:
            rd['non_target'].update({k: zone_l})
    return rd


def target_zone_page(fab_obj, tc, wb, sheet_name, sheet_i, sheet_title):
    """Creates a target zone detail worksheet for the Excel report.

    Note: Adding non_target_zone_page() was an after thought. I created _common_zone_page() to set up the worksheet but
    I certainly could have written target_zone_page() and non_target_zone_page() to share more code.

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param wb: Workbook object
    :type wb: class
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0.
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :rtype: None
    """
    global _target_zone_hdr, _std_font, _bold_font, _align_wrap, _border_thin

    sheet, row = _common_zone_page(fab_obj, tc, wb, sheet_name, 0 if sheet_i is None else sheet_i, sheet_title,
                                   _target_zone_hdr)

    # Determine all the targets and what is zoned to each target
    t_obj_l = brcddb_search.match(fab_obj.r_login_objects(),  # List of login objects that are targets
                                  brcdapi_util.bns_fc4_features,
                                  'Target',
                                  ignore_case=True,
                                  stype='regex-s')  # List of target objects in the fabric
    for t_login_obj in t_obj_l:
        wwn = t_login_obj.r_obj_key()
        zoned_to_d = _get_zoned_to(fab_obj, wwn)
        if len(zoned_to_d['target']) + len(zoned_to_d['non_target']) == 0:
            continue  # This is a single member zone if we get here

        # The sub-header for the target
        row, col = row+1, 1
        for k, d in _target_zone_hdr.items():
            font = report_utils.font_type(_filter_alerts([t_login_obj, t_login_obj.r_port_obj()])) if k == 'Comments' \
                else _bold_font
            buf = d['ha'](wwn, t_login_obj, t_login_obj.r_port_obj(), list(), zoned_to_d)
            excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
            col += 1

        # Everything zoned to this target that is not a target (in most cases, these are servers)
        for wwn, zone_l in zoned_to_d['non_target'].items():
            login_obj = fab_obj.r_login_obj(wwn)
            port_obj = None if login_obj is None else login_obj.r_port_obj()
            row, col = row+1, 1
            for k in _target_zone_hdr:
                font = report_utils.font_type(_filter_alerts([login_obj, port_obj])) if k == 'Comments' \
                    else _std_font
                buf = '' if k == 'Zoned Target' else \
                    _target_zone_hdr[k]['a'](wwn, login_obj, port_obj, zone_l, zoned_to_d)
                excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
                col += 1

        # Add any targets zoned to this target
        for wwn, zone_l in zoned_to_d['target'].items():
            login_obj = fab_obj.r_login_obj(wwn)
            port_obj = None if login_obj is None else login_obj.r_port_obj()
            row, col = row+1, 1
            for k in _target_zone_hdr:
                font = report_utils.font_type(_filter_alerts([login_obj, port_obj])) if k == 'Comments' \
                    else _std_font
                buf = '' if k == 'Non-Target' else \
                    _target_zone_hdr[k]['a'](wwn, login_obj, port_obj, zone_l, zoned_to_d)
                excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
                col += 1

        row += 1


def non_target_zone_page(fab_obj, tc, wb, sheet_name, sheet_i, sheet_title):
    """Creates a non-target zone detail worksheet for the Excel report.

    See comments with target_zone_page() for additional notes and input parameter definitions.
    """
    global _server_zone_hdr, _std_font, _bold_font, _align_wrap, _border_thin

    sheet, row = _common_zone_page(fab_obj, tc, wb, sheet_name, 0 if sheet_i is None else sheet_i, sheet_title,
                                   _server_zone_hdr)

    # Fill out all the zone information for each server
    s_obj_l = brcddb_search.match(fab_obj.r_login_objects(),
                                  brcdapi_util.bns_fc4_features,
                                  'Initiator',
                                  ignore_case=True,
                                  stype='regex-s')  # List of initiator objects in the fabric
    for s_login_obj in s_obj_l:
        wwn = s_login_obj.r_obj_key()
        zoned_to_d = _get_zoned_to(fab_obj, wwn)
        if len(zoned_to_d['target']) + len(zoned_to_d['non_target']) == 0:
            continue  # This is a single member zone if we get here

        # The server (initiator) sub-header
        row, col = row+1, 1
        for k, d in _server_zone_hdr.items():
            font = report_utils.font_type(_filter_alerts([s_login_obj, s_login_obj.r_port_obj()])) if k == 'Comments' \
                else _bold_font
            buf = d['ha'](wwn, s_login_obj, s_login_obj.r_port_obj(), list(), zoned_to_d)
            excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
            col += 1

        # Fill in all targets zoned to this initiator
        for wwn, zone_l in zoned_to_d['target'].items():
            login_obj = fab_obj.r_login_obj(wwn)
            port_obj = None if login_obj is None else login_obj.r_port_obj()
            row, col = row+1, 1
            for k in _server_zone_hdr:
                font = report_utils.font_type(_filter_alerts([login_obj, port_obj])) if k == 'Comments' \
                    else _std_font
                buf = '' if k == 'Zoned Server' else \
                    _server_zone_hdr[k]['a'](wwn, login_obj, port_obj, zone_l, zoned_to_d)
                excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
                col += 1

        # Add any initiators zoned to this initiator
        for wwn, zone_l in zoned_to_d['non_target'].items():
            login_obj = fab_obj.r_login_obj(wwn)
            port_obj = None if login_obj is None else login_obj.r_port_obj()
            row, col = row+1, 1
            for k in _server_zone_hdr:
                font = report_utils.font_type(_filter_alerts([login_obj, port_obj])) if k == 'Comments' \
                    else _std_font
                buf = '' if k == 'Target' else \
                    _server_zone_hdr[k]['a'](wwn, login_obj, port_obj, zone_l, zoned_to_d)
                excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin)
                col += 1

        row += 1


def group_zone_page(proj_obj, tc, wb, sheet_name, sheet_i, sheet_title):
    """Creates a zone group detail worksheet for the Excel report.

    See comments with target_zone_page() for additional notes and input parameter definitions.
    """
    global _zone_group_hdr_d, _std_font, _bold_font, _italic_font, _align_wrap, _border_thin, _lightblue_fill
    global _align_wrap_c, _align_wrap_r, UNGROUPED_TARGET, UNGROUPED_INITIATOR

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, align=_align_wrap, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(_zone_group_hdr_d))
    row, col = row+1, 2
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_zone_group_hdr_d))
    excel_util.cell_update(sheet,
                           row,
                           col,
                           'Port in group is highlighted and followed by all that is online and zoned to that port',
                           font=_std_font,
                           align=_align_wrap,
                           border=_border_thin)
    row, col = row+1, 1
    sheet.freeze_panes = sheet['A4']
    for k, d in _zone_group_hdr_d.items():
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['c']
        excel_util.cell_update(sheet, row, col, k, font=_hdr2_font, align=_align_wrap, border=_border_thin)
        col += 1

    # This list is to ensure the ungrouped targets and initiators come after all the groups
    zone_group_d = proj_obj.r_get('report_app/group_d')
    if not isinstance(zone_group_d, dict):
        return
    last_group_d = {UNGROUPED_TARGET: True, UNGROUPED_INITIATOR: True}
    zone_group_l = [k for k in zone_group_d.keys() if not bool(last_group_d.get(k))]
    zone_group_l.extend(last_group_d.keys())

    # Add each group
    for group_name in zone_group_l:
        group_d = zone_group_d[group_name]
        group_obj_l = brcddb_util.sort_ports(group_d['port_obj_l'])  # List of port objects in the zone group
        row, col = row+2, 1

        # The group name
        excel_util.cell_update(sheet, row, col, group_name, fill=_lightblue_fill, font=_bold_font, border=_border_thin)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_zone_group_hdr_d))

        # A common use for groups is to define storage enclosures. It's also common to put all the storage in the same
        # zone. This means the same WWNs can appear in multiple places. zone_d is used to ensure each group member is
        # displayed once.
        zone_d = dict()
        for port_obj in group_obj_l:
            for login_obj in port_obj.r_login_objects():
                zone_d.update({login_obj.r_obj_key(): True})

        # Each port in the zone group
        for port_obj in group_obj_l:
            # Information about the logins for this port.
            row, col, fab_obj = row+1, 1, port_obj.r_fabric_obj()
            if fab_obj is None:
                # I can't think of a way for fab_obj to be None here so this is just in case I'm overlooking something
                continue

            # The logins for this port - typically just one
            for login_obj in port_obj.r_login_objects():
                wwn = login_obj.r_obj_key()
                for d in _zone_group_hdr_d.values():
                    excel_util.cell_update(sheet,
                                           row,
                                           col,
                                           d['m'](login_obj, None, wwn, port_obj) if bool(d.get('g')) else '',
                                           font=_bold_font,
                                           border=_border_thin,
                                           align=_align_wrap_c if bool(d.get('v')) else _align_wrap)
                    col += 1

            # The logins zoned to this port
            if bool(last_group_d.get(group_name)):
                continue
            for lwwn, zl in brcddb_zone.eff_zoned_to_wwn(fab_obj, wwn, all_types=True).items():
                lz_obj = fab_obj.r_login_obj(lwwn)
                if lz_obj is None:
                    continue  # The coresponding device may not be logged in

                row, col = row+1, 1
                for d in _zone_group_hdr_d.values():
                    buf = d['m'](lz_obj, None, lwwn, port_obj, zl) if bool(d.get('zg')) else ''
                    excel_util.cell_update(sheet,
                                           row,
                                           col,
                                           buf,
                                           font=_std_font,
                                           border=_border_thin,
                                           align=_align_wrap_c if bool(d.get('v')) else _align_wrap)
                    col += 1

        row += 1
