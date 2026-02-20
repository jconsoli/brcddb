"""
Copyright 2023, 2024, 2025, 2026 Jack Consoli.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack_consoli@yahoo.com for
details.

** Description**

Creates a worksheet with a graph

**Public Methods & Data**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| chart_types           | Returns the supported chart types.                                                        |
+-----------------------+-------------------------------------------------------------------------------------------+
| graph                 | Inserts a worksheet into a workbook with a graph of specified ports and statistics.       |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 25 Aug 2025   | Updated email address in __email__ only.                                              |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 19 Oct 2025   | Updated comments only.                                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 20 Feb 2026   | Added supported_chart_types(). Set the chart size in graph()                          |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025, 2026 Jack Consoli'
__date__ = '20 Feb 2026'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.4'

from openpyxl.chart import AreaChart, AreaChart3D, BarChart, BarChart3D, LineChart, LineChart3D, Reference
from openpyxl.chart.axis import DateAxis
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.log as brcdapi_log

_std_font = excel_fonts.font_type('std')
_link_font = excel_fonts.font_type('link')
_align_wrap = excel_fonts.align_type('wrap')

_chart_width = 22
_chart_height = 15


def _area_chart():
    return AreaChart()


def _area_3d_chart():
    return AreaChart3D()


def _bar_chart():
    return BarChart()


def _bar_3d_chart():
    return BarChart3D()


def _line_chart():
    return LineChart()


def _line_3d_chart():
    return LineChart3D()


_chart_types = dict(
    area=_area_chart,
    area_3d=_area_3d_chart,
    bar=_bar_chart,
    bar_3d=_bar_3d_chart,
    line=_line_chart,
    line_3d=_line_3d_chart,
)

def supported_chart_types():
    """Returns a list of strings representing the supported chart types.

    :return: Supported chart types
    :rtype: list
    """

    return [str(k) for k in _chart_types.keys()]


def graph(wb, tc, sheet_name, sheet_i, data_ref_d, msg=None, title_in_data=False):
    """Inserts a worksheet into a workbook with a graph of specified ports and statistics.

    data_ref_d: The graph assumes that the first row of the data are the data labels

    +---------------+-----------+-----------------------------------------------------------------------------------+
    | Member        | type      | Description                                                                       |
    +===============+===========+===================================================================================+
    | title         | str       | Chart title                                                                       |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | type          | str       | Chart type. Must be one of the keys in _chart_types.                              |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | x_title       | str       | X Axis title.                                                                     |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | x_sheet       | pointer   | Pointer to the reference worksheet, from openpyxl.Workbook().create_sheet(), for  |
    |               |           | the X axis (column data). This is the sheet referenced by x_start_col,            |
    |               |           | x_end_col, x_start_row, and x_end_row. The values in these cells must be either a |
    |               |           | time stamp in hh:mm:ss      |
    |               |           | format as values in these cells.                           |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | x_start_col   | int       | Column number where the x data begins. See Note below.                            |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | x_start_row   | int       | Row number where the x data begins. See Note below.                               |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | x_end_row     | int       | Row number where the x data end. See Note below.                                  |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | y_title       | str       | Y Axis title.                                                                     |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | y_data_l      | list      | List of dictionaries as follows:                                                  |
    |               |           |                                                                                   |
    |               |           | sheet       Pointer to the reference worksheet, from                              |
    |               |           |             openpyxl.Workbook().create_sheet().                                   |
    |               |           |                                                                                   |
    |               |           | start_col     int Starting column for the Y axis. See Note below.                 |
    |               |           |                                                                                   |
    |               |           | end_col       int Column number where the x data ends. See Note below.            |
    |               |           |                                                                                   |
    |               |           | start_row     int Row number where the Y data begins. See Note below.             |
    |               |           |                                                                                   |
    |               |           | end_row       int ow number where the x data end. See Note below.                 |
    +---------------+-----------+-----------------------------------------------------------------------------------+

    **Note** Keep in mind that cell references in Excel begin with 1, not 0. The row number should include the row
    header, whether title_in_data is True or False. The X axis can only have one reference, so there isn't an x_end_col.

    :param wb: Workbook
    :type wb: Workbook object
    :param tc: Link to table of context page.
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed.
    :type sheet_i: int
    :param data_ref_d: Chart type. See chart_types
    :type data_ref_d: dict
    :param msg: Optional message to add to top of chart. Not the title. The chart title is in data_ref_d.
    :type msg: str, None
    :param title_in_data: If True: Title is included in the data (row starts with the header)
    :type title_in_data: bool
    :rtype: None
    """
    global _std_font, _link_font, _chart_types, _chart_width, _chart_height

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    col = 10
    sheet.column_dimensions['A'].width = col
    sheet.column_dimensions['B'].width = 120 - col
    if isinstance(tc, str):
        excel_util.cell_update(sheet, 1, 1, 'Contents', link=tc, font=_link_font, align=_align_wrap)
    if isinstance(msg, str):
        excel_util.cell_update(sheet, 1, 2, msg, font=_std_font, align=_align_wrap)

    # Setup the chart
    chart = _chart_types.get(data_ref_d['type'], 'line')()
    chart.title = data_ref_d['title']
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.title = data_ref_d['x_title']
    chart.y_axis.title = data_ref_d['y_title']
    chart.y_axis.crossAx = 500
    chart.width = _chart_width
    chart.height = _chart_height

    # Add the Y axis (statistics) data
    for y_data_d in data_ref_d['y_data_l']:
        y_data_ref = Reference(
            y_data_d['sheet'],
            min_col=y_data_d['start_col'],
            min_row=y_data_d['start_row'],
            max_col=y_data_d['end_col'],
            max_row=y_data_d['end_row'],
        )
        chart.add_data(y_data_ref, titles_from_data=title_in_data)

    # Add the X axis (date) reference
    chart.set_categories(
        Reference(
            data_ref_d['x_sheet'],
            min_col=data_ref_d['x_start_col'],
            min_row=data_ref_d['x_start_row'],
            max_row=data_ref_d['x_end_row']
        )
    )

    sheet.add_chart(chart, 'A3')
