# Copyright 2019, 2020, 2021, 2022 Jack Consoli.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""

:mod:`report.port` - Includes methods to create port page and performance dsahboard Excel worksheets

Public Methods & Data::

    +-----------------------+---------------------------------------------------------------------------------------+
    | Method                | Description                                                                           |
    +=======================+=======================================================================================+
    | performance_dashboard | Creates a dashboard worksheet for the Excel report.                                   |
    +-----------------------+---------------------------------------------------------------------------------------+
    | port_page             | Creates a port detail worksheet for the Excel report.                                 |
    +-----------------------+---------------------------------------------------------------------------------------+

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 1.x.x     | 03 Jul 2019   | Experimental                                                                      |
    | 2.x.x     |               |                                                                                   |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.0     | 19 Jul 2020   | Initial Launch                                                                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.1     | 02 Aug 2020   | PEP8 Clean up                                                                     |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.2     | 29 Sep 2020   | Standardized time stamp and added return type documentation to port_page()        |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.3     | 26 Jan 2021   | Miscellaneous cleanup. No functional changes                                      |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.4     | 13 Feb 2021   | Removed the shebang line                                                          |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.5     | 13 Feb 2021   | Changed is to == for literal compare                                              |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.6     | 14 Nov 2021   | No functional changes. Added defaults for display tables and sheet indices.       |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.7     | 31 Dec 2021   | Made all exception clauses explicit                                               |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.8     | 28 Apr 2022   | Added hyperlinks for ports                                                        |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 3.0.9     | 25 Jul 2022   | Fixe case where link may not have been assigned in port_page()                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2019, 2020, 2021, 2022 Jack Consoli'
__date__ = '25 Jul 2022'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack.consoli@broadcom.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '3.0.9'

import datetime
import collections
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcddb.brcddb_common as brcddb_common
import brcddb.util.util as brcddb_util
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_switch as brcddb_switch
import brcddb.report.utils as report_utils
import brcddb.app_data.report_tables as rt

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')


def performance_dashboard(wb, tc, sheet_name, sheet_i, sheet_title, content):
    """Creates a dashboard worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param content: Caller defined content. List of lists or tuples to add to the title page. For example:
            dashboard = collections.OrderedDict()
            dashboard['class-3-discards'] =     {'title' : 'Top 10 C3 Discards', 'port_list' : []}
            dashboard['in-crc-errors'] =        {'title' : 'Top 10 CRC With Good EOF', 'port_list' : []}
    :type content: list, tuple
    :rtype: None
    """
    global _border_thin, _std_font, _align_wrap, _link_font, _bold_font, _hdr2_font

    hdr = collections.OrderedDict()
    # Key is the column header and value is the width
    hdr['Count'] = 14
    hdr['Fabric'] = 22
    hdr['Switch'] = 22
    hdr['Port'] = 7
    hdr['Type'] = 8
    hdr['Description'] = 117 - (hdr['Count'] + hdr['Switch'] + hdr['Switch'] + hdr['Port'] + hdr['Type'])

    # Create the worksheet, add the title, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    for k, v in hdr.items():
        sheet.column_dimensions[xl.get_column_letter(col)].width = v
        col += 1
    max_col = col - 1
    col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    row += 2

    # Now add each dashboard item
    for statistic in content.keys():
        db = content.get(statistic)
        col = 1

        # The individual dashboard title
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        for i in range(col, max_col+1):
            excel_util.cell_update(sheet, row, i, None, border=_border_thin)
        excel_util.cell_update(sheet, row, col, db.get('title'), font=_hdr2_font,
                                 fill=excel_fonts.fill_type('lightblue'))
        row, col = row + 1, 1

        # Now the individual dashboard headers
        for k in hdr.keys():
            excel_util.cell_update(sheet, row, col, k, font=_bold_font, border=_border_thin)
            col += 1

        # Now add the dashboard content
        port_list = db.get('port_list')
        if len(port_list) == 0:
            row += 1
        for port_obj in port_list:
            col = 1
            temp_l = [port_obj.r_get('fibrechannel-statistics/' + statistic),
                      brcddb_fabric.best_fab_name(port_obj.r_switch_obj().r_fabric_obj()),
                      brcddb_switch.best_switch_name(port_obj.r_switch_obj())]
            for buf in temp_l:
                excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap, border=_border_thin)
                col += 1
            # Port
            link = port_obj.r_get('report_app/hyperlink_ps')
            if link is not None:
                excel_util.cell_update(sheet, row, col, port_obj.r_obj_key(), font=_link_font, align=_align_wrap,
                                         link=link, border=_border_thin)
            else:
                excel_util.cell_update(sheet, row, col, port_obj.r_obj_key(), font=_std_font, align=_align_wrap,
                                         border=_border_thin)
            col += 1
            for buf in [port_obj.c_login_type(), brcddb_port.port_best_desc(port_obj)]:
                excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap, border=_border_thin)
                col += 1
            row += 1
        row += 1

##################################################################
#
#        Port case statements in port_page()
#
###################################################################


def _p_switch_name_case(port_obj, k, wwn):
    return brcddb_switch.best_switch_name(port_obj.r_switch_obj(), False)


def _p_switch_name_and_wwn_case(port_obj, k, wwn):
    return brcddb_switch.best_switch_name(port_obj.r_switch_obj(), True)


def _p_switch_wwn_case(port_obj, k, wwn):
    return port_obj.r_switch_key()


def _p_port_number_case(port_obj, k, wwn):
    return '\'' + port_obj.r_obj_key()  # Add ' in front of it to make sure Excel doesn't treat it like a date


def _p_port_desc_case(port_obj, k, wwn):
    return brcddb_port.port_best_desc(port_obj)


def _p_port_config_link_case(port_obj, k, wwn):
    return 'Config', port_obj.r_get('report_app/hyperlink/pc')


def _p_port_stats_link_case(port_obj, k, wwn):
    return 'Stats', port_obj.r_get('report_app/hyperlink/ps')


def _p_port_zone_link_case(port_obj, k, wwn):
    return 'Zone', port_obj.r_get('report_app/hyperlink/pz')


def _p_port_sfp_link_case(port_obj, k, wwn):
    return 'SFP', port_obj.r_get('report_app/hyperlink/sfp')


def _p_port_rnid_link_case(port_obj, k, wwn):
    return 'RNID', port_obj.r_get('report_app/hyperlink/pr')


def _p_port_maps_group_case(port_obj, k, wwn):
    maps_groups = port_obj.r_get('_maps_fc_port_group')
    maps_groups.extend(port_obj.r_get('_maps_sfp_group'))
    return ', '.join(maps_groups)


def _p_time_generated_case(port_obj, k, wwn):
    x = port_obj.r_get('fibrechannel-statistics/time-generated')
    return None if x is None else datetime.datetime.fromtimestamp(x).strftime('%d %b %Y, %H:%M:%S')


def _p_los_tov_mode_case(port_obj, k, wwn):
    los = port_obj.r_get('fibrechannel/los-tov-mode-enabled')
    if los is None:
        return ''
    v = brcddb_common.port_conversion_tbl['fibrechannel/los-tov-mode-enabled'][los]
    return 'Unknown: ' + str(los) if v is None else v


def _p_alias_case(port_obj, k, wwn):
    return '\n'.join(port_obj.r_fabric_obj().r_alias_for_wwn(wwn))


def _p_comment_case(port_obj, k, wwn):
    if k == '_PORT_COMMENTS':  # It's for the port itself
        return report_utils.combined_alerts(port_obj, wwn)
    else:  # It's an NPIV login
        return report_utils.combined_login_alerts(port_obj.r_fabric_obj().r_login_obj(wwn), wwn)


def _p_login_wwn_case(port_obj, k, wwn):
    return wwn


def _p_login_addr_case(port_obj, k, wwn):
    if wwn is None:
        return ''
    try:
        return port_obj.r_fabric_obj().r_login_obj(wwn).r_get('brocade-name-server/port-id')
    except AttributeError:
        brcdapi_log.exception('No login address for ' + wwn + '. Switch ' + port_obj.r_switch_obj().r_obj_key() +
                              ', port ' + port_obj.r_obj_key(), True)
    return 'Unknown'


def _p_zones_def_case(port_obj, k, wwn):
    # You can have an alias that includes a WWN + the same WWN as a WWN in the zone in which case this would return
    # the same zone twice. As a practical matter, I've never seen it done and nothing breaks in the code so this is
    # good enough
    fab_obj = port_obj.r_fabric_obj()
    return '\n'.join(fab_obj.r_zones_for_wwn(wwn) +
                     fab_obj.r_zones_for_di(port_obj.r_switch_obj().r_did(), port_obj.r_index()))


def _zones_eff_case(port_obj, k, wwn):
    fab_obj = port_obj.r_fabric_obj()
    return '\n'.join(fab_obj.r_eff_zones_for_wwn(wwn) + fab_obj.r_eff_di_zones_for_addr(port_obj.r_addr()))


def _p_name_server_node_case(port_obj, k, wwn):
    try:
        return port_obj.r_fabric_obj().r_login_obj(wwn).r_get('brocade-name-server/node-symbolic-name')
    except AttributeError:
        return None


def _p_name_server_port_case(port_obj, k, wwn):
    try:
        return port_obj.r_fabric_obj().r_login_obj(wwn).r_get('brocade-name-server/port-symbolic-name')
    except AttributeError:
        return None


def _p_fdmi_node_case(port_obj, k, wwn):
    try:
        buf = port_obj.r_fabric_obj().r_fdmi_node_obj(wwn).r_get('brocade-fdmi/node-symbolic-name')
    except AttributeError:
        buf = ''
    return '' if buf is None else buf


def _p_fdmi_port_case(port_obj, k, wwn):
    try:
        buf = port_obj.r_fabric_obj().r_fdmi_port_obj(wwn).r_get('brocade-fdmi/port-symbolic-name')
    except AttributeError:
        buf = ''
    return '' if buf is None else buf


def _p_media_uptime_case(port_obj, k, wwn):
    try:
        return int(port_obj.r_get('media-rdp/power-on-time')/24 + .5)
    except (TypeError, ValueError):
        return None


def _p_media_distance_case(port_obj, k, wwn):
    try:
        return ', '.join(brcddb_util.get_key_val(port_obj, 'media-rdp/media-distance/distance'))
    except TypeError:
        return None


def _p_media_speed_case(port_obj, k, wwn):
    if port_obj.r_get('media-rdp/media-speed-capability/speed') is None:
        return None
    else:
        return ', '.join([str(i) for i in port_obj.r_get('media-rdp/media-speed-capability/speed')])


def _p_media_rspeed_case(port_obj, k, wwn):
    if port_obj.r_get('media-rdp/remote-media-speed-capability/speed') is None:
        return None
    else:
        return ', '.join([str(i) for i in port_obj.r_get('media-rdp/media-speed-capability/speed')])


def _p_operational_status_case(port_obj, k, wwn):
    os = port_obj.r_get('fibrechannel/operational-status')
    if os is not None:
        try:
            return brcddb_common.port_conversion_tbl['fibrechannel/operational-status'][os]
        except KeyError:
            pass
    return brcddb_common.port_conversion_tbl['fibrechannel/operational-status'][0]


def _p_port_type_case(port_obj, k, wwn):
    try:
        return brcddb_common.port_conversion_tbl['fibrechannel/port-type'][port_obj.r_get('fibrechannel/port-type')]
    except KeyError:
        return brcddb_common.port_conversion_tbl['fibrechannel/port-type'][brcddb_common.PORT_TYPE_UNKNOWN]


def _p_port_speed_case(port_obj, k, wwn):
    return port_obj.c_login_speed()


_port_case = {
    '_FABRIC_NAME': report_utils.fabric_name_case,
    '_FABRIC_NAME_AND_WWN': report_utils.fabric_name_or_wwn_case,
    '_FABRIC_WWN': report_utils.fabric_wwn_case,
    '_PORT_NUMBER': _p_port_number_case,
    '_BEST_DESC': _p_port_desc_case,
    '_MAPS_GROUP': _p_port_maps_group_case,
    '_SWITCH_NAME': _p_switch_name_case,
    '_SWITCH_NAME_AND_WWN': _p_switch_name_and_wwn_case,
    '_SWITCH_WWN': _p_switch_wwn_case,
    'media-rdp/media-distance': _p_media_distance_case,
    'media-rdp/media-speed-capability': _p_media_speed_case,
    'media-rdp/remote-media-speed-capability': _p_media_rspeed_case,
    'fibrechannel/operational-status': _p_operational_status_case,
    'fibrechannel/port-type': _p_port_type_case,
    'fibrechannel/speed': _p_port_speed_case,
    'fibrechannel-statistics/time-generated': _p_time_generated_case,
    'media-rdp/power-on-time': _p_media_uptime_case,
    '_ALIAS': _p_alias_case,
    '_PORT_COMMENTS': _p_comment_case,
    '_LOGIN_WWN': _p_login_wwn_case,
    '_LOGIN_ADDR': _p_login_addr_case,
    '_ZONES_DEF': _p_zones_def_case,
    '_ZONES_EFF': _zones_eff_case,
    '_NAME_SERVER_NODE': _p_name_server_node_case,
    '_NAME_SERVER_PORT': _p_name_server_port_case,
    '_FDMI_NODE': _p_fdmi_node_case,
    '_FDMI_PORT': _p_fdmi_port_case,
    'los-tov-mode-enabled': _p_los_tov_mode_case,
}
_port_link_case = {
    '_CONFIG_LINK': _p_port_config_link_case,
    '_STATS_LINK': _p_port_stats_link_case,
    '_ZONE_LINK': _p_port_zone_link_case,
    '_SFP_LINK': _p_port_sfp_link_case,
    '_RNID_LINK': _p_port_rnid_link_case,
}


def port_page(wb, tc, sheet_name, sheet_i, sheet_title, p_list, in_display=None, in_port_display_tbl=None,
              login_flag=False, link_type=None):
    """Creates a port detail worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: I don't think this is used anymore but if present, it will take precidence over what in the port object
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param p_list: List of port objects (PortObj) to display
    :type p_list: list, tuple
    :param in_display: List of parameters to display. If None, default is brcddb.app_data.report_tables.port_config_tbl
    :type in_display: list, tuple
    :param in_port_display_tbl: Display control table. If None, default is brcddb.report.report_tables.port_display_tbl
    :type in_port_display_tbl: dict, None
    :param login_flag: When True, include NPIV logins below the base port.
    :type login_flag: bool
    :param link_type: Type of link, see brcddb.apps.report. Valid types are: pc, ps, pz, pr, and sfp
    :type link_type: str, None
    :return: openpyxl sheet
    :rtype: Worksheet
    """
    global _port_case, _link_font, _bold_font, _hdr1_font, _std_font, _align_wrap, _align_wrap_c

    # I forgot all about logins and had to shoe horn in something to report the logins. This also meant I had to shoe
    # horn in a way to get the comments associated with logins, but only if login information was requested, into the
    # display. $ToDo: Clean this up

    addl_row = dict()

    # Validate the user input
    err_msg = list()
    if p_list is None:
        err_msg.append('p_list was not defined.')
    elif not isinstance(p_list, (list, tuple)):
        err_msg.append('p_list was type ' + str(type(p_list)) + '. Must be a list or tuple.')
    if len(err_msg) > 0:
        brcdapi_log.exception(err_msg, True)
        return None
    display = rt.Port.port_config_tbl if in_display is None else in_display
    port_display_tbl = rt.Port.port_display_tbl if in_port_display_tbl is None else in_port_display_tbl

    # Create the worksheet
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

    # Add the headers and set up the column widths
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    sheet.freeze_panes = sheet['A3']
    row, col = row+1, 1
    for k in display:
        if k in port_display_tbl:
            if 'dc' in port_display_tbl[k] and port_display_tbl[k]['dc'] is True:
                continue
            col_width = port_display_tbl[k].get('c')
            if isinstance(col_width, int):
                sheet.column_dimensions[xl.get_column_letter(col)].width = col_width
            buf = port_display_tbl[k]['d'] if bool(port_display_tbl[k].get('d')) else k
            alignment = _align_wrap_vc if bool(port_display_tbl[k].get('v')) else _align_wrap
        else:
            brcdapi_log.exception('Item ' + k + ' not in port_display_tbl.', True)
            alignment = _align_wrap
            buf = k
        excel_util.cell_update(sheet, row, col, buf, font=_bold_font, align=alignment, border=_border_thin)
        col += 1

    # Add the ports
    last_switch, row = '', row + 1
    for port_obj in p_list:
        if port_obj is None:
            continue  # For lazy programmers :-)
        col = 1
        cell = xl.get_column_letter(col) + str(row)
        if link_type is not None:
            if port_obj.r_switch_key() != last_switch:
                last_switch = port_obj.r_switch_key()
                brcddb_util.add_to_obj(port_obj.r_switch_obj(),
                                       'report_app/hyperlink/' + link_type,
                                       '#' + sheet_name + '!' + cell)
            brcddb_util.add_to_obj(port_obj, 'report_app/hyperlink/' + link_type, '#' + sheet_name + '!' + cell)
        # We don't want SIM ports
        login = [wwn for wwn in port_obj.r_login_keys() if port_obj.c_login_type() == 'F-Port']
        lwwn = None if len(login) == 0 else login[0]
        calc_font = report_utils.font_type(report_utils.combined_alert_objects(port_obj, lwwn))
        for k in display:
            if k in port_display_tbl and 'dc' in port_display_tbl[k] and port_display_tbl[k]['dc'] is True:
                continue
            if len(login) > 1:
                addl_row.update({k: col})
            alignment = _align_wrap_c if k in port_display_tbl and bool(port_display_tbl[k].get('m')) else _align_wrap
            link = None
            if k in _port_case:
                font = calc_font
                buf = _port_case[k](port_obj, k, lwwn)
            elif k in _port_link_case:
                buf, link = _port_link_case[k](port_obj, k, lwwn)
                font = calc_font if link is None else _link_font
            elif k in port_display_tbl:
                font = calc_font
                v = port_obj.r_get(k)
                if v is None:
                    v1 = ''
                else:
                    try:
                        v1 = brcddb_common.port_conversion_tbl[k][v]
                    except KeyError:
                        v1 = v
                if isinstance(v1, bool):
                    buf = '\u221A' if v1 else ''
                else:
                    buf = v1
            else:
                font = _std_font
                buf = '' if port_obj.r_get(k) is None else str(port_obj.r_get(k))
            excel_util.cell_update(sheet, row, col, buf, font=font, align=alignment, border=_border_thin, link=link)
            col += 1
        row += 1

        # Now add the remaining logins
        if login_flag:
            for i in range(1, len(login)):
                calc_font = report_utils.font_type(report_utils.combined_login_alert_objects(
                    port_obj.r_fabric_obj().r_login_obj(login[i])))
                for k1 in addl_row:
                    if k1 in _port_link_case:
                        buf, link = _port_link_case[k1](port_obj, k1, login[i])
                        font = calc_font if link is None else _link_font
                    else:
                        link, font = None, calc_font
                        if k1 in _port_case:
                            buf = _port_case[k1](port_obj, '', login[i]) if k1 == '_PORT_COMMENTS' \
                                else _port_case[k1](port_obj, k1, login[i])
                        else:
                            buf = port_obj.r_fabric_obj().r_get(k1)
                    excel_util.cell_update(sheet, row, col, '' if buf is None else buf, font=font, align=_align_wrap,
                                             border=_border_thin, link=link)
                row += 1

    return sheet
