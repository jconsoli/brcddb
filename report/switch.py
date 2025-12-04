"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack_consoli@yahoo.com for
details.

**To Add Port Highlighting**

Conditional highlighting is setup in brcddb.report.chassis.chassis_page(). Search for "_conditional_highlight". The
switch page, brcddb.report.switch.switch_page() is called first so that the links for the switch pages are known and
added to the chassis page. The switch page, however, leverages the port highlighting setup on the chassis page. This
circular conundrum was an "Oh crap" moment.

Processing pages and the order they are processed in is in brcddb.apps.report._add_chassis(). The pointer to
_add_chassis is loaded into brcddb.apps.report._chassis_control_d. The dictionary _chassis_control_d is loaded into a
list, report_l, in brcddb.apps.report.report().

The easiest thing to do is probably to insert a step that sets up the chassis page first, then go back and add the
switch page links. Perhaps it will make more sense to just copy and paste the chassis code to the switch page. I have
not fully thought this out yet.

**Public Methods & Data**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| add_port_map          | Adds the port map to the switch page                                                      |
+---------------------------+---------------------------------------------------------------------------------------+
| switch_page           | Creates a switch detail worksheet for the Excel report.                                   |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 20 Oct 2024   | Redesigned switch page                                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 26 Dec 2024   | Documentation updates only                                                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 25 Aug 2025   | Added SCC_POLICY.                                                                     |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 19 Oct 2025   | Updated comments only.                                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.6     | 04 Dec 2025   | Added note referencing the port page for port highlighting.                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025 Consoli Solutions, LLC'
__date__ = '04 Dec 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.6'

import openpyxl.utils.cell as xl
import copy
import collections
import brcddb.brcddb_common as brcddb_common
import brcdapi.util as brcdapi_util
import brcdapi.gen_util as gen_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.port as brcdapi_port
import brcddb.brcddb_switch as brcddb_switch
import brcddb.report.utils as report_utils
import brcddb.app_data.alert_tables as alert_table
import brcddb.util.util as brcddb_util

# Below is for efficiency
_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_gray_font = excel_fonts.font_type('gray')
_link_font = excel_fonts.font_type('link')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')
_common_default_d = dict(font=_std_font, align=_align_wrap, border=_border_thin)
_non_switch_default_d = dict(font=_gray_font, align=_align_wrap, border=_border_thin, fill=None, sw_cond_link=None)

_SWITCH_PARAM_C = 7  # Number of columns switch comments should span
_SWITCH_PARAM_P = 7  # Number of columns switch parameters should span
_SWITCH_PARAM_V = 12  # Number of columns switch values should span
# _cell_span_l is used when a=report_utils.cell_content, _switch_type, or _conv_to_bool
_cell_span_l = (_SWITCH_PARAM_C, _SWITCH_PARAM_P, _SWITCH_PARAM_V)
_SWITCH_LINK_C = 7  # Number of columns for the links at the top of the sheet
_PORT_SPEED = 4  # Number of columns for the port login speed.
_CONDITIONAL_C = 4  # Number of columns for the conditional port formatting
_PORT_COLS = 2  # Number of columns for the port map display
_PORT_SUM_C = 4  # Number of columns for the ports in the "Port Summary"
_common_width = 3  # The switch worksheet uses a single common column width

# _extra_columns: I forgot that the ports are inserted in brcddb.report.utils.port_map(). The port_map() utility,
# however, does not count the needed columns for the ports. Even if it did, the port map is determined long after the
# column widths are set. Some day, I may fix this, but for now, I just slammed in a somewhat arbitrary number to add to
# _num_columns when setting up the column widths.
_extra_columns = 12


class Found(Exception):
    pass


#########################################################
#                                                       #
#         Row insert functions for _contents            #
#                                                       #
#########################################################
def _links(switch_obj, col_d):
    """Useful links.

    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param col_d: Dictionary in the sub-list of _contents
    :type col_d: dict
    :return: List of dictionaries to insert while processing _contents
    :rtype: list
    """
    global _SWITCH_LINK_C, _align_wrap_c, _std_font, _link_font
    
    chassis_obj = switch_obj.r_chassis_obj()
    chassis_link = None if chassis_obj is None else chassis_obj.r_get('report_app/hyperlink/chassis')
    chassis_font = _std_font if chassis_link is None else _link_font
    fabric_obj = switch_obj.r_fabric_obj()
    fabric_link = None if fabric_obj is None else fabric_obj.r_get('report_app/hyperlink/fab')
    fabric_font = _std_font if fabric_link is None else _link_font
    port_link = switch_obj.r_get('report_app/hyperlink/pl')
    port_font = _std_font if port_link is None else _link_font

    return report_utils.add_content_defaults(
        [
            [
                dict(buf='Chassis Page', link=chassis_link, font=chassis_font, align=_align_wrap_c, border=None),
                dict(buf='Fabric Page', link=fabric_link, font=fabric_font, align=_align_wrap_c, border=None),
                dict(buf='Port Links Page', link=port_link, font=port_font, align=_align_wrap_c, border=None),
            ],
        ],
        dict(align=_align_wrap, span=_SWITCH_LINK_C)
    )


def _port_statistics(switch_obj, col_d):
    """Returns the port statistics to insert in _contents. See _links() for parameter definitions."""
    global _SWITCH_LINK_C, _PORT_SPEED

    return report_utils.port_statistics(switch_obj, _SWITCH_LINK_C, _PORT_SPEED)


def _conditional_highlight(switch_obj, col_d):
    """Returns the conditional highlighting to insert in _contents. See _links() for parameter definitions."""
    global _CONDITIONAL_C

    return report_utils.conditional_highlight(switch_obj, _CONDITIONAL_C)


#########################################################
#                                                       #
#        Column insert functions for _contents          #
#                                                       #
#########################################################
def _active_maps(switch_obj, col_d):
    """Returns a list of columns to display for the active MAPS policy.

    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param col_d: Dictionary in the sub-list of _contents
    :type col_d: dict
    :return: List of dictionaries with the comments, parameter, and setting values.
    :rtype: list
    """
    global _align_wrap

    maps_d = switch_obj.r_active_maps_policy()
    maps_policy = maps_d.get('name', None) if isinstance(maps_d, dict) else None
    span_l = report_utils.span_to_list(col_d.get('span'))

    return report_utils.add_content_defaults(
        [
            [
                dict(buf=None, span=span_l[0]),
                dict(buf='Active MAPS Policy', align=_align_wrap, span=span_l[1]),
                dict(buf=maps_policy, span=span_l[2]),
            ],
        ],
        dict(font=col_d.get('font'), align=col_d.get('align'), border=col_d.get('border'))
    )[0]


def _switch_type(switch_obj, col_d):
    """Returns a list of columns to display for the switch type. See _active_maps() for parameters"""
    r_col_d = col_d.copy()
    try:
        for k in (brcdapi_util.bfls_base_sw_en, brcdapi_util.bfls_def_sw_status, brcdapi_util.bfls_ficon_mode_en):
            if switch_obj.r_get(k, False):
                raise Found
    except Found:
        r_col_d.update(key=k)

    return report_utils.cell_content(switch_obj, r_col_d)


def _conv_to_bool(switch_obj, col_d):
    """Returns a list of columns to display for the switch type. See _active_maps() for parameters"""
    return report_utils.cell_content(switch_obj, col_d, old_bool=True)


def _defined_scc_policy(switch_obj, col_d):
    span_l = report_utils.span_to_list(col_d.get('span'))
    return report_utils.add_content_defaults(
        [
            [
                dict(buf=None, span=span_l[0]),
                dict(buf='Defined SC_POLICY', align=_align_wrap, span=span_l[1]),
                dict(buf='\n'.join(switch_obj.r_defined_scc(list())), span=span_l[2]),
            ],
        ],
        dict(font=col_d.get('font'), align=col_d.get('align'), border=col_d.get('border'))
    )[0]


def _active_scc_policy(switch_obj, col_d):
    global _std_font

    span_l = report_utils.span_to_list(col_d.get('span'))
    active_scc_l = switch_obj.r_active_scc(list())
    alert_obj = switch_obj.r_alert_obj(alert_table.ALERT_NUM.SWITCH_SCC_MATCH)
    if alert_obj is None:
        font = col_d.get('font')
        comment = None
    else:
        font = report_utils.alert_font_d.get(alert_obj.sev(), _std_font)
        comment = alert_obj.fmt_msg()

    return report_utils.add_content_defaults(
        [
            [
                dict(buf=comment, span=span_l[0]),
                dict(buf='Active SC_POLICY', align=_align_wrap, span=span_l[1]),
                dict(buf='\n'.join(active_scc_l), span=span_l[2]),
            ],
        ],
        dict(font=font, align=col_d.get('align'), border=col_d.get('border'))
    )[0]


_contents = [  # Search for **add_contents()** in brcddb.report.utils for an explanation
    # Links, switch name with DID and WWN, summary of alerts, and login banner at the top of the page.
    list(),
    _links,
    list(),
    [dict(buf='Summary of Alerts', font=_bold_font, border=None, span=0)],
    report_utils.alert_summary,
    list(),
    [dict(buf='Switch Login Banner', font=_bold_font, border=None, span=0)],
    [dict(buf=None, key=brcdapi_util.bfs_banner, font=_std_font, border=None, span=0)],
    list(),
    [  # Column headers for the remaining parameters
        dict(buf='Comment', font=_bold_font, span=_SWITCH_PARAM_C),
        dict(buf='Parameter', font=_bold_font, span=_SWITCH_PARAM_P),
        dict(buf='Setting', font=_bold_font, span=_SWITCH_PARAM_V),
        dict(font=None, align=None, border=None),
        dict(buf='Comment', font=_bold_font, span=_SWITCH_PARAM_C),
        dict(buf='Parameter', font=_bold_font, span=_SWITCH_PARAM_P),
        dict(buf='Setting', font=_bold_font, span=_SWITCH_PARAM_V),
    ],
    [  # Insistent Domain ID, Status
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_idid),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfs_op_status_str),
    ],
    [  # Switch Type, Access Gateway Mode
        dict(a=_switch_type, span=_cell_span_l, key='_open_switch'),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfs_ag_mode_str),
    ],
    [  # Fabric Principal,
        dict(a=_conv_to_bool, span=_cell_span_l, key=brcdapi_util.bfs_principal),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_port_id_mode),
    ],
    [
        dict(a=_conv_to_bool, span=_cell_span_l, key=brcdapi_util.bfls_isl_enabled),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_max_flogi_rate_port),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_free_fdisc),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_fport_enforce_login_str),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_stage_interval),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.seconds_to_days, border=_border_thin, span=_cell_span_l, key=brcdapi_util.bfc_up_time),
    ],
    [
        dict(a=_active_maps, span=_cell_span_l),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key='brocade-maps/system-resources/memory-usage'),
    ],
    [
        dict(a=report_utils.cell_content,
             span=_cell_span_l,
             key='brocade-maps/dashboard-misc/congestion-db-start-time'
             ),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key='brocade-maps/dashboard-misc/maps-start-time'),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_lc),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_bfid),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_ifl),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_sp_tags),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_mm),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_rc_ptde),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_lz_in_use),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_mld),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_mpds),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_ld_in_use),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_pds_in_use),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_mpd),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfr_stats_max_nr),
        dict(font=None, align=None, border=None),
        dict(
            a=report_utils.list_to_str,
            span=_cell_span_l,
            key='brocade-fibrechannel-switch/fibrechannel-switch/ip-static-gateway-list/ip-static-gateway'
        ),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfs_domain_name),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfs_fcid_hex),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfsw_uri + '/path-count'),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_portname_format)
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_max_logins),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_portname_mode),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfc_area_mode),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bfs_edge_hold)
    ],
    [dict(buf='FICON CUP & SCC_POLICY', font=_bold_font, border=None, span=0)],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_cup_en),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_posc),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_uam),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_asm),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_dcam),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_mihpto),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_uam_fru),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_uam_hsc),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_uam_invalid_attach),
        dict(font=None, align=None, border=None),
        dict(font=None, align=None, border=None),
    ],
    [
        dict(a=_defined_scc_policy, span=_cell_span_l, key=None),
        dict(font=None, align=None, border=None),
        dict(a=_active_scc_policy, span=_cell_span_l, key=None),
    ],
    [dict(buf='RNID', font=_bold_font, border=None, span=0)],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_flags),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_node_params),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_type),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_model),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_mfg),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_pant),
    ],
    [
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_seq),
        dict(font=None, align=None, border=None),
        dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.ficon_sw_rnid_tag),
    ],
    list(),
    _port_statistics,
    list(),
    _conditional_highlight,
    list(),
]
# Determine the maximum number of columns needed
_num_columns, _temp_span_l = 0, list()
for _temp_l in [_l for _l in _contents if isinstance(_l, list)]:
    _temp_columns = 0
    for _temp_d in [_d for _d in _temp_l if isinstance(_d, dict)]:
        _temp_span = _temp_d.get('span')
        if _temp_span is None:
            _temp_columns += 1
        elif isinstance(_temp_span, int):
            if _temp_span == 0:
                _temp_span_l.append(_temp_d)
            else:
                _temp_columns += _temp_span
        else:  # Assume it's a list or tuple
            for _x in _temp_span:
                _temp_columns += _x
    _num_columns = max(_num_columns, _temp_columns)
# Fill in the maximum columns
for _temp_d in _temp_span_l:
    _temp_d['span'] = _num_columns


def add_port_map(switch_obj):
    """Adds the port map to the switch page

    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :rtype: None
    """
    global _non_switch_default_d, _extra_columns

    content_l = list()
    chassis_obj = switch_obj.r_chassis_obj()
    if chassis_obj is None:
        content_l.append([
            dict(buf='Not enough chassis data collected to determine the port map',
                 span=switch_obj.r_get('report_app/worksheet/num_columns', _extra_columns)
                 )])
    else:
        port_map_l = gen_util.convert_to_list(chassis_obj.r_get('report_app/port_map_l'))

        # Copy the port map from the chassis. For any port not in the switch, display in gray with minimal formatting
        for row_l in port_map_l:
            col_l = list()
            content_l.append(col_l)
            for col_d in row_l:
                port_obj = col_d.get('port_obj')
                if port_obj is not None:
                    if switch_obj.r_obj_key() == port_obj.r_switch_obj().r_obj_key():
                        col_l.append(col_d)
                    else:
                        temp_d = _non_switch_default_d.copy()
                        temp_d.update(dict(buf=col_d['buf'], span=col_d['span'], comments=col_d['comments']))
                        col_l.append(temp_d)
                else:
                    col_l.append(col_d)

    report_utils.add_contents(switch_obj, content_l)


def _all_ports(port_obj, port_l):
    """Adds the port number to port_l

    :param port_obj: Port object
    :type port_obj: brcddb.classes.port.PortObject
    :param port_l: The list the matching port number should be added to.
    :type port_l: list
    """
    port_l.append(port_obj.r_obj_key())


def _enabled_logins(port_obj, port_l):
    """If the port is enabled with logins, the port number is added to port_l. See _all_ports() for parameters."""
    if port_obj.r_is_enabled() and len(port_obj.r_login_keys()) > 0:
        port_l.append(port_obj.r_obj_key())


def _enabled_no_logins(port_obj, port_l):
    """If the port is enabled with no logins, the port number is added to port_l. See _all_ports() for parameters."""
    if port_obj.r_is_enabled() and len(port_obj.r_login_keys()) == 0:
        port_l.append(port_obj.r_obj_key())


def _disabled(port_obj, port_l):
    """If the port is enabled with no logins, the port number is added to port_l. See _all_ports() for parameters."""
    if not port_obj.r_is_enabled():
        port_l.append(port_obj.r_obj_key())


def _e_port(port_obj, port_l):
    """If the port is an E-Port (ISL), the port number is added to port_l. See _all_ports() for parameters."""
    if port_obj.c_login_type() == 'E-Port':
        port_l.append(port_obj.r_obj_key())


def _f_port(port_obj, port_l):
    """If the port is an F-Port, the port number is added to port_l. See _all_ports() for parameters."""
    if port_obj.c_login_type() == 'F-Port':
        port_l.append(port_obj.r_obj_key())


def _u_or_n_port(port_obj, port_l):
    """If the port is a U-Port or N-Port, the port number is added to port_l. See _all_ports() for parameters."""
    port_type = port_obj.c_login_type()
    if port_type == 'U-Port' or port_type == 'N-Port':
        port_l.append(port_obj.r_obj_key())


def _fc4_target(port_obj, port_l):
    """If the attached device is storage, the port number is added to port_l. See _all_ports() for parameters."""
    attach_type = port_obj.r_get(brcdapi_util.bns_fc4_features)
    if isinstance(attach_type, str) and 'Target' in attach_type and 'Initiator' not in attach_type:
        port_l.append(port_obj.r_obj_key())


def _fc4_initiator(port_obj, port_l):
    """If the attached device is storage, the port number is added to port_l. See _all_ports() for parameters."""
    attach_type = port_obj.r_get(brcdapi_util.bns_fc4_features)
    if isinstance(attach_type, str) and 'Initiator' in attach_type and 'Target' not in attach_type:
        port_l.append(port_obj.r_obj_key())


def _fc4_mixed(port_obj, port_l):
    """If the attached device is storage, the port number is added to port_l. See _all_ports() for parameters."""
    attach_type = port_obj.r_get(brcdapi_util.bns_fc4_features)
    if isinstance(attach_type, str) and 'Initiator' in attach_type and 'Target' in attach_type:
        port_l.append(port_obj.r_obj_key())


def _fc4_other(port_obj, port_l):
    """If the attached device is storage, the port number is added to port_l. See _all_ports() for parameters."""
    attach_type = port_obj.r_get(brcdapi_util.bns_fc4_features)
    if isinstance(attach_type, str) and 'Initiator' not in attach_type and 'Target' not in attach_type:
        port_l.append(port_obj.r_obj_key())


def port_summary(switch_obj):
    """Returns the conditional highlighting to insert in _contents. See _links() for parameter definitions."""
    global _extra_columns, _std_font, _bold_font, _align_wrap_c, _align_wrap, _border_thin, _PORT_SUM_C
    global _common_default_d

    # Get a list of each port parameter type
    port_d = collections.OrderedDict()  # Key is the header. Value is the list of port numbers
    port_d['All'] = dict(l=list(),a=_all_ports)
    port_d['Enabled Logins'] = dict(l=list(),a=_enabled_logins)
    port_d['Enabled No Logins'] = dict(l=list(),a=_enabled_no_logins)
    port_d['Disabled'] = dict(l=list(),a=_disabled)
    port_d['E-Port'] = dict(l=list(),a=_e_port)
    port_d['F-Port'] = dict(l=list(),a=_f_port)
    port_d['U or N Port'] = dict(l=list(),a=_u_or_n_port)
    port_d['FC4 Storage'] = dict(l=list(),a=_fc4_target)
    port_d['FC4 Server'] = dict(l=list(),a=_fc4_initiator)
    port_d['Mixed FC4'] = dict(l=list(),a=_fc4_mixed)
    port_d['Other FC4'] = dict(l=list(),a=_all_ports)
    for port_obj in switch_obj.r_port_objects():
        for hdr, d in port_d.items():
            d['a'](port_obj, d['l'])

    # Add the headers to the worksheet
    # The Main Header
    full_span = switch_obj.r_get('report_app/worksheet/num_columns', _extra_columns)
    content_l = [
        [dict(buf='Port Summary', font=_bold_font, align=_align_wrap, border=None, span=full_span)],
        list()
    ]
    # Add the sub-headers
    row_l = list()
    for hdr in [str(k) for k in port_d.keys()]:
        row_l.append(dict(buf=hdr, font=_bold_font, align=_align_wrap_c, border=_border_thin, span=_PORT_SUM_C))
    content_l.append(row_l)
    # Update the worksheet
    report_utils.add_contents(switch_obj, content_l)

    # The ports should be in sorted order and 'All' should be the max. Below is future proofing. It also removes
    # duplicate port numbers. Duplicate port numbers shouldn't be in here either, so that is also future proofing.
    max_ports = 0
    for d in port_d.values():
        d['l'] = brcdapi_port.sort_ports(d['l'])
        max_ports = max(max_ports, len(d['l']))

    # Add the port numbers
    content_l = list()
    for i in range(0, max_ports):
        row_l = list()
        for d in port_d.values():
            try:
                row_l.append(dict(buf=d['l'][i], span=_PORT_SUM_C))
            except IndexError:
                row_l.append(dict(buf=None, span=_PORT_SUM_C))
        content_l.append(row_l)
    report_utils.add_contents(switch_obj, report_utils.add_content_defaults(content_l, _common_default_d))


def switch_page(wb, tc, sheet_name, sheet_i, sheet_title, switch_obj, in_display=None):
    """Creates a switch detail worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param in_display: Display table for switch parameters
    :type in_display: None, dict
    :rtype: None
    """
    global _contents, _num_columns, _extra_columns

    # Create the worksheet and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.add_data_validation(report_utils.highlight_stats_dv)
    for col in range(0, _num_columns + _extra_columns):
        sheet.column_dimensions[xl.get_column_letter(col+1)].width = _common_width
    sheet.freeze_panes = sheet['A2']
    sheet_d = dict(sheet=sheet, num_columns=_num_columns, row=1)  # Add cond_format_d=dict()
    switch_obj.rs_key('report_app/worksheet', sheet_d)

    # Add the table of contents and title to the first row.
    col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, sheet_d['row'], col, 'Contents', font=_link_font, link=tc)
        sheet.merge_cells(start_row=sheet_d['row'], start_column=col, end_row=sheet_d['row'], end_column=col+3)
        col += 4
    excel_util.cell_update(
        sheet,
        sheet_d['row'],
        col,
        brcddb_switch.best_switch_name(switch_obj, wwn=True, did=True, fid=True),
        font=_hdr1_font,
        align=_align_wrap_c
    )
    sheet.merge_cells(
        start_row=sheet_d['row'],
        start_column=col,
        end_row=sheet_d['row'],
        end_column=sheet_d['num_columns']
    )
    sheet_d['row'] += 1

    # Add the contents
    report_utils.add_contents(switch_obj, report_utils.add_content_defaults(_contents, _common_default_d))
