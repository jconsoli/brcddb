# Copyright 2023 Consoli Solutions, LLC.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
mod:`brcddb.report.bp` - Best practice violations

Public Methods::

    +-----------------------+---------------------------------------------------------------------------------------+
    | Method                | Description                                                                           |
    +=======================+=======================================================================================+
    | bp_page               | Creates a best practice violation worksheet for the Excel report.                     |
    +-----------------------+---------------------------------------------------------------------------------------+

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 4.0.0     | 04 Aug 2023   | Re-Launch                                                                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""

__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023 Consoli Solutions, LLC'
__date__ = '04 August 2023'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.0'

import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcddb.util.util as brcddb_util
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_chassis as brcddb_chassis
import brcddb.brcddb_switch as brcddb_switch
import brcddb.classes.alert as al
import brcddb.classes.util as class_util
import brcddb.brcddb_port as brcddb_port

_obj_type_link = dict(
    AliasObj='report_app/hyperlink/ali',
    ChassisObj='report_app/hyperlink/chassis',
    FabricObj='report_app/hyperlink/fab',
    LoginObj='report_app/hyperlink/log',
    FdmiNodeObj='report_app/hyperlink/log',
    FdmiPortObj='report_app/hyperlink/log',
    PortObj='report_app/hyperlink/pl',
    ProjectObj='report_app/hyperlink/tc',
    SwitchObj='report_app/hyperlink/sw',
    ZoneCfgObj='report_app/hyperlink/za',
    ZoneObj='report_app/hyperlink/za',
    # IOCPObj='report_app/hyperlink/',
    # ChpidObj='report_app/hyperlink/',
    )
_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')


def _add_alerts(obj, alert_type, sev, area_1, area_2):
    rl = list()
    for al_obj in obj.r_alert_objects():
        if al_obj.sev() == sev:
            class_type = class_util.get_simple_class_type(obj)
            rl.append(dict(type=alert_type,
                           sev=al_obj.fmt_sev(),
                           area_1=area_1,
                           area_2=area_2,
                           link=obj.r_get(_obj_type_link[class_type]) if class_type in _obj_type_link else None,
                           desc=al_obj.fmt_msg()))
    return rl


##################################################################
#
#        Best practice case statements in bp_page()
#
###################################################################
def _type_case(ad):
    """Returns the type value

    :param ad: Alert parameters dictionary as returned from _add_alerts()
    :type ad: dict
    :return desc: Value for element 'type'
    :rtype desc: str
    :return link: Hyperlink associated with desc being returned
    :rtype link: None, str
    """
    return ad.get('type'), None


def _sev_case(ad):
    """Returns the sev value. See _type_case() for parameters"""
    return ad.get('sev'), None


def _area_1_case(ad):
    """Returns the area_1 value. See _type_case() for parameters"""
    return ad.get('area_1'), None


def _area_2_case(ad):
    """Returns the area_2 value. See _type_case() for parameters"""
    return ad.get('area_2'), None


def _link_case(ad):
    """Returns the link value. See _type_case() for parameters"""
    return 'link', ad.get('link')


def _desc_case(ad):
    """Returns the desc value. See _type_case() for parameters"""
    return ad.get('desc'), None


bp_case = dict(
    _TYPE=_type_case,
    _SEV=_sev_case,
    _AREA_1=_area_1_case,
    _AREA_2=_area_2_case,
    _LINK=_link_case,
    _DESCRIPTION=_desc_case,
)


def bp_page(wb, tc, sheet_name, sheet_i, sheet_title, obj, display, display_tbl):
    """Creates a best practice violation worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param obj: Project or fabric object.
    :type obj: brcddb.classes.project.ProjectObj, brcddb.classes.fabric.FabricObj
    :param display: List of parameters to display.
    :type display: list, tuple
    :param display_tbl: Display control table. See brcddb.report.report_tables.display_tbl
    :rtype: None
    """
    global _link_font, _hdr1_font, _hdr2_font, _border_thin, _align_wrap, _align_wrap_vc

    # Validate the user input
    err_msg = list()
    if obj is None:
        err_msg.append('obj was not defined.')
    elif not bool('ProjectObj' in str(type(obj)) or 'FabricObj' in str(type(obj))):
        err_msg.append('Invalid object type: ' + str(type(obj)) + '.')
    if display is None:
        err_msg.append('display not defined.')
    if len(err_msg) > 0:
        brcdapi_log.exception(err_msg, True)
        return

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    brcddb_util.add_to_obj(obj, 'report_app/hyperlink/bp', '#' + sheet_name + '!A1')
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    sheet.freeze_panes = sheet['A3']
    row, col = row+1, 1
    for k in display:
        if k in display_tbl:
            if 'dc' in display_tbl[k] and display_tbl[k]['dc']:
                continue
            if 'c' in display_tbl[k]:
                sheet.column_dimensions[xl.get_column_letter(col)].width = display_tbl[k]['c']
            alignment = _align_wrap_vc if 'v' in display_tbl[k] and display_tbl[k]['v'] else _align_wrap
            buf = display_tbl[k]['d'] if 'd' in display_tbl[k] else k
        else:  # This happens when a new key is introduced before the display tables have been updated
            buf, alignment = k, _align_wrap
        excel_util.cell_update(sheet, row, col, buf, font=_bold_font, border=_border_thin, align=alignment)
        col += 1

    # Get a list of fabric objects and initialize alert_list
    if 'ProjectObj' in str(type(obj)):
        fab_list = obj.r_fabric_objects()
        chassis_list = obj.r_chassis_objects()
        switch_list = obj.r_switch_objects()
    else:
        fab_list = [obj]
        chassis_list = obj.r_project_obj().r_chassis_objects()
        switch_list = obj.r_project_obj().r_switch_objects()
    alert_list = list()

    # Get all the chassis alerts
    for tobj in chassis_list:
        tl = list()
        for sev in (al.ALERT_SEV.ERROR, al.ALERT_SEV.WARN):  # Display errors first
            tl.extend(_add_alerts(tobj, 'Chassis', sev, brcddb_chassis.best_chassis_name(tobj), ''))
        if len(tl) > 0:
            alert_list.append(dict(t=True, desc='Chassis: ' + brcddb_chassis.best_chassis_name(tobj)))

    # Get all the fabric alerts
    for fab_obj in fab_list:
        tl = list()
        for sev in (al.ALERT_SEV.ERROR, al.ALERT_SEV.WARN):  # Display errors first
            tl.extend(_add_alerts(fab_obj, 'Fabric', sev, brcddb_fabric.best_fab_name(fab_obj), ''))
            for tobj in fab_obj.r_zonecfg_objects():
                tl.extend(_add_alerts(tobj, 'ZoneCfg', sev, tobj.r_obj_key(), ''))
            for tobj in fab_obj.r_zone_objects():
                tl.extend(_add_alerts(tobj, 'Zone', sev, tobj.r_obj_key(), ''))
            for tobj in fab_obj.r_alias_objects():
                tl.extend(_add_alerts(tobj, 'Alias', sev, tobj.r_obj_key(), ''))
            for tobj in fab_obj.r_login_objects():
                tl.extend(_add_alerts(tobj, 'Login', sev, tobj.r_obj_key(), ''))
            for tobj in fab_obj.r_fdmi_node_objects():
                tl.extend(_add_alerts(tobj, 'FDMI_Node', sev, tobj.r_obj_key(), ''))
            for tobj in fab_obj.r_fdmi_port_objects():
                tl.extend(_add_alerts(tobj, 'FDMI_Port', sev, tobj.r_obj_key(), ''))
        if len(tl) > 0:
            alert_list.append(dict(t=True, desc='Fabric: ' + brcddb_fabric.best_fab_name(fab_obj, True)))
            alert_list.extend(tl)

    # Get all the switch and port alerts
    for switch_obj in switch_list:
        tl = list()
        for sev in (al.ALERT_SEV.ERROR, al.ALERT_SEV.WARN):  # Display errors first
            tl.extend(_add_alerts(switch_obj, 'Switch', sev, brcddb_switch.best_switch_name(switch_obj), ''))
            for tobj in switch_obj.r_port_objects():
                tl.extend(_add_alerts(tobj, 'Port', sev, brcddb_port.best_port_name(tobj), tobj.r_obj_key()))
        if len(tl) > 0:
            alert_list.append(dict(t=True, desc='Switch: ' + brcddb_switch.best_switch_name(switch_obj, True)))
            alert_list.extend(tl)

    # Now add the alerts to the worksheet

    # Add all alerts to the worksheet
    for d in alert_list:
        row, col = row+1, 1
        if 't' in d and d['t']:
            row += 1
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(display))
            excel_util.cell_update(sheet, row, col, d.get('desc'), font=_hdr2_font,
                                   fill=excel_fonts.fill_type('lightblue'))
        else:
            for k in display:
                buf, link = bp_case[k](d)
                excel_util.cell_update(sheet, row, col, buf, font=_std_font if link is None else _link_font,
                                       align=_align_wrap, border=_border_thin, link=link)
                col += 1
