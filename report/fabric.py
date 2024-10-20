"""
Copyright 2023, 2024 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

:mod:`brcddb.report.fabric` - Creates a fabric page to be added to an Excel Workbook

**Public Methods & Data**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| fabric_page           | Creates the fabric summary page                                                           |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 20 Oct 2024   | PEP 8 corrections to login speeds in brcddb.util.search.login_xxx                     |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024 Consoli Solutions, LLC'
__date__ = '20 Oct 2024'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.2'

import collections
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcddb.brcddb_common as brcddb_common
import brcddb.brcddb_switch as brcddb_switch
import brcddb.app_data.alert_tables as al
import brcddb.classes.util as brcddb_class_util
import brcddb.util.search as brcddb_search

_hdr = collections.OrderedDict()
# Key is the column header and value is the width
_hdr['Name'] = 30
_hdr['WWN'] = 22
_hdr['DID'] = 12
_hdr['Fabric ID'] = 12
_hdr['Firmware'] = 22
_zone_key_conv = {
    'cfg-action': 'Configuration actions',
    'cfg-name': 'Effective Configuration',
    'checksum': 'Checksum',
    'db-avail': 'Available zone database size (bytes)',
    'db-committed': 'Size of committed defined zone database (bytes)',
    'db-max': 'Maximum permitted zone database size (bytes)',
    'db-transaction': 'Size (bytes) to commit the pending transaction',
    'transaction-token': 'Transaction token',
    'default-zone-access': 'All Access',
}
_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')


def _setup_worksheet(wb, tc, sheet_i, sheet_name, sheet_title):
    """Creates a fabric detail worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :return sheet: Sheet object
    :rtype sheet: Worksheet
    :return row: Next row number
    :rtype row: int
    """
    global _hdr, _link_font, _hdr1_font

    # Create the worksheet, add the headers, and set up the column widths
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        col += 1
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr1_font)
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
    sheet.freeze_panes = sheet['A2']

    # Set the column width
    col = 1
    for k, v in _hdr.items():
        sheet.column_dimensions[xl.get_column_letter(col)].width = v
        col += 1

    return sheet, row + 1


def _fabric_summary(sheet, row, fabric_obj):
    """Adds a fabric summary to the worksheet.

    :param sheet: Workbook sheet object
    :type sheet: worksheet
    :param row: Starting row number
    :param row: int
    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :return: Next row number
    :rtype: int
    """
    global _hdr, _std_font, _border_thin, _bold_font, _link_font

    # Add the headers
    col = 1
    for k in _hdr.keys():
        excel_util.cell_update(sheet, row, col, k, font=_bold_font, align=_align_wrap, border=_border_thin)
        col += 1
    row += 1

    # Add the switch summary
    for switch_obj in fabric_obj.r_switch_objects():
        col = 1

        # Switch name
        buf = brcddb_switch.best_switch_name(switch_obj, False)
        if switch_obj.r_is_principal():
            buf = '*' + buf
        link = switch_obj.r_get('report_app/hyperlink/switch')
        font = _std_font if link is None else _link_font
        excel_util.cell_update(sheet, row, col, buf, font=font, align=_align_wrap, border=_border_thin, link=link)
        col += 1

        # Switch WWN
        excel_util.cell_update(sheet, row, col, switch_obj.r_obj_key(), font=_std_font, align=_align_wrap,
                               border=_border_thin)
        col += 1

        # Switch DID
        buf = switch_obj.r_get(brcdapi_util.bfs_did)
        if buf is None:
            buf = switch_obj.r_get('brocade-fabric/fabric-switch/domain-id')
        excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap, border=_border_thin)
        col += 1

        # Switch FID
        excel_util.cell_update(sheet, row, col, brcddb_switch.switch_fid(switch_obj), font=_std_font,
                               align=_align_wrap, border=_border_thin)
        col += 1

        # Firmware version
        buf = switch_obj.r_get(brcdapi_util.bfs_fw_version)
        if buf is None:
            buf = switch_obj.r_get(brcdapi_util.bf_fw_version)
        excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap, border=_border_thin)
        row += 1

    # Principal switch footnote
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
    excel_util.cell_update(sheet, row, col, '* indicates principal switch', font=_std_font)

    return row + 1


def _maps_dashboard(sheet, row, fabric_obj):
    """Adds the MAPS dashboard to the worksheet.

    :param sheet: Workbook sheet object
    :type sheet: worksheet
    :param row: Starting row number
    :param row: int
    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :return: Next row number
    :rtype: int
    """
    global _hdr, _align_wrap, _bold_font

    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
    excel_util.cell_update(sheet, row, col, 'MAPS Dashboard Alerts', font=_bold_font, align=_align_wrap)
    i = 0
    for alert_obj in fabric_obj.r_alert_objects():
        if alert_obj.alert_num() in al.AlertTable.maps_alerts:
            row += 1
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
            for col in range(1, len(_hdr)):
                excel_util.cell_update(sheet, row, col, None, border=_border_thin)
            col = 1
            excel_util.cell_update(sheet, row, col, alert_obj.fmt_msg(), font=_std_font, align=_align_wrap)
            i += 1
    if i == 0:
        row += 1
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
        for col in range(1, len(_hdr) + 1):
            excel_util.cell_update(sheet, row, col, None, border=_border_thin)
        col = 1
        excel_util.cell_update(sheet, row, col, 'None', font=_std_font, align=_align_wrap)

    return row + 1


def _fabric_statistics(sheet, row, fabric_obj):
    """Adds the zone statistics summary to the worksheet.

    :param sheet: Workbook sheet object
    :type sheet: worksheet
    :param row: Starting row number
    :param row: int
    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :return: Next row number
    :rtype: int
    """
    global _hdr, _border_thin, _align_wrap, _bold_font, _std_font

    # Add the section header
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr)+1)
    excel_util.cell_update(sheet, row, col, 'Fabric Statistics', font=_bold_font, align=_align_wrap)

    # Figure out what to put in the statistics summary section
    fab_stats_d = collections.OrderedDict()
    port_obj_l = fabric_obj.r_port_objects()
    fab_stats_d['Physical Ports'] = len(port_obj_l)
    fab_stats_d['ICL-Ports'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.icl_ports))
    fab_stats_d['ISL (E-Ports)'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.e_ports))
    fab_stats_d['FC-Lag Ports'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.fc_lag_ports))
    port_obj_l = brcddb_search.match_test(port_obj_l, brcddb_search.f_ports)
    sum_logins = 0
    for port_obj in port_obj_l:
        sum_logins += len(port_obj.r_login_keys())
    fab_stats_d['Name Server Logins'] = sum_logins
    fab_stats_d['Port Logins at 1G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_1g))
    fab_stats_d['Port Logins at 2G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_2g))
    fab_stats_d['Port Logins at 4G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_4g))
    fab_stats_d['Port Logins at 8G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_8g))
    fab_stats_d['Port Logins at 16G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_16g))
    fab_stats_d['Port Logins at 32G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_32g))
    fab_stats_d['Port Logins at 64G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_64g))
    fab_stats_d['Port Logins at 128G'] = len(brcddb_search.match_test(port_obj_l, brcddb_search.login_128g))

    # Add the statistics summary items to the sheet
    for k, v in fab_stats_d.items():
        row, col = row+1, 1
        excel_util.cell_update(sheet, row, col, k, font=_std_font, align=_align_wrap)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        col += 2
        excel_util.cell_update(sheet, row, col, v, font=_std_font)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
        for col in range(1, len(_hdr) + 1):
            excel_util.cell_update(sheet, row, col, None, border=_border_thin)

    return row+1


def _zone_configuration(sheet, row, fabric_obj):
    """Adds the zone configuration summary to the worksheet.

    :param sheet: Workbook sheet object
    :type sheet: worksheet
    :param row: Starting row number
    :param row: int
    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :return: Next row number
    :rtype: int
    """
    global _hdr, _zone_key_conv, _border_thin, _align_wrap, _bold_font, _std_font

    obj = fabric_obj.r_eff_zone_cfg_obj()

    # Add the Defined Configurations
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr)+1)
    excel_util.cell_update(sheet, row, col, 'Defined Configurations', font=_bold_font, align=_align_wrap,
                           border=_border_thin)
    for obj in fabric_obj.r_zonecfg_objects():
        buf = obj.r_obj_key()
        if buf == '_effective_zone_cfg':
            continue
        row += 1
        if obj.r_is_effective():
            buf = '*' + buf
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
        for col in range(1, len(_hdr)+1):
            excel_util.cell_update(sheet, row, col, None, border=_border_thin)
        col = 1
        excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap)

    # Effective zone footnote
    row, col = row+1, 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
    excel_util.cell_update(sheet, row, col, '* indicates effective zone configuration', font=_std_font)

    # Links to zoning pages
    row += 2
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
    excel_util.cell_update(sheet, row, col, 'Zone Analysis & Login Links', font=_bold_font)
    row += 1
    for d in (dict(t=fabric_obj.r_get('report_app/control/za/tc'), l=fabric_obj.r_get('report_app/hyperlink/za')),
              dict(t=fabric_obj.r_get('report_app/control/zt/tc'), l=fabric_obj.r_get('report_app/hyperlink/zt')),
              dict(t=fabric_obj.r_get('report_app/control/znt/tc'), l=fabric_obj.r_get('report_app/hyperlink/znt')),
              dict(t=fabric_obj.r_get('report_app/control/log/tc'), l=fabric_obj.r_get('report_app/hyperlink/log'))):
        if d['l'] is not None:
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
            excel_util.cell_update(sheet, row, col, d['t'], font=_link_font, link=d['l'])
            row += 1

    # Zone configuration summary
    if obj is not None:
        row, col = row+1, 1
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
        excel_util.cell_update(sheet, row, col, 'Zone Configuration Summary', font=_bold_font)
        ec_obj = fabric_obj.r_get('brocade-zone/effective-configuration')
        if isinstance(ec_obj, dict):
            for k in ec_obj:
                row, col = row+1, 1
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
                excel_util.cell_update(sheet, row, col, _zone_key_conv[k] if k in _zone_key_conv else k,
                                       font=_std_font, align=_align_wrap)
                col += 2
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
                v = ec_obj.get(k)
                if isinstance(v, (str, int)):
                    try:
                        buf = brcddb_common.zonecfg_conversion_tbl[k][v]
                    except KeyError:
                        buf = v
                    excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_align_wrap)
                for col in range(1, len(_hdr)+1):
                    excel_util.cell_update(sheet, row, col, None, border=_border_thin)
        else:
            row, col = row+1, 1
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            excel_util.cell_update(sheet, row, col, 'No effective configuration', font=_std_font, align=_align_wrap)

        # Total alias, zone, and zone configuration summary
        zone_l = [dict(d='Total Aliases', v=len(fabric_obj.r_alias_keys())),
                  dict(d='Total Zones', v=len(fabric_obj.r_zone_keys())),
                  dict(d='Total Zone Configurations', v=len(fabric_obj.r_zonecfg_keys()))]
        for d in zone_l:
            row, col = row+1, 1
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            excel_util.cell_update(sheet, row, col, d['d'], font=_std_font, align=_align_wrap, border=_border_thin)
            excel_util.cell_update(sheet, row, col+1, None, border=_border_thin)
            col += 2
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_hdr))
            excel_util.cell_update(sheet, row, col, d['v'], font=_std_font, align=_align_wrap, border=_border_thin)
            for col in range(col+1, len(_hdr)+1):
                excel_util.cell_update(sheet, row, col, None, border=_border_thin)

    return row + 1


def fabric_page(wb, tc, sheet_i, sheet_name, sheet_title, fabric_obj):
    """Creates the fabric summary page

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_i: Relative location for this worksheet. Default is 0
    :type sheet_i: int
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_title: Title for sheet
    :type sheet_title: str
    :param fabric_obj: Fabric object
    :type fabric_obj: FabricObj
    :rtype: None
    """
    # Validate the user input
    err_msg = list()
    if fabric_obj is None:
        err_msg.append('FabricObj was not defined.')
    elif brcddb_class_util.get_simple_class_type(fabric_obj) != 'FabricObj':
        err_msg.append('Wrong object type, ' + str(type(fabric_obj)) + 'Must be brcddb.classes.fabric.FabricObj.')
    if len(err_msg) > 0:
        brcdapi_log.exception(err_msg, echo=True)
        return

    # Set up the worksheet and add the sections to the fabric sheet
    sheet, row = _setup_worksheet(wb, tc, 0 if sheet_i is None else sheet_i, sheet_name, sheet_title)
    for method in (_fabric_summary, _maps_dashboard, _zone_configuration, _fabric_statistics):
        row = method(sheet, row+1, fabric_obj)

    return
