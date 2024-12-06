"""
Copyright 2023, 2024 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

:mod:`brcddb.report.chassis` - Creates a chassis page to be added to an Excel Workbook

**Public Methods**

+---------------------------+---------------------------------------------------------------------------------------+
| Method                    | Description                                                                           |
+===========================+=======================================================================================+
| add_port_map              | Adds the port map to chassis page                                                     |
+---------------------------+---------------------------------------------------------------------------------------+
| chassis_page              | Creates a chassis detail worksheet for the Excel report.                              |
+---------------------------+---------------------------------------------------------------------------------------+
| chassis_hidden_port_page  | Adds the hidden chassis port page which is used for port conditional highlighting     |
+---------------------------+---------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 26 Jun 2024   | Added firmware version and missing alerts to chassis report.                          |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 20 Oct 2024   | Added hidden chassis port page, chassis_hidden_port_page()                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 06 Dec 2024   | Added entitlement S/N to chassis report.                                              |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024 Consoli Solutions, LLC'
__date__ = '06 Dec 2024'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.4'

import collections
import copy
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.excel_util as excel_util
import brcdapi.port as brcdapi_port
import brcddb.brcddb_chassis as brcddb_chassis
import brcddb.brcddb_switch as brcddb_switch
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_login as brcddb_login
import brcddb.report.utils as report_utils
import brcddb.app_data.alert_tables as al
import brcddb.classes.alert as alert_class
import brcddb.util.obj_convert as obj_convert
import brcddb.util.iocp as brcddb_iocp

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr1_font = excel_fonts.font_type('hdr_1')
_error_font = excel_fonts.font_type('error')
_gray_font = excel_fonts.font_type('gray')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')
_fill_lightblue = excel_fonts.fill_type('lightblue')
_fill_lightgreen = excel_fonts.fill_type('lightgreen')
_fill_lightgold = excel_fonts.fill_type('lightgold')
_fill_lightred = excel_fonts.fill_type('lightred')
_fill_yellow = excel_fonts.fill_type('yellow')

_common_default_d = dict(font=_std_font, align=_align_wrap, border=_border_thin)
_fru_span = 6  # The number of columns to span for FRU components
_fru_default_d = dict(font=_std_font, align=_align_wrap, border=_border_thin, span=_fru_span)
_hdr_span = 30  # Unlike the switch page, there are too many columns to automatically span them all
_comment_span = 10
_parameter_span = 10
_setting_span = 24
_detail_param_span = 6
_port_speed = 4
_detail_value_span = 4
_port_col_span = 2  # Number of columns for the port map display
_fixed_ge_port_col_span = 3  # NUmber of columns for the GE ports in the map display
_bladed_ge_port_col_span = 4  # Number of columns for the GE ports on a blade in the map display

_conditional_c = 4  # Number of columns for the conditional port formatting
_cell_span_l = (_comment_span, _parameter_span, _setting_span)
_detail_span_l = (_comment_span,  # Comments
                  _detail_param_span,  # Part Number
                  _detail_param_span,  # S/N
                  _detail_param_span,  # Blade ID, Airflow, or None for Power Supplies
                  _detail_value_span,  # State
                  _detail_value_span,  # Type, Speed, or Source
                  _detail_value_span,  # Port Count or Unit Number
                  _detail_value_span,  # Time Alive
                  _detail_value_span,  # Time Awake
                  # The remainder apply to Blades only
                  _detail_value_span,  # Extension Enabled
                  _detail_value_span,  # Slot
                  _detail_value_span,  # Power Consumed
                  _detail_value_span)  # Power Used
# _extra_columns: I forgot that the port map is determined long after the column widths are set. Some day, I may fix
# this, but for now, I just slammed in a somewhat arbitrary number to add to _num_columns when setting column widths.
_extra_columns = 78
_common_width = 3  # The chassis worksheet uses a single common column width

# _blade_keys_d is used in _blades() to determine the column display header and the order of the display.
_blade_keys_d = collections.OrderedDict()
# _blade_keys_d['slot-number']  - Added manually with some cell highlighting at the top of each column
_blade_keys_d['part-number'] = 'Part Number'
_blade_keys_d['serial-number'] = 'Serial Number'
_blade_keys_d['firmware-version'] = 'Firmware'
_blade_keys_d['blade-id'] = 'Blade ID'
_blade_keys_d['blade-state'] = 'State'
_blade_keys_d['blade-type'] = 'Type'
_blade_keys_d['fc-port-count'] = 'Port Count'
_blade_keys_d['time-alive'] = 'Time Alive (days)'
_blade_keys_d['time-awake'] = 'Time Awake (days)'
_blade_keys_d['extension-enabled'] = 'Extension Enabled'
_blade_keys_d['power-consumption'] = 'Power Consumed'
_blade_keys_d['power-usage'] = 'Power Usage'
_blade_keys_d['manufacturer'] = 'Manufacturer'
_blade_keys_d['vendor-id'] = 'Vendor ID'
_blade_skip_keys_d = {'slot-number': True}  # Keys to skip in the general listing of parameters and values

_fan_keys_d = collections.OrderedDict()
# _fan_keys_d['unit-number'] - Added manually with some cell highlighting at the top of each column
_fan_keys_d['airflow-direction'] = 'Airflow'
_fan_keys_d['part-number'] = 'Part Number'
_fan_keys_d['serial-number'] = 'Serial Number'
_fan_keys_d['operational-state'] = 'State'
_fan_keys_d['speed'] = 'Speed'
_fan_keys_d['manufacture-date'] = 'Manufactured'
_fan_keys_d['time-awake'] = 'Time Awake (days)'
_fan_skip_keys_d = {'unit-number': True}  # Keys to skip in the general listing of parameters and values

_ps_keys_d = collections.OrderedDict()
# _ps_keys_d['unit-number'] - Added manually with some cell highlighting at the top of each column
_ps_keys_d['part-number'] = 'Part Number'
_ps_keys_d['serial-number'] = 'Serial Number'
_ps_keys_d['operational-state'] = 'State'
_ps_keys_d['input-voltage'] = 'Input Voltage'
_ps_keys_d['power-source'] = 'Power Source'
_ps_keys_d['power-usage'] = 'Power Usage'
_ps_keys_d['manufacture-date'] = 'Manufactured'
_ps_keys_d['time-awake'] = 'Time Awake (days)'
_ps_keys_d['temperature-sensor-supported'] = 'Temp Sensor Supported'
_ps_skip_keys_d = {'unit-number': True}  # Keys to skip in the general listing of parameters and values

_sensor_keys_d = collections.OrderedDict()
# _sensor_keys_d[id'] - Added manually with some cell highlighting at the top of each column
_sensor_keys_d['category'] = 'Sensor Type'
_sensor_keys_d['state'] = 'State'
_sensor_keys_d['temperature'] = 'Temperature C'
_sensor_skip_keys_d = dict(id=True, index=True)  # Keys to skip in the general listing of parameters and values

_wwn_keys_d = collections.OrderedDict()
# _wwn_keys_d['unit-number''] - Added manually with some cell highlighting at the top of each column
_wwn_keys_d['airflow-direction'] = 'Airflow'
_wwn_keys_d['part-number'] = 'Part Number'
_wwn_keys_d['serial-number'] = 'Serial Number'
_wwn_keys_d['header-version'] = 'Header Version'
_wwn_keys_d['last-header-update-date'] = 'Last Header Update'
_wwn_keys_d['manufacture-date'] = 'Manufactured'
_wwn_keys_d['time-alive-in-hours'] = 'Time Alive (hours)'
_wwn_keys_d['time-awake-in-hours'] = 'Time Awake (hours)'
_wwn_keys_d['vendor-id'] = 'Vendor ID'
_wwn_keys_d['vendor-part-number'] = 'Vendor Part Number'
_wwn_keys_d['vendor-revision-number'] = 'Vendor Revision'
_wwn_keys_d['vendor-serial-number'] = 'Vendor Serial Number'
_wwn_skip_keys_d = {'unit-number': True}  # Keys to skip in the general listing of parameters and values

_port_map_default_d = dict(font=_std_font, align=_align_wrap_c, border=_border_thin)
_fixed_port_map_l = [  # Used in add_port_map()
    [0, 1, 2, 3,
     8, 9, 10, 11,
     16, 17, 18, 19,
     24, 25, 26, 27,
     32, 33, 34, 35,
     40, 41, 42, 43,
     48, 49, 50, 51,
     56, 57, 58, 59,],
]
_fixed_port_map_l.append([_p + 4 for _p in _fixed_port_map_l[0]])
_fixed_port_map_l.append([_p + 64 for _p in _fixed_port_map_l[0]])
_fixed_port_map_l.append([_p + 64 for _p in _fixed_port_map_l[1]])
_fill_d = {  # Used in _insert_port() to determine the fill color
    'fcoe-port': _fill_lightblue,
    'e-port': _fill_lightblue,
    'ex-port': _fill_lightblue,
    've-port': _fill_lightblue,
    'ethernet-port': _fill_lightblue,
    'universal-port': _fill_lightgold,
    'd-port': _fill_lightgold,
    'fc-lag-port': _fill_lightblue,
    'f-port': _fill_lightgreen,
    'n-port': _fill_lightgreen,
    # Pre FOS 9.0 - As converted by brcddb.brcddb_common.port_conversion_tbl.fibrechannel.port-type
    'E-Port': _fill_lightblue,
    'G-Port': _fill_lightgold,
    'U-Port': _fill_lightgold,
    'F-Port': _fill_lightgreen,
    'L-Port': _fill_lightred,  # Loop port - no longer supported
    'FCoE-Port': _fill_lightblue,
    'EX-Port': _fill_lightblue,
    'D-Port': _fill_lightgold,
    'SIM-Port': _fill_lightgold,
    'AF-Port': _fill_lightgold,
    'AE-Port': _fill_lightgold,
    'VE-Port': _fill_lightblue,
    'Ethernet Flex Port': _fill_lightblue,
    'Flex Port': _fill_lightblue,
    'N-Port': _fill_lightgreen,
    'Mirror': _fill_lightgold,
    'ICL': _fill_lightblue,
    'FC-LAG': _fill_lightblue,
    'LB-Port': _fill_lightgold,
}
_port_legend_span = 12
_port_legend = [  # Used in add_port_map()
    [
        dict(buf='x', font=_std_font, align=_align_wrap_c, border=_border_thin, fill=_fill_lightgreen),
        dict(buf='F or N Port', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_std_font, align=_align_wrap_c, border=_border_thin, fill=_fill_lightblue),
        dict(buf='E, FCOE, GigE, or Lag Port', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_std_font, align=_align_wrap_c, border=_border_thin, fill=_fill_lightgold),
        dict(buf='Enabled, no logins (offline)', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_bold_font, align=_align_wrap_c, border=_border_thin, fill=_fill_lightred),
        dict(buf='Port faulty, testing, or unknown', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_bold_font, align=_align_wrap_c, border=_border_thin, fill=_fill_yellow),
        dict(
            buf='Highlighted (met port highlighting criteria)',
            font=_std_font,
            align=_align_wrap,
            span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_std_font, align=_align_wrap_c, border=_border_thin),
        dict(buf='Disabled', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
    [
        dict(buf='x', font=_gray_font, align=_align_wrap_c, border=_border_thin),
        dict(buf='Not in this logical switch', font=_std_font, align=_align_wrap, span=_port_legend_span),
    ],
]


class Found(Exception):
    pass


def _insert_port(chassis_obj, port_obj, span):
    """Determines the column dictionary for a specific port.

    :param chassis_obj: Switch or chassis object
    :type chassis_obj: brcddb.classes.chassis.ChassisObj
    :param port_obj: Port number in s/p notation
    :type port_obj: brcddb.classes.port.PortObj
    :param span: Number of columns to span (merge) for the port.
    :type span: int
    :return: The cell definitions for a single port
    :rtype: dict
    """
    global _std_font, _link_font, _align_wrap_c, _border_thin, _fill_lightred

    fill, link = None, None

    # Excel comment for this port - Includes status, port type, and alerts
    status = port_obj.r_status()
    comment_l = [
        'Port:   ' + brcddb_port.best_port_name(port_obj, port_num=True),
        'Status: ' + status,
        'Switch: ' + brcddb_switch.best_switch_name(port_obj.r_switch_obj(), fid=True),
    ]
    rnid_d = port_obj.r_get('rnid')
    if isinstance(rnid_d, dict):
        comment_l.extend([
            'RNID:',
            ' Mfg:   ' + rnid_d.get('manufacturer', ''),
            ' Model: ' + rnid_d.get('model-number', ''),
            ' S/N:   ' + rnid_d.get('sequence-number', ''),
            ' Tag:   ' + rnid_d.get('tag', ''),
            ' Type:  ' + brcddb_iocp.dev_type_desc(rnid_d.get('type-number')),
            ' Link Addr: ' + port_obj.r_get(brcdapi_util.fc_fcid_hex, 'Unknown').replace('0x', '')[:4].upper()
        ])
    fabric_obj = port_obj.r_fabric_obj()
    login_l = [' ' + brcddb_login.best_login_name(fabric_obj, obj.r_obj_key()) for obj in port_obj.r_login_objects()]
    comment_l.extend(['Logins: None'] if len(login_l) == 0 else ['Logins:'] + login_l)
    alert_l = ['  ' + obj.fmt_msg() for obj in port_obj.r_alert_objects() if obj.sev() >= alert_class.ALERT_SEV.WARN]
    comment_l.extend(['Alerts: None'] if len(alert_l) == 0 else ['Alerts:'] + alert_l)

    # Determine how to display the port - fill color, font, link, and additional comments
    port_num = port_obj.r_obj_key().split('/')[1]
    if port_num.isnumeric():
        port_num = int(port_num)
    fill = _fill_lightred if 'line' not in status else None  # Anything other than "offline" or "online" is a port error
    port_type = port_obj.c_login_type()
    if isinstance(port_type, str):
        if fill is None:
            fill = _fill_d.get(port_type, None)
        if fill is None:
            buf = 'ERROR: Unknown port type. ' + port_type + ' for port ' + str(port_num) + ' in chassis: ' + \
                  brcddb_chassis.best_chassis_name(chassis_obj, wwn=True)
            brcdapi_log.log(buf, echo=True)
    elif port_type is not None:
        buf = 'ERROR: Invalid data type for port type: ' + str(type(port_type)) + '. Chassis: ' + \
              brcddb_chassis.best_chassis_name(chassis_obj, wwn=True) + '. Port: ' + str(port_num)
        brcdapi_log.exception(buf, echo=True)
    link = port_obj.r_get('report_app/hyperlink/pl')

    return dict(
        buf=port_num,
        font=_std_font if link is None else _link_font,
        align=_align_wrap_c,
        border=_border_thin,
        link=link,
        fill=fill,
        span=span,
        comments='\n'.join(comment_l),
        ch_cond_link=port_obj.r_get('report_app/hyperlink/ch_highlight/total'),
        sw_cond_link=port_obj.r_get('report_app/hyperlink/sw_highlight/total'),  # For future use
        fab_cond_link=port_obj.r_get('report_app/hyperlink/fab_highlight/total'),  # For future use
        port_obj=port_obj,
    )


def _fru(chassis_obj, key_d, fru_key, unit, unit_hdr, state, default_d, skip_d):
    """Returns list of rows and list of columns for the FRU.

    :param chassis_obj: Chassis object
    :type chassis_obj: brcddb.classes.chassis.ChassisObj
    :param key_d: Key table. _blade_keys_d, _fan_keys_d, _ps_keys_d
    :type key_d: dict
    :param fru_key: Key in chassis object for the FRU. brocade-fru/fru_key
    :type fru_key: str
    :param unit: Unit identifier in the FRU data returned form FOS. 'unit-number', 'unit', 'slot-number', or 'id'
    :type unit: str
    :param unit_hdr: Title for the unit's header. Typically, 'Unit' or 'Slot'
    :type unit_hdr: str
    :param state: FRU state identifier in the FRU data returned form FOS. 'operational-state' or 'blade-state'
    :type state: str
    :param default_d: Common defaults for the returned dictionaries.
    :type default_d: dict
    :param skip_d: Skip keys that are True in this dictionary
    :type skip_d: dict
    :return: Row list of column list of dictionaries to insert while processing _contents
    :rtype: list
    """
    global _fill_lightblue, _error_font, _bold_font

    row_l = list()
    fru_l = gen_util.sort_obj_num(chassis_obj.r_get(fru_key, list()), unit)

    # Add the top row column headers - slot numbers
    col_l = [dict(border=None)]  # Leave the upper left cell blank with no border
    row_l.append(col_l)
    for fru_d in fru_l:
        col_l.append(
            dict(
                buf=unit_hdr + ' ' + str(fru_d.get(unit, '?')),
                font=_bold_font,
                fill=_fill_lightblue,
                align=_align_wrap_c
            )
        )

    # Add the row headers and values
    for key, buf in key_d.items():
        col_l = [dict(buf=buf)]  # Row header
        row_l.append(col_l)
        for fru_d in fru_l:
            col_l.append(dict(buf=fru_d.get(key)))
            if key == state and 'ault' in fru_d.get(key, '').lower():
                col_l[len(col_l)-1].update(font=_error_font)

    # key_d contained all the known keys in FOS v9.1.1c. Below looks for anything new added
    key_l = list()
    for fru_d in fru_l:  # slot-number is skipped because it is the column header
        key_l.extend([str(k) for k in fru_d.keys() if not skip_d.get(k, False) and k not in key_d])
    for key in gen_util.remove_duplicates(key_l):
        row_l.append([dict(buf=key)] + [dict(buf=fru_d.get(key)) for fru_d in fru_l])

    return report_utils.add_content_defaults(row_l, default_d)


#########################################################
#                                                       #
#         Row insert functions for _contents            #
#                                                       #
#########################################################
def _firmware_version(chassis_obj, col_d):
    """Returns list of columns for the firmware version row.

    :param chassis_obj: Chassis object
    :type chassis_obj: brcddb.class.chassis.ChassisObj
    :param col_d: This is for the column definition dictionaries but since this function works on rows, its None.
    :type col_d: None
    :return: Row list of column list of dictionaries to insert while processing _contents
    :rtype: list
    """
    global _comment_span, _parameter_span, _setting_span, _common_default_d

    alert_obj = chassis_obj.r_alert_obj(al.ALERT_NUM.CHASSIS_FIRMWARE)

    return report_utils.add_content_defaults(
        [
            [
                dict(buf=None if alert_obj is None else alert_obj.fmt_msg(), span=_comment_span),
                dict(buf='Firmware Version', span=_parameter_span),
                dict(buf=brcddb_chassis.firmware_version(chassis_obj), span=_setting_span),
            ]
        ],
        _common_default_d
        )


def _logical_switches(chassis_obj, col_d):
    """Returns list of rows and list of columns for the logical switches. See _firmware_version() for parameters"""
    global _link_font, _num_columns

    rl = list()
    for switch_obj in chassis_obj.r_switch_objects():
        buf = brcddb_switch.best_switch_name(switch_obj, wwn=True, did=True, fid=True)
        rl.append([dict(
            buf=buf,
            font=_link_font,
            link=switch_obj.r_get('report_app/hyperlink/switch'),
            span=_num_columns)]
        )

    return rl


def _blades(chassis_obj, col_d):
    """Returns list of rows and list of columns for the blades in a chassis. See _firmware_version() for parameters"""
    global _blade_keys_d, _blade_skip_keys_d, _fru_default_d

    return _fru(
        chassis_obj,
        _blade_keys_d,
        brcdapi_util.fru_blade,
        'slot-number',
        'Slot',
        'blade-state',
        _fru_default_d,
        _blade_skip_keys_d
    )


def _fans(chassis_obj, col_d):
    """Returns list of rows and list of columns for the fans in the chassis. See _firmware_version() for parameters"""
    global _fan_keys_d, _fan_skip_keys_d, _fru_default_d

    return _fru(
        chassis_obj,
        _fan_keys_d,
        brcdapi_util.fru_fan,
        'unit-number',
        'Fan',
        'operational-state',
        _fru_default_d,
        _fan_skip_keys_d
    )


def _power_supplies(chassis_obj, col_d):
    """Returns list of rows and list of columns for the power supplies. See _firmware_version() for parameters"""
    global _ps_keys_d, _ps_skip_keys_d, _fru_default_d

    return _fru(
        chassis_obj,
        _ps_keys_d,
        brcdapi_util.fru_ps,
        'unit-number',
        'PS',
        'operational-state',
        _fru_default_d,
        _ps_skip_keys_d
    )


def _sensors(chassis_obj, col_d):
    """Returns list of rows and list of columns for the power supplies. See _firmware_version() for parameters"""
    global _sensor_keys_d, _sensor_skip_keys_d, _fru_default_d

    return _fru(
        chassis_obj,
        _sensor_keys_d,
        brcdapi_util.fru_sensor,
        'id',
        'Sensor',
        'state',
        _fru_default_d,
        _sensor_skip_keys_d
    )


def _wwn(chassis_obj, col_d):
    """Returns list of rows and list of columns for the power supplies. See _firmware_version() for parameters"""
    global _wwn_keys_d, _wwn_skip_keys_d, _fru_default_d

    return _fru(
        chassis_obj,
        _wwn_keys_d,
        brcdapi_util.fru_wwn,
        'unit-number',
        'WWN',
        'state',
        _fru_default_d,
        _wwn_skip_keys_d
    )


def _port_statistics(chassis_obj, col_d):
    """Returns the port statistics to insert in _contents. See _links() for parameter definitions."""
    global _detail_param_span, _port_speed

    return report_utils.port_statistics(chassis_obj, _detail_param_span, _port_speed)


def _conditional_highlight(chassis_obj, col_d):
    """Returns the conditional highlighting to insert in _contents. See _links() for parameter definitions."""
    global _conditional_c

    return report_utils.conditional_highlight(chassis_obj, _conditional_c)


_contents = [  # Search for **add_contents()** in brcddb.report.utils for an explanation
    list(),
    [dict(buf='Logical Switches', font=_bold_font, border=None, span=0)],
    _logical_switches,
    list(),
    [
        dict(buf='Comment', font=_bold_font, span=_comment_span),
        dict(buf='Parameter', font=_bold_font, span=_parameter_span),
        dict(buf='Setting', font=_bold_font, span=_setting_span),
    ],
    _firmware_version,
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_license_id)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_mfg)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_org_name)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_org_reg_date)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_pn)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_serial_num)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_support_sn)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_max_blades)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_vendor_pn)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_vendor_sn)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_vendor_rev_num)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_product_name)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_date)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_enabled)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_motd)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_shell_to)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_session_to)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_usb_enbled)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_usb_avail_space)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_tcp_to_level)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_bp_rev)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bp_vf_enabled)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_vf)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_ha)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_fcr_en)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_fcr_supported)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_heartbeat)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_sync)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_active_cp)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_active_slot)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_ha_recovery)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_ha_standby_cp)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_ha_standby_health)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_ha_standby_slot)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_rest_enabled)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_https_enabled)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_eff_protocol)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_max_rest)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_https_ka)],
    [dict(a=report_utils.cell_content, span=_cell_span_l, key=brcdapi_util.bc_https_ka_to)],
    list(),
    [dict(buf='Summary of Alerts', font=_bold_font, border=None, span=0)],
    report_utils.alert_summary,
    list(),
    [dict(buf='Blades', font=_bold_font, border=None, span=0)],
    _blades,
    list(),
    [dict(buf='Fans', font=_bold_font, border=None, span=0)],
    _fans,
    list(),
    [dict(buf='Power Supplies', font=_bold_font, border=None, span=0)],
    _power_supplies,
    list(),
    [dict(buf='Sensors', font=_bold_font, border=None, span=0)],
    _sensors,
    list(),
    [dict(buf='WWN Cards', font=_bold_font, border=None, span=0)],
    _wwn,
    list(),
    _port_statistics,
    list(),
    _conditional_highlight,
]

# Determine the maximum number of columns needed
_num_columns, _temp_span_l = 0, list()
for _temp_l in [_l for _l in _contents if isinstance(_l, list)]:
    _temp_columns = 0
    for _temp_d in [_d for _d in _temp_l if isinstance(_d, dict)]:
        _temp_span = _temp_d.get('span')
        if _temp_span is None:
            _temp_columns += 1
        elif isinstance(_temp_span, int):
            if _temp_span == 0:
                _temp_span_l.append(_temp_d)
            else:
                _temp_columns += _temp_span
        else:  # Assume it's a list or tuple
            for _x in _temp_span:
                _temp_columns += _x
    _num_columns = max(_num_columns, _temp_columns)
# Fill in the maximum columns
for _temp_d in _temp_span_l:
    _temp_d['span'] = _hdr_span

""" _hidden_hdr_d and _hidden_highlight_d are dictionaries of dictionaries used to determine the hidden chassis port
page. The key is the column header. The sub-dictionaries are as follows:

+-------+-----------+-----------------------------------------------------------------------------------------------+
| Key   | Type      | Description                                                                                   |
+=======+===========+===============================================================================================+
| w     | int       | Column width                                                                                  |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| v     | str       | Cell value. If the string begins with '=' it is assumed to e a formula. Formulas may contain  |
|       |           | the variables. Variables begin with $ and are substituted as follows:                         |
|       |           |   $alias   Aliases associated with this port.                                                 |          
|       |           |   $col     See notes with $sheet.                                                             |
|       |           |   $ealert  Number of error level alerts. Nothing else may be in v.                            |
|       |           |   $port    Port number in s/p notation. Nothing else may be in v.                             |
|       |           |   $ref     Followed by the column header in parenthesis. It is replaced by the excel col-row  |
|       |           |            reference. The column is the column for the header and the row is the active row   |
|       |           |            in the worksheet.                                                                  |
|       |           |   $row     See notes with $sheet.                                                             |
|       |           |   $sheet   Filled in by brcddb.report.switch.switch_page, or more likely a function called by |
|       |           |            switch_page(). This is the name of the switch page sheet. It is followed by $col   |
|       |           |            and $row which reference the associated cells in the "Highlight Filters" section.  |
|       |           |   $sw_ref  The switch and cell reference for the associated item in the "Highlight Filters"   |
|       |           |            section. This item is the key in sw_sheet_cond_format_d.                           |
|       |           |   $ch_ref  Same as $sw_ref, but for the chassis                                               |
|       |           |   $fab_ref Same as $sw_ref, but for the fabric                                                |               
|       |           |   $walert  Number of warning level alerts. Nothing else may be in v.                          |
|       |           |   $zone    Zones the port participates in.                                                    |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| c     | int       | Excel column number matching this header.                                                     |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| cl    | str       | Excel column letter                                                                           |
+-------+-----------+-----------------------------------------------------------------------------------------------+
"""
_hidden_hdr_d = collections.OrderedDict()
_hidden_hdr_d['Port'] = dict(w=10, v='$port')
_hidden_hdr_d['C3 Frame Discards'] = dict(w=10, v=brcdapi_util.stats_c3)
_hidden_hdr_d['Rx C3 Frames discarded'] = dict(w=10, v=brcdapi_util.stats_c3_in)
_hidden_hdr_d['Tx C3 Frame Discards'] = dict(w=10, v=brcdapi_util.stats_c3_out)
_hidden_hdr_d['Total C3 Discards'] = dict(w=10, v='=SUM($ref(C3 Frame Discards):$ref(Tx C3 Frame Discards))')
_hidden_hdr_d['Zero BB Credit'] = dict(w=10, v=brcdapi_util.stats_bb_credit)
_hidden_hdr_d['CRC Errors'] = dict(w=10, v=brcdapi_util.stats_crc)
_hidden_hdr_d['Incoming CRC Errors'] = dict(w=10, v=brcdapi_util.stats_in_crc)
_hidden_hdr_d['Total Bit Errors'] = dict(w=10, v='=SUM($ref(CRC Errors):$ref(Incoming CRC Errors))')
_hidden_hdr_d['Invalid Ordered Sets'] = dict(w=10, v=brcdapi_util.stats_ios)
_hidden_hdr_d['Bad EOF'] = dict(w=10, v=brcdapi_util.stats_bad_eof)
_hidden_hdr_d['Truncated Frames'] = dict(w=10, v=brcdapi_util.stats_tunc)
_hidden_hdr_d['Delimiter Errors'] = dict(w=10, v=brcdapi_util.stats_delimiter)
_hidden_hdr_d['Encoding Disparity Errors'] = dict(w=10, v=brcdapi_util.stats_enc_disp)
_hidden_hdr_d['Encoding Errors Outside Frame'] = dict(w=10, v=brcdapi_util.stats_enc)
_hidden_hdr_d['Frame Busy'] = dict(w=10, v=brcdapi_util.stats_f_busy)
_hidden_hdr_d['Frame Rejects'] = dict(w=10, v=brcdapi_util.stats_f_rjt)
_hidden_hdr_d['Frames Too Long'] = dict(w=10, v=brcdapi_util.stats_long)
_hidden_hdr_d['Total Frame Errors'] = dict(w=10, v='=SUM($ref(Invalid Ordered Sets):$ref(Frames Too Long))')
_hidden_hdr_d['Link Resets'] = dict(w=10, v=brcdapi_util.stats_in_reset)
_hidden_hdr_d['Link Failures'] = dict(w=10, v=brcdapi_util.stats_link_fail)
_hidden_hdr_d['Loss Of Signal'] = dict(w=10, v=brcdapi_util.stats_loss_sig)
_hidden_hdr_d['Loss Of Sync'] = dict(w=10, v=brcdapi_util.stats_loss_sync)
_hidden_hdr_d['Too Many RDYs'] = dict(w=10, v=brcdapi_util.stats_rdy)
_hidden_hdr_d['Warn Alerts'] = dict(w=10, v='$walert')
_hidden_hdr_d['Error Alerts'] = dict(w=10, v='$ealert')
_hidden_hdr_d['Zone'] = dict(w=28, v='$zones')
_hidden_hdr_d['Alias'] = dict(w=28, v='$aliases')
_hidden_hdr_d['Chassis Highlight Discarded Frames'] = dict(w=10, v='=IF($ref(Total C3 Discards)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight BB Credit Starvation'] = dict(w=10, v='=IF($ref(Zero BB Credit)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Bit Errors'] = dict(w=10, v='=IF($ref(Total Bit Errors)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Warn Alerts'] = dict(w=10, v='=IF($ref(Warn Alerts)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Error Alerts'] = dict(w=10, v='=IF($ref(Error Alerts)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Frame Errors'] = dict(w=10, v='=IF($ref(Total Frame Errors)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Link Resets'] = dict(w=10, v='=IF($ref(Link Resets)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Link Failures'] = dict(w=10, v='=IF($ref(Link Failures)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Loss Of Signal'] = dict(w=10, v='=IF($ref(Loss Of Signal)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Loss Of Sync'] = dict(w=10, v='=IF($ref(Loss Of Sync)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Too Many RDYs'] = dict(w=10, v='=IF($ref(Too Many RDYs)>$ch_ref,1,0)')
_hidden_hdr_d['Chassis Highlight Zone EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Zone))>0,IF(EXACT($ref(Zone),$ch_ref),1,0),0)'))
_hidden_hdr_d['Chassis Highlight Zone FIND'] = (
     dict(w=10, v='=IF(AND(LEN($ch_ref)>0,LEN($ref(Zone))>0),IFERROR(FIND($ch_ref,$ref(Zone)),0),0)'))
_hidden_hdr_d['Chassis Highlight Alias EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Alias))>0,IF(EXACT($ref(Alias),$ch_ref),1,0),0)'))
_hidden_hdr_d['Chassis Highlight Alias FIND'] = (
    dict(w=10, v='=IF(AND(LEN($ch_ref)>0,LEN($ref(Alias))>0),IFERROR(FIND($ch_ref,$ref(Alias)),0),0)')
)
_hidden_hdr_d['Chassis Highlight Total'] = dict(w=10, v='n/a')
_hidden_hdr_d['Switch Highlight Discarded Frames'] = dict(w=10, v='=IF($ref(Total C3 Discards)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight BB Credit Starvation'] = dict(w=10, v='=IF($ref(Zero BB Credit)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Bit Errors'] = dict(w=10, v='=IF($ref(Total Bit Errors)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Warn Alerts'] = dict(w=10, v='=IF($ref(Warn Alerts)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Error Alerts'] = dict(w=10, v='=IF($ref(Error Alerts)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Frame Errors'] = dict(w=10, v='=IF($ref(Total Frame Errors)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Link Resets'] = dict(w=10, v='=IF($ref(Link Resets)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Link Failures'] = dict(w=10, v='=IF($ref(Link Failures)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Loss Of Signal'] = dict(w=10, v='=IF($ref(Loss Of Signal)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Loss Of Sync'] = dict(w=10, v='=IF($ref(Loss Of Sync)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Too Many RDYs'] = dict(w=10, v='=IF($ref(Too Many RDYs)>$sw_ref,1,0)')
_hidden_hdr_d['Switch Highlight Zone EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Zone))>0,IF(EXACT($ref(Zone),$sw_ref),1,0),0)'))
_hidden_hdr_d['Switch Highlight Zone FIND'] = (
    dict(w=10, v='=IF(AND(LEN($sw_ref)>0,LEN($ref(Zone))>0),IFERROR(FIND($sw_ref,$ref(Zone)),0),0)'))
_hidden_hdr_d['Switch Highlight Alias EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Alias))>0,IF(EXACT($ref(Alias),$sw_ref),1,0),0)'))
_hidden_hdr_d['Switch Highlight Alias FIND'] = (
    dict(w=10, v='=IF(AND(LEN($sw_ref)>0,LEN($ref(Alias))>0),IFERROR(FIND($sw_ref,$ref(Alias)),0),0)'))
_hidden_hdr_d['Switch Highlight Total'] = (
    dict(w=10, v='n/a'))
_hidden_hdr_d['Fabric Highlight Discarded Frames'] = dict(w=10, v='=IF($ref(Total C3 Discards)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight BB Credit Starvation'] = dict(w=10, v='=IF($ref(Zero BB Credit)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Bit Errors'] = dict(w=10, v='=IF($ref(Total Bit Errors)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Warn Alerts'] = dict(w=10, v='=IF($ref(Warn Alerts)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Error Alerts'] = dict(w=10, v='=IF($ref(Error Alerts)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Frame Errors'] = dict(w=10, v='=IF($ref(Total Frame Errors)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Link Resets'] = dict(w=10, v='=IF($ref(Link Resets)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Link Failures'] = dict(w=10, v='=IF($ref(Link Failures)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Loss Of Signal'] = dict(w=10, v='=IF($ref(Loss Of Signal)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Loss Of Sync'] = dict(w=10, v='=IF($ref(Loss Of Sync)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Too Many RDYs'] = dict(w=10, v='=IF($ref(Too Many RDYs)>$fab_ref,1,0)')
_hidden_hdr_d['Fabric Highlight Zone EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Zone))>0,IF(EXACT($ref(Zone),$fab_ref),1,0),0)'))
_hidden_hdr_d['Fabric Highlight Zone FIND'] = (
    dict(w=10, v='=IF(AND(LEN($fab_ref)>0,LEN($ref(Zone))>0),IFERROR(FIND($fab_ref,$ref(Zone)),0),0)'))
_hidden_hdr_d['Fabric Highlight Alias EXACT'] = (
    dict(w=10, v='=IF(LEN($ref(Alias))>0,IF(EXACT($ref(Alias),$fab_ref),1,0),0)'))
_hidden_hdr_d['Fabric Highlight Alias FIND'] = (
    dict(w=10, v='=IF(AND(LEN($fab_ref)>0,LEN($ref(Alias))>0),IFERROR(FIND($fab_ref,$ref(Alias)),0),0)'))
_hidden_hdr_d['Fabric Highlight Total'] = (
    dict(w=10, v='n/a'))

# Add the Excel column number to _hidden_hdr_d and _hidden_highlight_d
_col = 1
for _d in _hidden_hdr_d.values():
    _d.update(c=_col, cl=xl.get_column_letter(_col))
    _col += 1


def add_port_map(chassis_obj):
    """Adds the port map to chassis page

    :param chassis_obj: Chassis object
    :type chassis_obj: brcddb.classes.chassis.ChassisObj
    :rtype: None
    """
    global _port_col_span, _fixed_port_map_l, _port_legend, _fixed_ge_port_col_span, _bladed_ge_port_col_span
    global _fill_lightblue

    # Get all the ports on a per-blade basis.
    slot_port_d = collections.OrderedDict()  # The ports are sorted, so this will be in slot order
    for key, port_l in dict(
            port_l=brcdapi_port.sort_ports(chassis_obj.r_port_keys()),
            ge_port_l=brcdapi_port.sort_ports(chassis_obj.r_ge_port_keys()),
    ).items():
        for port in port_l:
            port_split_l = port.split('/')
            port_d = slot_port_d.get(port_split_l[0])
            if port_d is None:
                port_d = dict(port_l=list(), ge_port_l=list())
                slot_port_d.update({port_split_l[0]: port_d})
            port_d[key].append(port_split_l[1])

    # Sort and find the maximum number of ports per blade
    max_ports = max_ge_ports = 0
    slot_keys_l = [str(k) for k in slot_port_d.keys()]
    for slot in slot_port_d.keys():
        max_ports = max(max_ports, len(slot_port_d[slot]['port_l']))
        max_ge_ports = max(max_ge_ports, len(slot_port_d[slot]['ge_port_l']))

    # Add the lists and dictionaries with the data for the worksheet to pass to add_contents().
    rl = list([dict()])
    if len(slot_keys_l) == 0:  # This happens when a partial data collection is taken
        span = chassis_obj.r_get('report_app/worksheet/num_columns', _port_legend_span)
        rl.append([dict(buf='No ports found', span=span)])
        rl.append([dict()])

    elif slot_keys_l[0] == '0':  # It's a fixed port chassis
        try:
            for row_l in _fixed_port_map_l:
                col_l = list()
                rl.append(col_l)
                for port in row_l:
                    port_obj = chassis_obj.r_port_obj('0/' + str(port))
                    if port_obj is None:
                        # Add the GigE ports
                        for port_obj in chassis_obj.r_ge_port_objects():
                            col_l.append(_insert_port(chassis_obj, port_obj, _port_col_span))
                        raise Found  # There are no ports in the chassis, so stop here.
                    col_l.append(_insert_port(chassis_obj, port_obj, _port_col_span))
        except Found:
            pass

    else:  # It's a bladded chassis.
        # max_ports will always be 64 because that's how many ports are on a core blade and the highest port count for
        # a port blade is 64. I'm using the calculated value for max_ports in case that ever changes.
        port_map_d, col_l = dict(), list()

        # The slot headers & set up the port map, port_map_d.
        for slot, port_d in slot_port_d.items():

            # The slot headers
            col_l.append(dict(font=None, border=None, align=None))
            col_l.append(dict(
                buf='Slot ' + slot,
                font=_bold_font,
                align=_align_wrap_c,
                fill=_fill_lightblue,
                span=_port_col_span * 2
            ))

            # The port map. I'm counting on the fact that there will always be an even number of ports per column. I add
            # the GE ports to col_0_l after adding all the ports so that when the corresponding value at the same index
            # in col_1_l is None, I know it's a GE port. This is significant because ports are displayed in two columns
            # on port cards while GE ports are displayed as a single entry and therefore spans both columns
            x = int(len(port_d['port_l'])/2)
            port_d.update(col_0_l=port_d['port_l'][0:x], col_1_l=port_d['port_l'][x:])
            port_d['col_0_l'].reverse()
            for port in port_d['ge_port_l']:
                port_d['col_0_l'].insert(0, port)
            port_d['col_1_l'].reverse()
            for temp_l in [port_d['col_0_l'], port_d['col_1_l']]:
                for x in range(len(temp_l), int(max_ports/2)):
                    temp_l.insert(0, None)
        rl.append(col_l)

        # Add the ports to the return list
        row_l = list()
        for x in range(0, int(max_ports/2)):
            col_l = list()
            row_l.append(col_l)
            for slot, port_d in slot_port_d.items():
                col_l.append(dict(font=None, border=None, align=None))
                if port_d['col_1_l'][x] is None:
                    if port_d['col_0_l'][x] is None:
                        col_l.append(dict(span=_port_col_span))
                        col_l.append(dict(span=_port_col_span))
                    else:
                        col_l.append(_insert_port(
                            chassis_obj,
                            chassis_obj.r_ge_port_obj(slot + '/' + port_d['col_0_l'][x]),
                            _port_col_span * 2
                        ))
                else:
                    for port in (port_d['col_0_l'][x], port_d['col_1_l'][x]):
                        col_l.append(_insert_port(
                            chassis_obj,
                            chassis_obj.r_port_obj(slot + '/' + port),
                            _port_col_span
                        ))
        rl.extend(row_l)
        rl.append(list())

    # Add the defaults, the legend, and add it to report_app in the chassis object
    report_utils.add_content_defaults(rl, _port_map_default_d)
    rl.extend(_port_legend)
    chassis_obj.rs_key('report_app/port_map_l', rl)
    report_utils.add_contents(chassis_obj, rl)


def chassis_page(wb, tc, sheet_name, sheet_i, chassis_obj):
    """Creates a chassis detail worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: openpyxl.workbook.workbook.Workbook
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Default is 0
    :type sheet_i: int, None
    :param chassis_obj: Chassis object
    :type chassis_obj: brcddb.classes.chassis.ChassisObj
    :rtype: None
    """
    global _contents, _num_columns, _common_default_d, _link_font, _align_wrap_c, _hdr1_font

    # Create the worksheet, set the column widths, and add the report control structure to the chassis object
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.add_data_validation(report_utils.highlight_stats_dv)
    for col in range(0, _num_columns + _extra_columns):
        sheet.column_dimensions[xl.get_column_letter(col+1)].width = _common_width
    sheet_d = dict(sheet=sheet, num_columns=_num_columns, row=1, cond_format_d=dict())
    chassis_obj.rs_key('report_app/worksheet', sheet_d)

    # Add the link to the table of contents and title
    col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, sheet_d['row'], col, 'Contents', font=_link_font, link=tc)
        sheet.merge_cells(start_row=sheet_d['row'], start_column=col, end_row=sheet_d['row'], end_column=col+3)
        col += 4
    excel_util.cell_update(
        sheet,
        sheet_d['row'],
        col,
        brcddb_chassis.best_chassis_name(chassis_obj, wwn=True),
        font=_hdr1_font,
        align=_align_wrap_c
    )
    sheet.merge_cells(
        start_row=sheet_d['row'],
        start_column=col,
        end_row=sheet_d['row'],
        end_column=_hdr_span-col
    )
    sheet_d['row'] += 1
    sheet.freeze_panes = sheet['A' + str(sheet_d['row'])]

    # Add the contents
    report_utils.add_contents(chassis_obj, report_utils.add_content_defaults(_contents, _common_default_d))


def _solve_ref(in_buf, row):
    """Used in chassis_hidden_port_page to resolve $ref

    :param in_buf: Text to search for $ref and resolve reference
    :type in_buf: str
    :param row: Current Excel row
    :type row: int
    :return: in_buf with $ref converted to a cell identifier
    :rtype: str
    """
    global _hidden_hdr_d

    while '$ref' in in_buf:
        begin_i = in_buf.index('$ref')
        w_buf = in_buf[begin_i:]
        hdr_key, w_buf = gen_util.paren_content(w_buf[len('$ref'):], p_remove=True)
        in_buf = in_buf[0:begin_i] + _hidden_hdr_d[hdr_key]['cl'] + str(row) + w_buf
    return in_buf


def chassis_hidden_port_page(wb, sheet_name, sheet_i, chassis_obj):
    """Adds the hidden chassis port page which is used for port conditional highlighting

    :param wb: Workbook object
    :type wb: openpyxl.workbook.workbook.Workbook
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Location for this sheet.
    :type sheet_i: int
    :param chassis_obj: Chassis object
    :type chassis_obj: brcddb.classes.chassis.ChassisObj
    :rtype: None
    """
    global _hidden_hdr_d, _align_wrap_c, _border_thin

    sheet_name = chassis_obj.r_get('report_app/control/chassis_p')['sn']
    brcdapi_log.log('    Adding ' + sheet_name, echo=True)
    cond_format_d = chassis_obj.rs_key('report_app/worksheet/cond_format_d', dict())

    # Create the worksheet, set up the column widths, and add the headers
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.sheet_state = 'hidden'
    row = 1
    for buf, d in _hidden_hdr_d.items():
        excel_util.cell_update(sheet, row, d['c'], buf, font=_bold_font, align=_align_wrap_c)
        sheet.column_dimensions[d['cl']].width = d['w']
    row += 1

    # Add all the ports. They are sorted for readability.
    for port in brcdapi_port.sort_ports(chassis_obj.r_port_keys()):
        port_obj = chassis_obj.r_port_obj(port)

        # Resolve the remaining reference values
        w_alert = e_alert = 0
        zone_l = [obj.r_obj_key() for obj in obj_convert.obj_extract(port_obj, 'ZoneObj')]
        zone = zone_l.pop(0) if len(zone_l) > 0 else ''
        alias_l = [obj.r_obj_key() for obj in obj_convert.obj_extract(port_obj, 'AliasObj')]
        alias = alias_l.pop(0) if len(alias_l) > 0 else ''

        # Figure out how many warn and error level alerts are associated with the port.
        for alert_obj in port_obj.r_alert_objects():
            alert_sev = alert_obj.sev()
            if alert_sev == alert_class.ALERT_SEV.WARN:
                w_alert += 1
            elif alert_sev == alert_class.ALERT_SEV.ERROR:
                e_alert += 1

        # Fill in the cell value for each column
        for key, d in _hidden_hdr_d.items():
            ref_value = cond_format_d.get(key.replace('Chassis Highlight ', ''))
            v = d['v'] if ref_value is None else d['v'].replace('$ch_ref', ref_value)
            if '$walert' in v:
                v = w_alert
            elif '$ealert' in v:
                v = e_alert
            elif '$port' in v:
                v = port
            elif '$zones' in v or '$aliases' in v:
                v = d['v'].replace('$zones', zone).replace('$aliases', alias)
            else:
                # Resolve formula references for $ref (current page col-row).
                v = _solve_ref(v, row) if v[0] == '=' else port_obj.r_get(d['v'])

            # Add the value to the cell
            if isinstance(v, str):
                if len(v) == 0 or '$fab_ref' in v or '$sw_ref' in v:  # Fabric and switch highlighting not implemented
                    v = None  # There isn't anything for the cell.
                elif '$' in v:
                    brcdapi_log.exception('Unresolved reference in v: ' + str(v), echo=True)
                    v = None
            excel_util.cell_update(sheet, row, d['c'], v, border=_border_thin)

        # Add the remaining aliases and zones
        # Programmers Note: I forgot to add the formulas for the conditional highlighting cells, see k_l below. In
        # retrospect, I would have broken out the loop "for key, d in _hidden_hdr_d.items():". I would have done it that
        # way so that I could use it to insert rows and columns, which is the same as how I handle all other content.
        # That also means the updates to the spreadsheet would have to be broken out. As with any programmer, I took a
        # few ill-advised short-cuts. To patch in this fix without re-writing what was working, I took advantage of the
        # fact that I know the only reference variable, text beginning with '$', in _hidden_hdr_d for the keys in k_l is
        # $ch_ref. So rather than re-write the code to resolve references other than $ch_ref, since I know no other
        # references need to be resolved, I took the cheesy way out.

        max_row = row
        for d in [
            dict(
                l=alias_l,
                h_key='Alias',
                k_d={  # Key is the column header. Value is the $ch_ref
                    'Chassis Highlight Alias EXACT': cond_format_d['Alias EXACT'],
                    'Chassis Highlight Alias FIND': cond_format_d['Alias FIND'],
                }
            ),
            dict(
                l=zone_l,
                h_key='Zone',
                k_d={  # Same notes as with h_key='Alias'
                    'Chassis Highlight Zone EXACT': cond_format_d['Zone EXACT'],
                    'Chassis Highlight Zone FIND': cond_format_d['Zone FIND'],
                }
            ),
        ]:
            # These are the remaining aliases and zones
            temp_row = row + 1
            for buf in d['l']:
                excel_util.cell_update(sheet, temp_row, _hidden_hdr_d[d['h_key']]['c'], buf)  # The zone or alias
                for key, ref in d['k_d'].items():  # The zones and aliases that need additional rows
                    excel_util.cell_update(
                        sheet,
                        temp_row,
                        _hidden_hdr_d[key]['c'],
                        # Below only works for the chassis. Need to re-think this when I handle switches & fabrics
                        _solve_ref(_hidden_hdr_d[key]['v'].replace('$ch_ref', ref), temp_row),
                        border=_border_thin)
                temp_row += 1
        max_row = max(max_row, max_row + max(len(alias_l), len(zone_l)))

        # Resolve the final "Highlight Total" cells. ToDo - Add switch and fabric
        for d in (dict(ref='Chassis', link='ch'),):
            buf = '=SUM(' + \
                  _hidden_hdr_d[d['ref'] + ' Highlight Discarded Frames']['cl'] + str(row) + ':' + \
                  _hidden_hdr_d[d['ref'] + ' Highlight Alias FIND']['cl'] + str(max_row) + ')'
            col = _hidden_hdr_d[d['ref'] + ' Highlight Total']['c']
            excel_util.cell_update(sheet, max_row, col, buf)
            port_obj.rs_key(
                'report_app/hyperlink/' + d['link'] + '_highlight/total',
                sheet_name + '!' + xl.get_column_letter(col) + str(max_row)
            )
        row = max_row + 1

    return
