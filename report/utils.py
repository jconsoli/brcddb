"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Common functions used for reading and generating Excel reports.

**add_contents()**

add_contents() adds cell content and formatting to Excel sheets. The content list, contents_l, passed to add_contents()
is a list of rows with either:

    * A function pointer. This allows you to use a function to insert rows.
    * A list of dictionaries describing the cell content and formatting for each column. See "Column Dictionaries" and
      "Extended Column Dictionaries" below for details.

Column Dictionaries:

These key/value pairs are used by add_contents() to determine what and how to add cell contents.

+-----------+-----------------------+-------------------------------------------------------------------------------+
| Key       | Type                  | Description                                                                   |
+===========+=======================+===============================================================================+
| a         | Function pointer      | Allows you to insert a column data from a function rather than a dictionary   |
|           |                       | or hard coded values.                                                         |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| align     | xl_styles.Alignment,  | One of the align types from brcdapi.excel_util. The default is _align_wrap.   |
|           | None                  | If explicitly set to None, the alignment for the cell is not set. In that     |
|           |                       | case, the alignment for the cell will be the default for the Excel workbook.  |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| border    | xl_styles.Border,     | One of the border types from brcdapi.excel_util. The default is _border_thin. |
|           | None                  | See notes with align regarding a None value.                                  |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| buf       | See description       | Determines what to print in the cell as follows:                              |
|           |                       |   Function pointer    The function is called. The returned value is used as   |
|           |                       |                       cell content.                                           |
|           |                       |   int, float, str     Unless converted to a string, see note at bottom, these |
|           |                       |                       values are used directly as cell content.               |
|           |                       |   dict                The key is used to retrieve the cell value from the     |
|           |                       |                       dictionary.                                             |
|           |                       |   Object type         In this module, the object type is a switch object. The |
|           |                       |                       key is used to retrieve the cell value from the object. |
|           |                       |   list                May be a list of the above. If the length of the list   |
|           |                       |                       is greater than 1, after all items are converted to     |
|           |                       |                       strings and concatenated. WARNING: Supporting a list    |
|           |                       |                       was planned, but never tested. I'm not sure if I even   |
|           |                       |                       finished coding it.                                     |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| comments  | str                   | Cell comments                                                                 |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| dv        | DataValidation        | Data validation object from openpyxl.worksheet.datavalidation.DataValidation  |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| fill      | xl_styles.PatternFill | One of the fill types from brcdapi.excel_util.                                |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| font      | xl_styles.Font,       | One of the font types from brcdapi.excel_util. The default is _std_font       |
|           | None                  | See notes with align regarding a None value.                                  |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| key       | str, key              | Key to use when the buf item is a dictionary or a brcddb class object. It is  |
|           |                       | also passed to functions when buf is a function pointer.                      |
+-----------+-----------------------+-------------------------------------------------------------------------------+
| span      | int, list, tuple      | Number of columns to span (merge). If 0, fill in with the maximum column.     |
|           |                       | Using a list or tuple only makes sense when the function in 'a' knows what to |
|           |                       | do with it. At the time this was written, brcddb.report.utils.cell_content()  |
|           |                       | was the only function that worked with a list of tuple.                       |
+-----------+-----------------------+-------------------------------------------------------------------------------+

Extended Column Dictionaries:

These key/value pairs are ignored by add_contents(). They are used to customize pre-determined content. At the time this
was written, using this data was limited to the port map created when the chassis page is created. The switch page
displays this same port map except ports not in the logical switch are greyed out, the highlighting links are for the
switch instead of the chassis, and highlighting operations cannot be applied to ports not in that logical switch. Some
data was also added for fabrics with future plans to add port highlighting to the fabric port list.

+-------------------+-----------+-----------------------------------------------------------------------------------+
| Key               | Type      | Description                                                                       |
+===================+===========+===================================================================================+
| ch_cond_link      | str       | The reference to the "Chassis Highlight Total" cell.                              |
+-------------------+-----------+-----------------------------------------------------------------------------------+
| sw_cond_link      | str       | The reference to the "Switch Highlight Total" cell.                               |
+-------------------+-----------+-----------------------------------------------------------------------------------+
| fab_cond_link     | str       | The reference to the "Fabric Highlight Total" cell.                               |
+-------------------+-----------+-----------------------------------------------------------------------------------+
| port_obj          | PortObj   | Port object for the port.                                                         |
+-------------------+-----------+-----------------------------------------------------------------------------------+

**Public Functions**

+-------------------------------+-----------------------------------------------------------------------------------+
| Method                        | Description                                                                       |
+===============================+===================================================================================+
| add_content_defaults          | Adds the default values for font, align, and border to a list of contents for     |
|                               | add_contents().                                                                   |
+-------------------------------+-----------------------------------------------------------------------------------+
| add_contents                  | Adds contents to the worksheet                                                    |
+-------------------------------+-----------------------------------------------------------------------------------+
| alert_eval                    | Determines if there are alerts associated with an object.                         |
+-------------------------------+-----------------------------------------------------------------------------------+
| cell_content                  | Builds a list of column dictionaries for rows that are Comment, Parameter, Value  |
+-------------------------------+-----------------------------------------------------------------------------------+
| combined_alerts               | Converts alerts associated with a port object and the login and FDMI objects      |
|                               | for lwwn to a human-readable string.                                              |
+-------------------------------+-----------------------------------------------------------------------------------+
| combined_alert_objects        | Combines alerts associated with a port object and the login and FDMI objects      |
|                               | for wwn.                                                                          |
+-------------------------------+-----------------------------------------------------------------------------------+
| comments_for_alerts           | Converts alerts associated with an object to a human-readable string.             |
|                               | Multiple comments separated with /n                                               |
+-------------------------------+-----------------------------------------------------------------------------------+
| combined_login_alerts         | Converts alerts associated with a login object and the login and FDMI objects     |
|                               | for lwwn to a human-readable string                                               |
+-------------------------------+-----------------------------------------------------------------------------------+
| combined_login_alert_objects  | Combines login alert objects with FDMI HBA and FDMI port alerts objects           |
+-------------------------------+-----------------------------------------------------------------------------------+
| conditional_highlight         | Adds the port highlighting filters for a worksheet.                               |
+-------------------------------+-----------------------------------------------------------------------------------+
| fabric_name_case              | Return the fabric name. Typically used in case statements for page reports in     |
|                               | brcddb.report.                                                                    |
+-------------------------------+-----------------------------------------------------------------------------------+
| fabric_name_or_wwn_case       | Return the fabric name with (wwn). Typically used in case statements for page     |
|                               | reports in brcddb.report.*                                                        |
+-------------------------------+-----------------------------------------------------------------------------------+
| fabric_wwn_case               | Return the fabric WWN. Typically used in case statements for page reports in      |
|                               | brcddb.report.                                                                    |
+-------------------------------+-----------------------------------------------------------------------------------+
| font_type                     | Determines the display font based on alerts.                                      |
+-------------------------------+-----------------------------------------------------------------------------------+
| font_type_for_key             | Determines the display font based on alerts associated with an object.            |
+-------------------------------+-----------------------------------------------------------------------------------+
| get_next_switch_d             | Finds the first match in sl list returned from excel_util.read_sheet() and        |
|                               | returns the next entry                                                            |
+-------------------------------+-----------------------------------------------------------------------------------+
| list_to_str                   | Converts a list to a \n seperated list. Returns the column list.                  |
+-------------------------------+-----------------------------------------------------------------------------------+
| alert_summary                 | Inserts MAPS dashboard rows.                                                      |
+-------------------------------+-----------------------------------------------------------------------------------+
| parse_sfp_file                | Parses Excel file with the new SFP rules. See sfp_rules_rx.xlsx                   |
+-------------------------------+-----------------------------------------------------------------------------------+
| parse_switch_file             | Parses Excel switch configuration Workbook.                                       |
+-------------------------------+-----------------------------------------------------------------------------------+
| port_map                      | Adds the port map.                                                                |
+-------------------------------+-----------------------------------------------------------------------------------+
| port_statistics               | Adds port statistics for the worksheet.                                           |
+-------------------------------+-----------------------------------------------------------------------------------+
| seconds_to_days               | Converts a value in seconds to days. Returns the comments, parameter, and value   |
|                               | in a list for column insertion.                                                   |
+-------------------------------+-----------------------------------------------------------------------------------+
| title_page                    | Creates a title page for the Excel report.                                        |
+-------------------------------+-----------------------------------------------------------------------------------+

**Public Data**

+-----------------------+-------+-----------------------------------------------------------------------------------+
| Data                  | Type  | Description                                                                       |
+=======================+=======+===================================================================================+
| default_stats         | int   | Default value for the highlighting section of port stats. This is also the        |
|                       |       | maximum allowable value in highlight_stats_dv.                                    |
+-----------------------+-------+-----------------------------------------------------------------------------------+
| highlight_stats_dv    | class | Data validation class: openpyxl.worksheet.datavalidation.DataValidation           |
+-----------------------+-------+-----------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 20 Oct 2024   | Added functions to support the combined switch and chassis formats.                   |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 06 Dec 2024   | Fixed bugs: FC address on ICL ports for FICON switches, allowing XISL on base switch, |
|           |               | and setting the default routing policy.                                               |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 26 Dec 2024   | Fixed grammar mistake in help message.                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 04 Jan 2025   | Added Slot_x_Terms is list of sheets to skip reading.                                 |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.6     | 12 Apr 2025   | FOS 9.2 updates.                                                                      |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.7     | 25 Aug 2025   | Made _alert_font_d public (changed to alert_font_d).                                  |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024, 2025 Consoli Solutions, LLC'
__date__ = '25 Aug 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.7'

import collections
import copy
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
import brcdapi.util as brcdapi_util
import openpyxl.utils.cell as xl_util
import fnmatch
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.port as brcdapi_port
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_login as brcddb_login
import brcddb.util.search as brcddb_search
import brcddb.app_data.alert_tables as al_table
import brcddb.app_data.report_tables as rt
import brcddb.classes.alert as alert_class
import brcddb.classes.util as brcddb_class_util
import brcddb.brcddb_chassis as brcddb_chassis
import brcddb.brcddb_switch as brcddb_switch

default_stats = 999999999
highlight_stats_dv = DataValidation(type='whole',
                                    operator='between',
                                    formula1=0,
                                    formula2=default_stats,
                                    allow_blank=False)
highlight_stats_dv.errorTitle = 'Invalid Entry'
highlight_stats_dv.error = 'Value must be an integer between 0 and ' + str(default_stats)
highlight_stats_dv.showErrorMessage = True

_MAX_FICON_PORT_NAME = 24
_apt_policy_d = dict(EBR='exchange-based', DBR='device-based')

# Below is for efficiency
_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_warn_font = excel_fonts.font_type('warn')
_error_font = excel_fonts.font_type('error')
_gray_font = excel_fonts.font_type('gray')
_link_font = excel_fonts.font_type('link')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_r = excel_fonts.align_type('wrap_right')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')
_lightgreen_fill = excel_fonts.fill_type('lightgreen')
_lightgold_fill = excel_fonts.fill_type('lightgold')
_lightred_fill = excel_fonts.fill_type('lightred')
_lightblue_fill = excel_fonts.fill_type('lightblue')
_yellow_fill = excel_fonts.fill_type('yellow')

# _obj_type_to_key_d is used in add_contents() to determine what "xxx_Highlight Total" to use for the object.
_obj_type_to_key_d = dict(ChassisObj='ch_cond_link', SwitchObj='sw_cond_link', FabricObj='fab_cond_link')

# About worksheet - used in function: about
_about_sheet_l = (
    dict(f=_hdr1_font, t='Generated By'),
    None,
    dict(f=_std_font, t='Script:  ', s='script_name'),
    dict(f=_std_font, t='Version: ', s='version'),
    None,
    dict(f=_hdr1_font, t='Description'),
    None,
    dict(f=_std_font, t='', s='description'),
    None,
    dict(f=_hdr1_font, t='Automation & Scripting Tools'),
    None,
    dict(f=_std_font,
         t='The script used to generate this workbook and other scripts for use with the Brocade FOS RESTConf API can '
           'be found at:'),
    dict(f=_link_font, l='https://github.com/jconsoli', t='https://github.com/jconsoli'),
    None,
    dict(f=_std_font,
         t='YouTube videos to assist with getting started and using some of the sample scripts are available at:'),
    dict(f=_link_font,
         l='https://youtu.be/BWz7L0QOtYQ',
         t='Getting Started: Install Python and configure Brocade switches for use with the API'),
    dict(f=_link_font, l='https://youtu.be/n9-Eni_AFCg', t='Capture Data & Generate Reports'),
    dict(f=_link_font, l='https://youtu.be/x1OvuRZRdA8', t='FOS API Zoning'),
    dict(f=_link_font, l='https://youtu.be/WGxXZrvhG2E', t='Chassis, Switch, Fabric, and Port Configuration'),
    None,
    dict(f=_std_font, t='email: jack@consoli-solutions.com'),
    None,
    dict(f=_hdr1_font, t='Disclaimer'),
    None,
    dict(f=_std_font,
         t='The script uses open source software covered by the GNU General Public License or other open source '
           'license agreements.'),
    None,
    dict(f=_std_font,
         t='The accuracy and completness of information contained herein are dependant of the FOS version, data '
           'collection method, and other factors. Consoli-Solutions does not assume any liability arising out of the '
           'application or use of this information, nor the application or use of any product or software described '
           'herein, neither does it convey any license under its patent rights nor the rights of others.'),
)


#########################################################################
#                                                                       #
# Common case statements in port_page(), switch_page(), chassis_page()  #
#                                                                       #
#########################################################################
def fabric_name_case(obj, k=None, wwn=None):
    """Return the fabric name. Typically used in case statements for page reports in brcddb.report.*

    :param obj: brcddb object that contains a fabric object
    :type obj: brcddb.classes.fabric.FabricObj, brcddb.classes.fabric.SwitchObj, brcddb.classes.fabric.PortObj,
               brcddb.classes.fabric.LoginObj
    :param k: Not used
    :type k: str
    :param wwn: Not used
    :type wwn: str
    """
    return brcddb_fabric.best_fab_name(obj.r_fabric_obj(), False)


def fabric_name_or_wwn_case(obj, k=None, wwn=None):
    """Return the fabric name with (wwn). Typically used in case statements for page reports in brcddb.report.*

    See fabric_name_case() for parameter definitions"""
    return brcddb_fabric.best_fab_name(obj.r_fabric_obj(), True)


def fabric_wwn_case(obj, k=None, wwn=None):
    """Return the fabric WWN. Typically used in case statements for page reports in brcddb.report.*

    See fabric_name_case() for parameter definitions"""
    return obj.r_fabric_key()


def font_type(obj_list):
    """Determines the display font based on alerts.

    :param obj_list: List of brcddb.classes.alert.AlertObj
    :param obj_list: list
    """
    font = excel_fonts.font_type('std')
    for alert_obj in obj_list:
        if alert_obj.is_error():
            return excel_fonts.font_type('error')
        elif alert_obj.is_warn():
            font = excel_fonts.font_type('warn')

    return font


def font_type_for_key(obj, k=None):
    """Determines the display font based on alerts associated with an object.

    :param obj: Any brcddb object
    :param obj: ChassisObj, FabricObj, SwitchObj, PortObj, ZoneObj, ZoneCfgObj, AliasObj
    :param k: Key to associate the alert with
    :type k: str
    """
    font = excel_fonts.font_type('std')
    a_list = obj.r_alert_objects() if k is None else [a_obj for a_obj in obj.r_alert_objects() if a_obj.key() == k]
    for a_obj in a_list:
        if a_obj.is_error():
            return excel_fonts.font_type('error')
        elif a_obj.is_warn():
            font = excel_fonts.font_type('warn')

    return font


def comments_for_alerts(gobj, k=None, wwn=None):
    """Converts alerts associated with an object to a human-readable string. Multiple comments separated with /n

    :param gobj: Any brcddb object
    :type gobj: ChassisObj, FabricObj, SwitchObj, PortObj, ZoneObj, ZoneCfgObj, AliasObj
    :param k: Key to associate the alert with
    :type k: str
    :param wwn: If nont None, the object is assumed to be a portObj and the comments returned are for the login
    :type wwn: None, str
    :return: /n seperated list of alert message(s) associated with obj
    :rtype: str
    """
    obj = gobj if wwn is None else gobj.r_fabric_obj().r_login_obj(wwn)
    if obj is None:
        return ''
    a_list = obj.r_alert_objects() if k is None else [a_obj for a_obj in obj.r_alert_objects() if a_obj.key() == k]
    return '\n'.join([a_obja_obj.fmt_msg() for a_obja_obj in a_list])


def combined_login_alert_objects(login_obj):
    """Combines login alert objects with FDMI HBA and FDMI port alerts objects

    :param login_obj: Login object
    :type : brcddb.classes.login.LoginObj
    :return: List of AlertObj
    :rtype: list
    """
    a_list = list()
    if login_obj is not None:
        a_list.extend(login_obj.r_alert_objects())  # Alerts tied to the login
        wwn = login_obj.r_obj_key()
        fab_obj = login_obj.r_fabric_obj()
        if fab_obj is not None:
            obj = fab_obj.r_fdmi_node_obj(wwn)  # FDMI HBA (node) alerts
            if obj is not None:
                a_list.extend(obj.r_alert_objects())
            obj = fab_obj.r_fdmi_port_obj(wwn)  # FDMI port alerts
            if obj is not None:
                a_list.extend(obj.r_alert_objects())
    return a_list


def combined_login_alerts(login_obj, wwn):
    """Converts alerts associated with a login object and the login and FDMI objects for lwwn to a human-readable string

    :param login_obj: Port object
    :type login_obj: brcddb.classes.port.PortObj
    :param wwn: Login wwn
    :type wwn: str
    :return: CSV list of alert message(s) associated with obj
    :rtype: str
    """
    a_list = combined_login_alert_objects(login_obj)
    return '\n'.join([obj.fmt_msg() for obj in a_list]) if a_list is not None and len(a_list) > 0 else ''


def combined_alert_objects(port_obj, wwn):
    """Combines alerts associated with a port object and the login and FDMI objects for wwn.

    :param port_obj: Port object
    :type port_obj: brcddb.classes.port.PortObj
    :param wwn: Login wwn
    :type wwn: str
    :return: List of AlertObj
    :rtype: list
    """
    a_list = [obj for obj in port_obj.r_alert_objects()]  # All port alerts
    fab_obj = port_obj.r_fabric_obj()
    if fab_obj is not None and wwn is not None:
        obj = fab_obj.r_login_obj(wwn)
        if obj is not None:
            a_list.extend(obj.r_alert_objects())
        obj = fab_obj.r_fdmi_node_obj(wwn)
        if obj is not None:
            a_list.extend(obj.r_alert_objects())
        obj = fab_obj.r_fdmi_port_obj(wwn)
        if obj is not None:
            a_list.extend(obj.r_alert_objects())
    return a_list


def combined_alerts(port_obj, wwn):
    """Converts alerts associated with a port object and the login and FDMI objects for lwwn to a human-readable string.

    :param port_obj: Port object
    :type port_obj: brcddb.classes.port.PortObj
    :param wwn: Login wwn
    :type wwn: str
    :return: CSV list of alert message(s) associated with obj
    :rtype: str
    """
    a_list = combined_alert_objects(port_obj, wwn)
    return '\n'.join([obj.fmt_msg() for obj in a_list]) if a_list is not None and len(a_list) > 0 else ''


def title_page(wb, tc, sheet_name, sheet_i, sheet_title, content, col_width):
    """Creates a title page for the Excel report.

    content is as defined as noted below. Any unspecified item is ignored which means the default is whatever default is
    configured in Excel.
        content_item = {
            'new_row':  False           # Default is True. Setting this to False is useful when using different cell
                                        # attributes for different columns. Otherwise, all cell attributes are applied
                                        # to all cells.
            'font': 'hdr_1',            # Predefined font type in font_tbl
            'fill': 'light_blue'        # Predefined cell fill type in fill_tbl
            'border': 'thin'            # Predefined border type in border_tbl
            'align': 'wrap'             # Predefined align type in align_tbl
            'merge': 2,                 # Number of columns to merge, starting from current column
            'disp': ('col_1', 'col_2', 'etc.')   # Str, list or tuple. Content to add to the worksheet
        }
        Example:
        content = (
            {'font': 'hdr_1', 'merge': 3, 'align': 'wrap_center', 'disp': 'Title Page'},
            {}, # Skip a row
            {'font': 'std', 'align': 'wrap', 'disp': ('col_1', 'col_2', 'col_3') }
        )

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1.
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed. Typically, 0. Default is 0
    :type sheet_i: int, None
    :param sheet_title: Title to be displayed in large font, hdr_1, with light blue fill at the top of the sheet
    :type sheet_title: str
    :param content: Caller defined content. List or tuple of dictionaries to add to the title page. See comments above
    :type content: list, tuple
    :param col_width: Column widths of list of column widths to set on sheet
    :type col_width: int, list, tuple
    :rtype: None
    """

    # Set up the sheet
    col = 1
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i,
                            title=excel_util.valid_sheet_name.sub('_', sheet_name))
    for i in gen_util.convert_to_list(col_width):
        sheet.column_dimensions[xl_util.get_column_letter(col)].width = i
        col += 1
    max_col = col-1
    row = col = 1
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', link='#' + tc + '!A1', font=excel_fonts.font_type('link'))
        col += 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)
    excel_util.cell_update(sheet, row, col, sheet_title, font=excel_fonts.font_type('hdr_1'),
                           fill=excel_fonts.fill_type('lightblue'))
    row += 2
    # Add the content. Intended for general use so lots of error checking.
    col = 1
    if isinstance(content, (list, tuple)):
        for obj in content:
            if isinstance(obj, dict):
                display = list()
                if 'disp' in obj:
                    if isinstance(obj.get('disp'), (str, int, float)):
                        display.append(obj.get('disp'))
                    elif isinstance(obj.get('disp'), (list, tuple)):
                        display = obj.get('disp')
                    elif obj.get('disp') is None:
                        pass
                    else:
                        brcdapi_log.exception('Unknown disp type, ' + str(type((obj.get('disp')))) + ', at row ' +
                                              str(row), echo=True)
                for buf in display:
                    excel_util.cell_update(sheet, row, col, buf,
                                           link=obj.get('hyper'),
                                           font=excel_fonts.font_type(obj.get('font')),
                                           align=excel_fonts.align_type(obj.get('align')),
                                           border=excel_fonts.border_type(obj.get('border')),
                                           fill=excel_fonts.fill_type(obj.get('fill')))
                    if 'merge' in obj:
                        if isinstance(obj.get('merge'), int):
                            if obj.get('merge') > 1:
                                sheet.merge_cells(start_row=row, start_column=col, end_row=row,
                                                  end_column=(col + obj.get('merge') - 1))
                                col += obj.get('merge')
                        else:
                            brcdapi_log.exception('Merge must be an integer. Type: ' + str(type(obj.get('merge'))),
                                                  echo=True)
                            col += 1
                    else:
                        col += 1
                if 'new_row' not in obj or ('new_row' in obj and obj.get('new_row')):
                    row, col = row+1, 1
            else:
                brcdapi_log.exception('Invalid type in content list, ' + str(type(obj)) + ', at row ' + str(row),
                                      echo=True)
    else:
        brcdapi_log.exception('Invalid content type: ' + str(type(content)), echo=True)


def get_next_switch_d(switch_list, val, test_type, ignore_case=False):
    """Finds the first match in sl list returned from excel_util.read_sheet() and returns the next entry

    :param switch_list: A list of dictionaries as returned from excel_util.read_sheet()
    :type switch_list: list, tuple
    :param val: The value to look for
    :type val: str, int, float
    :param test_type: The type of test to perform. See brcddb.util.search.match_test
    :type test_type: str
    :param ignore_case: If True, performs a case-insensitive search
    :return: Entry in switch_list. None if not found or if the match was the last entry in switch_list
    :rtype: dict, None
    """
    ml = brcddb_search.match_test(switch_list, dict(k='val', v=val, t=test_type, i=ignore_case))
    if len(ml) > 0:
        cell = ml[0]['cell']  # Find this cell
        for i in range(0, len(switch_list)):
            if switch_list[i]['cell'] == cell:
                return None if i + 1 >= len(switch_list) else switch_list[i+1]

    return None


# Rather than rely on the user to not move columns or worksheets around, we use the tables below to figure out where
# everything is. The first key is a common name for each rule which is only used internally in _parse_sfp_file() to sort
# out all the rules. The keys in 'mem' are used to match the column headers in the Excel workbook.
_convert_action = {  # Key is the CLI action and value is the API action
    'decom': 'decomission',  # Yes, decommission is spelled incorrectly in FOS
    'email': 'e-mail',
    'fence': 'port-fence',
    'toggle': 'port-toggle',
    'sfp_marginal': 'sfp-marginal',
    'snmp': 'snmp-trap',
    'sw_critical': 'sw-critical',
    'sw_marginal': 'sw-marginal',
    'unquar': 'un-quarantine',
    'uninstall_vtap': 'vtap-uninstall',
}


def _action_value(group, val):
    return {'action': [_convert_action[buf] if buf in _convert_action else buf for buf in val.split(';')]}


def _name_value(group, val):
    return val + group


def _gen_value(group, val):
    return val


def _int_str_value(group, val):
    return str(val)


_rules = {
    'current_h': {
        'monitoring-system': 'CURRENT',
        'logical-operator': 'ge',
        'mem': {
            'Current High (mA)': {'api': 'threshold-value', 'val': _int_str_value},
            'Current High Name Prefix': {'api':  'name', 'val': _name_value},
            'Current High Sev': {'api':  'event-severity', 'val': _gen_value},
            'Current High QT': {'api':  'quiet-time', 'val': _gen_value},
            'Current High Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'current_l': {
        'monitoring-system': 'CURRENT',
        'logical-operator': 'le',
        'mem': {
            'Current Low (mA)': {'api':  'threshold-value', 'val': _int_str_value},
            'Current Low Name Prefix': {'api':  'name', 'val': _name_value},
            'Current Low Sev': {'api':  'event-severity', 'val': _gen_value},
            'Current Low QT': {'api':  'quiet-time', 'val': _gen_value},
            'Current Low Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'voltage_h': {
        'monitoring-system': 'VOLTAGE',
        'logical-operator': 'ge',
        'mem': {
            'Voltage High (mV)': {'api':  'threshold-value', 'val': _int_str_value},
            'Voltage High Name Prefix': {'api':  'name', 'val': _name_value},
            'Voltage High Sev': {'api':  'event-severity', 'val': _gen_value},
            'Voltage High QT': {'api':  'quiet-time', 'val': _gen_value},
            'Voltage High Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'voltage_l': {
        'monitoring-system': 'VOLTAGE',
        'logical-operator': 'le',
        'mem': {
            'Voltage Low (mV)': {'api':  'threshold-value', 'val': _int_str_value},
            'Voltage Low Name Prefix': {'api':  'name', 'val': _name_value},
            'Voltage Low Sev': {'api':  'event-severity', 'val': _gen_value},
            'Voltage Low QT': {'api':  'quiet-time', 'val': _gen_value},
            'Voltage Low Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'temp_h': {
        'monitoring-system': 'SFP_TEMP',
        'logical-operator': 'ge',
        'mem': {
            'Temp High (C)': {'api':  'threshold-value', 'val': _int_str_value},
            'Temp High Name Prefix': {'api':  'name', 'val': _name_value},
            'Temp High Sev': {'api':  'event-severity', 'val': _gen_value},
            'Temp High QT': {'api':  'quiet-time', 'val': _gen_value},
            'Temp High Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'temp_l': {
        'monitoring-system': 'SFP_TEMP',
        'logical-operator': 'le',
        'mem': {
            'Temp Low (C)': {'api':  'threshold-value', 'val': _int_str_value},
            'Temp Low Name Prefix': {'api':  'name', 'val': _name_value},
            'Temp Low Sev': {'api':  'event-severity', 'val': _gen_value},
            'Temp Low QT': {'api':  'quiet-time', 'val': _gen_value},
            'Temp Low Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'tx_h': {
        'monitoring-system': 'TXP',
        'logical-operator': 'ge',
        'mem': {
            'Tx High (uW)': {'api':  'threshold-value', 'val': _int_str_value},
            'Tx High Name Prefix': {'api':  'name', 'val': _name_value},
            'Tx High Sev': {'api':  'event-severity', 'val': _gen_value},
            'Tx High QT': {'api':  'quiet-time', 'val': _gen_value},
            'Tx High Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'tx_l': {
        'monitoring-system': 'TXP',
        'logical-operator': 'le',
        'mem': {
            'Tx Low (uW)': {'api':  'threshold-value', 'val': _int_str_value},
            'Tx Low Name Prefix': {'api':  'name', 'val': _name_value},
            'Tx Low Sev': {'api':  'event-severity', 'val': _gen_value},
            'Tx Low QT': {'api':  'quiet-time', 'val': _gen_value},
            'Tx Low Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'rx_h': {
        'monitoring-system': 'RXP',
        'logical-operator': 'ge',
        'mem': {
            'Rx High (uW)': {'api':  'threshold-value', 'val': _int_str_value},
            'Rx High Name Prefix': {'api':  'name', 'val': _name_value},
            'Rx High Sev': {'api':  'event-severity', 'val': _gen_value},
            'Rx High QT': {'api':  'quiet-time', 'val': _gen_value},
            'Rx High Action': {'api':  'actions', 'val': _action_value},
        }
    },
    'rx_l': {
        'monitoring-system': 'RXP',
        'logical-operator': 'le',
        'mem': {
            'Rx Low (uW)': {'api':  'threshold-value', 'val': _int_str_value},
            'Rx Low Name Prefix': {'api':  'name', 'val': _name_value},
            'Rx Low Sev': {'api':  'event-severity', 'val': _gen_value},
            'Rx Low QT': {'api':  'quiet-time', 'val': _gen_value},
            'Rx Low Action': {'api':  'actions', 'val': _action_value},
        },
    },
}


def parse_sfp_file(file):
    """Parses Excel file with the new SFP rules. See sfp_rules_rx.xlsx

    :param file: Path and name of Excel Workbook with new SFP rules
    :type file: str
    :return: List of dictionaries. The key for each dictionary is the column header and the value is the cell value
    :rtype: list
    """
    rl = list()

    # Load the workbook & contents
    if file is not None:
        try:
            # Return everything up to '__END__'
            for d in excel_util.parse_parameters(sheet_name='new_SFP_rules', hdr_row=0, wb_name=file)['content']:
                if d['Group'] == '__END__':
                    break
                else:
                    rl.append(d)
        except FileNotFoundError:
            brcdapi_log.log('SFP rules workbook: ' + str(file) + ' not found.', echo=True)
            return list()
        except FileExistsError:
            brcdapi_log.log('Folder in ' + file + ' does not exist', echo=True)
            return list()

    return rl


def find_headers(hdr_row_l, hdr_l=None, warn=False):
    """Moved to brcdapi.excel_util.find_headers()"""
    return excel_util.find_headers(hdr_row_l, hdr_l=hdr_l, warn=warn)


def _parse_chassis_sheet(sheet_d):
    """Parses the "Chassis" worksheet from a switch configuration workbook into a dictionary. The key is the KPI. The
    value is the content.

    :param sheet_d: Output from excel_util.read_workbook() for the Chassis worksheet
    :type sheet_d: dict
    :return error_l: Error messages. Empty if no error encountered
    :rtype error_l: list
    :return: Dictionary for chassis.
    :rtype: dict
    """
    error_l, rd = list(), dict()
    al = sheet_d['al']
    if len(al) < 2:
        return error_l, rd

    # Find the 'Area' and 'Parameter' columns
    col_d = excel_util.find_headers(al[0], ('Area', 'Parameter'))
    for k, v in col_d.items():
        if v is None:
            error_l.append('Sheet ' + sheet_d['switch_info']['sheet_name'] + ' missing column ' + k)
    if len(error_l) > 0:
        return error_l, None

    # Add all the data to the return dictionary
    area_i, content_d, working_key = col_d['Area'], None, None
    for row in range(1, len(al)):
        new_key, row_l = False, al[row]
        for k in ('running', 'operations'):
            if len(row_l[area_i]) > len(k) and row_l[area_i][0:len(k)] == k:
                new_key, working_key, content_d = True, row_l[area_i], rd.get(working_key)
                if content_d is None:
                    content_d = dict()
                    rd.update({working_key: content_d})
                break
        if new_key:
            continue
        if content_d is None:
            error_l.append('Chassis sheet: Content found before a KPI was defined in row ' + str(row+1))
            continue
        val = row_l[col_d['Parameter']]
        if val is None:
            continue
        content_d.update({row_l[area_i]: val})

    return error_l, rd


# Case methods for _switch_d
def _conv_add(error_l, obj_d, switch_d, val):
    """Used in _switch_d to add a key/value pair to obj_d

    :param error_l: List of error messages to append error messages to
    :type error_l: list
    :param obj_d: Object to update with Rest API key and values
    :type obj_d: dict
    :param switch_d: Dictionary in _switch_d
    :type switch_d: dict
    :param val: Value from the configuration workbook to add to key
    :type val: None, int, float, str
    :rtype: None
    """
    add_val = switch_d['d'] if val is None else val
    if add_val is not None:
        gen_util.add_to_obj(obj_d, switch_d['k'], add_val)


def _conv_port_name(error_l, obj_d, switch_d, val):
    """Used in _switch_d to convert hex to decimal int. See _conv_add for parameter definitions"""
    gen_util.add_to_obj(obj_d, 'switch_info/port_name', val)
    port_name_mode = switch_d['d'] if not isinstance(val, str) else 'off' if val in ('open -n, ficon -n') else val
    try:
        _conv_add(error_l, obj_d, switch_d, port_name_mode)
    except (ValueError, TypeError):
        error_l.append('Value for ' + switch_d['k'] + ', ' + str(val) + ', is not a valid hex number.')


def _conv_hex(error_l, obj_d, switch_d, val):
    """Used in _switch_d to convert hex to decimal int. See _conv_add for parameter definitions"""
    try:
        _conv_add(error_l, obj_d, switch_d, int(val.replace('0x', ''), 16))
    except (ValueError, TypeError):
        error_l.append('Value for ' + switch_d['k'] + ', ' + str(val) + ', is not a valid hex number.')


def _conv_yn_bool(error_l, obj_d, switch_d, val):
    """Used in _switch_d to convert "yes" or "no" to a bool. See _conv_add for parameter definitions"""
    add_val = switch_d['d'] if val is None else val
    if isinstance(add_val, str):
        add_val = True if add_val.lower() == 'yes' else False
    try:
        _conv_add(error_l, obj_d, switch_d, add_val)
    except AttributeError:
        error_l.append('Expected type bool or str for ' + switch_d['k'] + '. Received type ' + str(type(val)))


def _conv_req_int(error_l, obj_d, switch_d, val):
    """Used in _switch_d. An int type val is required. See _conv_add for parameter definitions"""
    add_val = switch_d['d'] if val is None else val
    if isinstance(add_val, int):
        _conv_add(error_l, obj_d, switch_d, add_val)
    else:
        error_l.append('Expected type int for ' + switch_d['k'] + '. Received type ' + str(type(val)))


def _conv_cup(error_l, obj_d, switch_d, val):
    """Enable CUP. See _conv_add for parameter definitions"""
    val_bool = True if isinstance(val, str) and val.lower() == 'yes' else False
    if obj_d['switch_info']['switch_type'] == 'ficon':
        _conv_add(error_l, obj_d, switch_d, val_bool)
        if val_bool:
            gen_util.add_to_obj(obj_d, 'running/brocade-ficon/cup/active-equal-saved-mode', True)
            # At the time this was written, modifying mihpto was not supported in the API. It was GET only.
            # gen_util.add_to_obj(obj_d, 'running/brocade-ficon/cup/mihpto', 180)
    elif val_bool:
        error_l.append('Enable CUP is only supported on "ficon" switch types. Sheet: ' + '')


def _conv_routing_policy(error_l, obj_d, switch_d, val):
    """Used in _switch_d. To determine the routing policy. See _conv_add for parameter definitions"""
    global _apt_policy_d

    try:
        _conv_add(error_l, obj_d, switch_d, _apt_policy_d.get(val))
    except (TypeError, ValueError):
        _conv_add(error_l, obj_d, switch_d, 'default')


def _conv_allow_xisl(error_l, obj_d, switch_d, val):
    """Used in _switch_d. Determines if allow XISL should be set. See _conv_add for parameter definitions"""
    if obj_d['switch_info']['switch_type'] != 'base':
        _conv_yn_bool(error_l, obj_d, switch_d, val)


""" _switch_d is used in _parse_switch_sheet() to determine what should be parsed out of the spreadsheet and how it
should be interpreted. The key is the value in the cell in the "Area" column. The value is a dict as follows:
    +-------+-------------------------------------------------------------------------------------------------------+
    | key   | Description                                                                                           |
    +=======+=======================================================================================================+
    | k     | Rest API branch corresponding to "Area" in the switch configuration workbooks. At the time this was   |
    |       | written, only running branches were included. This is because "operations" branches require special   |
    |       | handling. There is also a special key "switch_info". Most of the information in switch_info is what   |
    |       | is passed to the brcdapi.switch.py module which handles "operations" branch and other switch actions  |
    |       | that require special handling.                                                                        | 
    +-------+-------------------------------------------------------------------------------------------------------+
    | d     | The default value to assign in the event the parameter is missing from the spreadsheet                |
    +-------+-------------------------------------------------------------------------------------------------------+
    | c     | Pointer to value conversion action.                                                                   |
    +-------+-------------------------------------------------------------------------------------------------------+
"""
_switch_d = {  # See above for definitions. This table is used in _parse_switch_sheet()
    'Fabric ID (FID)': dict(k='switch_info/fid', d=None, c=_conv_req_int),
    'Fabric Name': dict(k='running/' + brcdapi_util.bfs_fab_user_name, d=None, c=_conv_add),
    'Switch Name': dict(k='running/' + brcdapi_util.bfs_sw_user_name, d=None, c=_conv_add),
    'Domain ID (DID)': dict(k='running/' + brcdapi_util.bfs_did, d=None, c=_conv_hex),
    'Insistent DID': dict(k='running/' + brcdapi_util.bfc_idid, d=True, c=_conv_yn_bool),
    'Fabric Principal Enable': dict(k='running/' + brcdapi_util.bfc_principal_en, d=False, c=_conv_yn_bool),
    'Fabric Principal Priority': dict(k='running/' + brcdapi_util.bfc_principal_pri, d=None, c=_conv_add),
    'Allow XISL': dict(k='running/' + brcdapi_util.bfc_xisl_en, d=False, c=_conv_allow_xisl),
    'Enable Switch': dict(k='switch_info/enable_switch', d=False, c=_conv_yn_bool),
    'Enable Ports': dict(k='switch_info/enable_ports', d=False, c=_conv_yn_bool),
    'Login Banner': dict(k='running/' + brcdapi_util.bfs_banner, d=None, c=_conv_add),
    'Switch Type': dict(k='switch_info/switch_type', d=None, c=_conv_add),
    'Duplicate WWN': dict(k='running/' + brcdapi_util.bfc_fport_enforce_login, d=0, c=_conv_req_int),
    'Bind': dict(k='switch_info/bind', d=False, c=_conv_yn_bool),
    'Routing Policy': dict(k='running/' + brcdapi_util.bfs_adv_tuning, d=None, c=_conv_routing_policy),
    'Port Name': dict(k='running/' + brcdapi_util.bfc_portname_mode, d=None, c=_conv_port_name),
    'Port Name Format': dict(k='running/' + brcdapi_util.bfc_portname_format, d='S.T.I.A', c=_conv_add),
    'Enable CUP': dict(k='running/' + brcdapi_util.ficon_cup_en, d=False, c=_conv_cup),
    'Enable HTTPS Timeout': dict(k='switch_info/TBD', d=True, c=_conv_yn_bool),
}


def _parse_switch_sheet(sheet_d):
    """Parses a "Switch_x" worksheet from one of the witch configuration worksheets and returns a dictionary of Rest API
    branch and leaf names. The values are what to send to the switch. Also contains a dictionary whose key is

    +---------------+-------+-------------------------------------------------------------------------------+
    | Key           | type  | Value Description                                                             |
    +===============+=======+===============================================================================+
    | switch_info   | dict  | See switch_info below                                                         |
    +---------------+-------+-------------------------------------------------------------------------------+
    | err_msgs      | list  | Although no longer used by this module, the list is created and left empty    |
    |               |       | because other modules may use it and expect this key to be present.           |
    +---------------+-------+-------------------------------------------------------------------------------+
    | port_d        | dict  | Key is port in s/p notation. Value is a dict as read from the Slot x sheets   |
    +---------------+-------+-------------------------------------------------------------------------------+
    | running       | dict  | Branch and leaves in Rest API format for the running branch                   |
    +---------------+-------+-------------------------------------------------------------------------------+
    | operations    | dict  | Branch and leaves in Rest API format for the operations branch. As of         |
    |               |       | 29 Dec 2022 there was nothing to put in here so the branch was not present.   |
    +---------------+-------+-------------------------------------------------------------------------------+

    "switch_info" defined as follows:

    +---------------+-----------+-----------------------------------------------------------------------------------+
    | key           | type      | Description                                                                       |
    +===============+===========+===================================================================================+
    | bind          | bool      | If True, bind the port addresses.                                                 |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | enable_ports  | bool      | If True, enable the ports when done.                                              |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | enable_switch | bool      | If True, enable the switch when done.                                             |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | fid           | int       | Fabric ID                                                                         |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | sheet_name    | str       | Sheet name from switch configuration workbook                                     |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | switch_type   | str       | base, ficon, or open                                                              |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | port_name     | str       | Port Name from switch configuration workbook                                      |
    +---------------+-----------+-----------------------------------------------------------------------------------+

    :param sheet_d: Output from excel_util.read_workbook() for the Chassis worksheet
    :type sheet_d: dict
    :return error_l: List of error messages. Empty if no error encountered
    :rtype error_l: list
    :return: Dictionary for switch. None if an error was encountered.
    :rtype: dict, None
    """
    global _switch_d

    # Find the 'Area' and 'Parameter' columns
    error_l, rd, al = list(), dict(switch_info=dict(), port_d=dict(), err_msgs=list()), sheet_d['al']
    col_d = excel_util.find_headers(al[0], ('Area', 'Parameter'))
    for k, v in col_d.items():
        if v is None:
            error_l.append('Sheet ' + sheet_d['switch_info']['sheet_name'] + ' missing column ' + k)
    if len(error_l) > 0:
        brcdapi_log.log(error_l, echo=True)
        return None

    # Add and validate all the data from the worksheet and add it to the return dictionary, rd
    for row_l in al[1:]:  # Opps. I forgot I needed to determine the switch type before checking other parameters
        if row_l[col_d['Area']] == 'Switch Type':
            switch_d = _switch_d.get(row_l[col_d['Area']])
            if isinstance(switch_d, dict):  # error_l, k, val, def_val
                switch_d['c'](error_l, rd, switch_d, row_l[col_d['Parameter']])
            break
    for row_l in al[1:]:
        switch_d = _switch_d.get(row_l[col_d['Area']])
        if isinstance(switch_d, dict):  # error_l, k, val, def_val
            switch_d['c'](error_l, rd, switch_d, row_l[col_d['Parameter']])
        else:
            error_l.append('Invalid key "' + row_l[col_d['Area']] + '" sheet: ' + sheet_d['switch_info']['sheet_name'])

    # If it's a ficon switch, in order delivery and DLS needs to be set
    if rd['switch_info']['switch_type'] == 'ficon':
        gen_util.add_to_obj(rd, 'running/' + brcdapi_util.bfs_dls, 'two-hop-lossless-dls')

    return error_l, rd


"""
+-------+-----------------------------------------------------------------------------------------------+
| k     | Key used in the dictionaries returned from _parse_slot_sheet(). IDK why I didn't just use the |
|       | same key as is used in the API. That would have saved the trouble of needing a look up table, |
|       | _switch_d_to_api, in switch_config.py but I'm not changing working code.                      |
+-------+-----------------------------------------------------------------------------------------------+
| h     | If True, treat the cell value as a hex number and convert to a decimal int                    |
+-------+-----------------------------------------------------------------------------------------------+
| i     | If True, convert the cell value to an int.                                                    |
+-------+-----------------------------------------------------------------------------------------------+
| s     | If True, convert the cell value to a str                                                      |
+-------+-----------------------------------------------------------------------------------------------+
| slot  | If True, prepend the slot number + '/' to the value. Used to convert ports to s/p notation    |
+-------+-----------------------------------------------------------------------------------------------+
| p     | When the length of the cell value is less than that specified by p, 0s are prepended to make  |
|       | the value this length. Used for creating FC addresses.                                        | 
+-------+-----------------------------------------------------------------------------------------------+
| pn    | When True, check the port naming convention and adjust the port name accordingly.             |
+-------+-----------------------------------------------------------------------------------------------+
"""
_slot_d_find = {  # See table above
    'Port': dict(k='port', h=False, i=False, s=True, slot=True, p=0, pn=False),
    'DID (Hex)': dict(k='did', h=True, i=False, s=True, slot=False, p=0, pn=False),
    'Port Addr (Hex)': dict(k='port_addr', h=False, i=False, s=True, slot=False, p=2, pn=False),
    'Index': dict(k='index', h=False, i=True, s=True, slot=False, p=0, pn=False),
    'Link Addr': dict(k='link_addr', h=False, i=False, s=True, slot=False, p=4, pn=False),
    'FID': dict(k='fid', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'CLI': dict(k='cli', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'Attached Device': dict(k='desc', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'ICL Description': dict(k='desc', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'Port Name': dict(k='port_name', h=False, i=False, s=False, slot=False, p=0, pn=True),
    'Low Qos VC': dict(k='low_vc', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'Med Qos VC': dict(k='med_vc', h=False, i=False, s=False, slot=False, p=0, pn=False),
    'High Qos VC': dict(k='high_vc', h=False, i=False, s=False, slot=False, p=0, pn=False),
}


def _parse_slot_sheet(sheet_d, port_name_d):
    """Parses a "Slot x" worksheet from a configuration workbook into a dictionary. The key is the port number in s/p
    notation. The value for each port dictionary is as follows:

    +---------------+-----------+-----------------------------------------------------------------------------------+
    | key           | type      | Description                                                                       |
    +===============+===========+===================================================================================+
    | fid           | int       | Fabric ID as a decimal.                                                           |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | did           | int       | Domain ID. Although read from the workbook as a str in Hex, it is returned as a   |
    |               |           | decimal int                                                                       |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | port_addr     | str       | Port address in hex as read from the sheet (no leading 0x). Padded to 2 places.   |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | index         | int       | Port index                                                                        |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | link_addr     | str       | Link address in hex as read from the sheet (no leading 0x)                        |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | desc          | str       | Attached Device or ICL Description                                                |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | low_vc        | int       | Low Qos VC                                                                        |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | med_vc        | int       | Medium Qos VC                                                                     |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | high_vc       | int       | High Qos VC                                                                       |
    +---------------+-----------+-----------------------------------------------------------------------------------+

    :param sheet_d: Output from excel_util.read_workbook() for the Chassis worksheet
    :type sheet_d: dict
    :param port_name_d: Key is the FID as an int. Value is the port naming convention
    :type port_name_d: dict()
    :return error_l: List of error messages. Empty if no error encountered
    :rtype error_l: list
    :return: Dictionary for switch. None if an error was encountered.
    :rtype: dict, None
    """
    global _slot_d_find, _MAX_FICON_PORT_NAME

    error_l, rd, al, buf, slot = list(), dict(), sheet_d['al'], '', ''

    # Find the slot number
    for buf in al[0]:
        if buf is not None and 'slot' in buf.lower():
            slot = buf.split(' ')[1] + '/'
            break

    # Find the column headers
    i, hdr_l, col_d_l = 0, al[1], [dict(), dict()]
    for col in range(0, len(hdr_l)):
        try:
            r_key = _slot_d_find[hdr_l[col]]['k']
            if r_key in col_d_l[i]:
                i += 1  # A repeat means we're starting the second column of ports
                if i > len(col_d_l):
                    error_l.append('Invalid column header, "' + buf + '", on sheet ' + sheet_d['sheet'])
                    break
            col_d_l[i].update({r_key: col})
        except KeyError:
            pass  # Other columns for documentation purposes may be present so just skip them

    # Add the data from the worksheet
    for row in range(3, len(al)):

        if al[row][col_d_l[0]['port']] is None:
            break  # If the port is None, we reached the end of the workbook

        for i in range(0, 2):  # Director class products have 2 columns, each with the same headers
            if len(col_d_l[i]) == 0:
                break

            # Build the port dict and add to the return dict
            pd = dict()
            for k, d in _slot_d_find.items():
                r_key = _slot_d_find[k]['k']
                if r_key in col_d_l[i]:
                    v = str(al[row][col_d_l[i][r_key]]) if d['s'] else al[row][col_d_l[i][r_key]]
                    if d['h']:
                        try:
                            v = int(v, 16)
                        except ValueError:
                            error_l.append('Value for ' + k + ', ' + str(v) + ', is not a valid hex number.')
                    if d['i']:
                        if v != 'None':
                            try:
                                v = int(v)
                            except ValueError:
                                error_l.append('Value for ' + k + ', ' + str(v) + ', is not valid in '+sheet_d['sheet'])
                    if d['pn']:
                        try:
                            port_name = port_name_d[pd['fid']]
                            if port_name in ('open -n', 'ficon -n'):
                                if isinstance(v, str):
                                    if port_name == 'ficon -n' and len(v) > _MAX_FICON_PORT_NAME:
                                        buf = 'The port name, ' + v + ', for port ' + str(pd.get('port')) + ' is ' + \
                                              str(len(v)) + '. The maximum supported FICON Port name length is ' + \
                                              str(_MAX_FICON_PORT_NAME)
                                        error_l.append(buf)
                                        v = v[0: _MAX_FICON_PORT_NAME]
                                else:
                                    v = ''
                        except KeyError:
                            pass  # We're not configuring this FID

                    if isinstance(v, str):
                        while len(v) < d['p']:
                            v = '0' + v
                    v = None if isinstance(v, str) and v == 'None' else slot + v if d['slot'] else v
                    pd.update({r_key: v})
            rd.update({pd['port']: pd})

    # Return the slot dict or None if there was an error
    return error_l, rd


def parse_switch_file(file):
    """Parses Excel switch configuration Workbook.

    :param file: Path and name of Excel Workbook with switch configuration definitions
    :type file: str
    :return error_l: List of errors. Empty list if no errors found
    :rtype error_l: list
    :return chassis_d: Dictionary of chassis parameters. None if error reding workbook
    :rtype chassis_d: dict, None
    :return switch_l: List of logical switch dictionaries as described below.
    :rtype switch_l: list

    Note: This started as something very different. If I had to do it over again, I would have used the FOS API
    mnemonics for these values.

    +---------------+-----------+-----------------------------------------------------------------------------------+
    | key           | type      | Description                                                                       |
    +===============+===========+===================================================================================+
    | banner        | None, str | Login banner. Not set if None.                                                    |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | bind          | bool      | If True, bind the addresses to the ports.                                         |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | did           | int       | Although read from the workbook as a str in Hex, it is returned as a decimal int  |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | dup_wwn       | int       | 0 - First login takes precedence.                                                 |
    |               |           | 1 - Second FLOGI and FDISC takes precedence.                                      |
    |               |           | 2 - First FLOGI takes precedence. Second FDISC takes precedence                   |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | enable_ports  | bool      | If True, enable the ports after configuration is complete.                        |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | enable_switch | bool      | If True, enable the switch after configuration is complete.                       |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | fab_name      | None, str | Fabric name. Not set if None                                                      |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | port_name     | None, str | User friendly port name                                                           |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | ports         | dict      | Key is the port number and value is a dictionary as follows:                      |
    |               |           |  Key          Type    Value                                                       |
    |               |           |  did          int     The hexadecimal DID converted to decimal.                   |
    |               |           |  port_addr    str     The hexadecimal port address (middle byte of the FC address)|
    |               |           |  index        int     The port index                                              |
    |               |           |  link_addr    str     FICON link address in hex                                   |
    |               |           |  fid          int     Fabric ID                                                   |
    |               |           |  ad           str,None    Attached device description. None if left blank.        |
    |               |           |  low_vc       int     VC for QOSL zone                                            |
    |               |           |  med_vc       int     VC for QOSM zone                                            |
    |               |           |  high_vc      int     VC for QOSH zone                                            |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | sheet_name    | str       | Name of sheet in Workbook.                                                        |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | switch_name   | None, str | Switch name. Not set if None                                                      |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    | xisl          | bool      | If True, base switch usage is allowed.                                            |
    +---------------+-----------+-----------------------------------------------------------------------------------+
    """
    chassis_d, port_d, port_name_d, rd = None, dict(), None, dict()

    # Load the workbook
    skip_l = ('About', 'Summary', 'Sheet', 'Instructions', 'Slot_x_Terms', 'CLI_Bind', 'lists', 'VC')
    error_l, sheet_l = excel_util.read_workbook(file, dm=3, skip_sheets=skip_l)
    if len(error_l) > 0:
        return error_l, chassis_d, list()

    # Parse the "Chassis", "Sheet_x", and "Slot x" worksheets
    for sheet_d in sheet_l:
        title = sheet_d['sheet']
        if fnmatch.fnmatch(title, 'Chassis*'):
            ml, chassis_d = _parse_chassis_sheet(sheet_d)
            error_l.extend(ml)
        elif fnmatch.fnmatch(title, 'Switch*'):
            ml, d = _parse_switch_sheet(sheet_d)
            error_l.extend(ml)
            gen_util.add_to_obj(d, 'switch_info/sheet_name', title)
            fid = d['switch_info']['fid']
            if fid in rd:
                error_l.append('Duplicate FID, ' + str(fid) + '. Appears in ' + title + ' and ' +
                               rd[fid]['switch_info']['sheet_name'])
            else:
                rd.update({fid: d})
        elif fnmatch.fnmatch(title, 'Slot ?*'):
            if not isinstance(port_name_d, dict):
                # This could have been more elegant. I forgot that some port naming modes want all ports not explicitly
                # named to be set to an empty string. Furthermore, ficon port names need to be truncated to a maximum of
                # 24 characters, so I jammed this in and made a hack in _parse_slot_sheet() to set the names correctly.
                port_name_d = dict()
                for fid, temp_switch_d in rd.items():
                    port_name_d.update({fid: temp_switch_d['switch_info']['port_name']})
            ml, d = _parse_slot_sheet(sheet_d, port_name_d)
            error_l.extend(ml)
            port_d.update(d)

    # Add all the ports to the switch in the return dictionary, rd. I did it this way in case someone rearranged the
    # sheets such that a "Slot x" worksheet came before the corresponding "Switch x" worksheet.
    for k, d in port_d.items():
        switch_d = rd.get(d['fid'])
        if isinstance(switch_d, dict):
            try:
                if switch_d['switch_info']['switch_type'] == 'ficon' and int(d['port_addr'], 16) > 253:
                    ml = ['Skipping FID: ' + str(d['fid']) + ', Port: ' + d['port'] + ' Address: ' + d['port_addr'],
                          'Addresses greater than 0xFD are not supported in a ficon switch.']
                    brcdapi_log.log(ml, echo=True)
                    continue
            except TypeError:
                pass  # It's a GE port
            except KeyError:
                pass  # It doesn't have a port address. Typically, ICL ports or virtual ports

            switch_d['port_d'].update({k: d})

    return error_l, chassis_d, [switch_d for switch_d in rd.values()]


#####################################################################################
#                                                                                   #
#   New data and functions supporting the combined switch and chassis page formats  #
#                                                                                   #
#####################################################################################
def span_to_list(span):
    """Converts a span value into a list of length 3

    :param span: Value to convert to a list
    :type span: int, list, tuple
    :return: List of column span values
    :rtype: list
    """
    rl = gen_util.convert_to_list(span)
    x = rl[len(rl) - 1]
    while len(rl) < 3:
        rl.append(x)
    return rl


# _alert_d is a dictionary indexed by the URI. The value is a dictionary indexed by the alert numbers. By default, the
# alert dictionary is assumed to be empty if it's not in this table. The alerts in the switch object are searched for
# matching alert numbers in _alert_d. If found, that alert number is converted to a human-readable text message and
# inserted in the comments cell for the associated URI.
_alert_d = {
    brcdapi_util.bfc_idid: {al_table.ALERT_NUM.SWITCH_IDID: True},
    brcdapi_util.bfc_fport_enforce_login_str: {al_table.ALERT_NUM.HANDLE_DUP_WWN: True},
}
# alert_font_d is used to determine the font
alert_font_d = {
    # alert_class.ALERT_SEV.GENERAL: _std_font,
    alert_class.ALERT_SEV.WARN: _warn_font,
    alert_class.ALERT_SEV.ERROR: _error_font,
}
_p_display_d = {  # Used to override the parameter header in brcddb.app_data.report_tables.Switch.switch_display_tbl
    brcdapi_util.bfls_base_sw_en: dict(b='Switch Type', p='Base'),
    brcdapi_util.bfls_def_sw_status: dict(b='Switch Type', p='Default'),
    brcdapi_util.bfls_ficon_mode_en: dict(b='Switch Type', p='FICON'),
    '_open_switch': dict(b='Switch Type', p='Open'),
}


"""The sheet and cell location of the conditional highlighting parameters has to be stored in the switch or chassis
object. Rather than hard code it in _conditional_highlight_l, the list entries are determined in
conditional_highlight(). The key is the header. The value is a dictionary as follows:

+-----------+---------------------------------------------------+
| Key       | Value                                             |
+===========+===================================================+
| align     | One of the Excel align definitions.               |
+-----------+---------------------------------------------------+
| border    | One of the Excel border definitions.              |
+-----------+---------------------------------------------------+
| default   | The default (initial) value to fill in the cell   |
+-----------+---------------------------------------------------+
| dv        | Data validation                                   |
+-----------+---------------------------------------------------+
| font      | One of the Excel font definitions.                |
+-----------+---------------------------------------------------+

Formatting is always: font=_std_font, align=_align_wrap_r, border=_border_thin. The fill color is dynamically determined
based on the port type and if there are any errors."""
_default_stats = 999999999
_conditional_test_0_d = collections.OrderedDict()
_conditional_test_0_d['Discarded Frames'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['BB Credit Starvation'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Bit Errors'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Warn Alerts'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Error Alerts'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Frame Errors'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Link Resets'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Link Failures'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_0_d['Loss Of Signal'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d = collections.OrderedDict()
_conditional_test_1_d['Loss Of Sync'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d['Too Many RDYs'] = dict(
    default=_default_stats,
    dv=highlight_stats_dv,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d['Zone EXACT'] = dict(
    default=None,
    dv=None,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d['Zone FIND'] = dict(
    default=None,
    dv=None,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d['Alias EXACT'] = dict(
    default=None,
    dv=None,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_test_1_d['Alias FIND'] = dict(
    default=None,
    dv=None,
    font=_std_font,
    align=_align_wrap_r,
    border=_border_thin)
_conditional_highlight_l = [
    [dict(buf='Port Highlighting', font=_bold_font, align=_align_wrap)],
    [],
    [dict(
        buf='To highlight ports with zones and aliases using the EXACT or FIND functions in Excel, enter the zone '
            'object name in the box below. EXACT and FIND are case sensitive. A FIND in Excel is equivalent to '
            '*zone_name*. For all other highlighting options, enter an integer in the range 0 - 999,999,999. Numeric '
            'comparisons are greater than, ">". Highlighting is determined by a logical OR of all high lighting '
            'filters.',
        font=_std_font,
        align=_align_wrap
    )],
    [],
]


def conditional_highlight(obj, span):
    """Adds the port highlighting filters.

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param span: Number of columns to span when displaying the conditional formatting
    :type span: int
    :return: List of lists containing dictionaries to insert while processing _contents
    :rtype: list
    """
    global _conditional_highlight_l, _std_font, _align_wrap_r, _border_thin

    rl, sheet_d = list(), obj.r_get('report_app/worksheet')
    sheet_ref = sheet_d['sheet'].title + '!'
    num_columns = sheet_d['num_columns']

    # Insert the section header
    rl = copy.deepcopy(_conditional_highlight_l)
    for col_l in rl:
        for col_d in [d for d in col_l if isinstance(d, dict)]:
            col_d['span'] = num_columns
    row = len(rl) + sheet_d['row']

    # Insert the cells for the "Port Highlighting" section.
    cond_format_d = sheet_d.get('cond_format_d')
    if isinstance(cond_format_d, dict):
        for conditional_d in (_conditional_test_0_d, _conditional_test_1_d):
            col, col_h_l, col_v_l = 1, list(), list()
            rl.extend([col_h_l, col_v_l])
            for buf, d in conditional_d.items():
                col_h_l.append(dict(
                    buf=buf,
                    dv=None,
                    font=_bold_font,
                    align=d['align'],
                    border=d['border'],
                    span=span))
                col_v_l.append(dict(
                    buf=d['default'],
                    dv=d['dv'],
                    font=d['font'],
                    align=d['align'],
                    border=d['border'],
                    span=span))
                cond_format_d[buf] = sheet_ref + xl_util.get_column_letter(col) + str(row+1)  # Entry is the next row
                col += span
            row += 2
    else:
        rl = [
            [dict(buf='Port Highlighting', font=_bold_font, align=_align_wrap, span=num_columns)],
            list(),
            [dict(
                buf='Port highlighting is only supported on chassis sheets at this time.',
                font=_std_font,
                align=_align_wrap,
                span=num_columns
            )],
        ]

    return rl


def port_statistics(obj, type_span, speed_span):
    """Adds port statistics for the worksheet.

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param type_span: Number of columns to span when displaying the number of ports by login type
    :type type_span: int
    :param speed_span: Number of columns to span when displaying the number of ports by login speed
    :type speed_span: int
    :return: List of lists containing dictionaries to insert while processing _contents
    :rtype: list
    """
    global _border_thin, _align_wrap, _bold_font, _align_wrap_r

    # Figure out what to put in the statistics summary section
    port_obj_l = brcddb_search.match_test(obj.r_port_objects(), brcddb_search.port_online)
    sum_logins = 0
    for port_obj in brcddb_search.match_test(port_obj_l, brcddb_search.f_ports):
        sum_logins += len(port_obj.r_login_keys())
    sheet_d = obj.r_get('report_app/worksheet')

    # Add the ports by login type
    return [
        [dict(buf='Ports by Login Type', font=_bold_font, align=_align_wrap, span=sheet_d['num_columns'])],
        [
            dict(buf='Physical Ports', font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf='ICL-Ports', font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf='ISL (E-Ports)', font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf='FC-Lag Ports', font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf='Name Server Logins', font=_std_font, align=_align_wrap_r, border=_border_thin,
                 span=type_span),
        ],
        [
            dict(buf=len(obj.r_port_objects()),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.icl_ports)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.e_ports)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.fc_lag_ports)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
            dict(buf=sum_logins, font=_std_font, align=_align_wrap_r, border=_border_thin, span=type_span),
        ],
        [],

        # Add the ports by speed: Headers
        [dict(buf='Ports by Login Speed', font=_bold_font, align=_align_wrap, span=sheet_d['num_columns'])],
        [
            dict(buf='1G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='2G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='4G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='8G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='16G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='32G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='64G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf='128G', font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
        ],
        [  # Add the ports by speed: The values
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_1g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_2g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_4g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_8g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_16g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_32g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_64g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
            dict(buf=len(brcddb_search.match_test(port_obj_l, brcddb_search.login_128g)),
                 font=_std_font, align=_align_wrap_r, border=_border_thin, span=speed_span),
        ],
    ]


def add_content_defaults(contents_l, default_d):
    """Intended for adding default values for things such as font in a list of contents for add_contents()

    :param contents_l: Worksheet contents. This list will be modified.
    :type contents_l: list
    :param default_d: Contains keys and values (typically openpyxl.styles.fonts.Font)
    :type default_d: dict
    :return: Modified contents_l. WARNING: contents_l is returned as a convenience. contents_l is modified.
    :rtype: list
    """
    for content_l in [row_l for row_l in contents_l if isinstance(row_l, list)]:
        for d in [v0 for v0 in content_l if isinstance(v0, dict)]:
            for key, v1 in default_d.items():
                if key not in d:
                    d.update({key: v1})

    return contents_l


# _lookup_d is used in cell_content to determine which lookup table should be used to resolve descriptions.
_lookup_d = dict(
    ChassisObj=rt.Chassis.chassis_display_tbl,
    SwitchObj=rt.Switch.switch_display_tbl,
    FabricObj=rt.Fabric.fabric_display_tbl
)


def cell_content(obj, col_d, old_bool=False):
    """Builds a list of column dictionaries for rows that are Comment, Parameter, Value

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param col_d: Dictionary defining the column. Search for _contents: in brcddb.report.switch for details.
    :type col_d: dict
    :param old_bool: If True, treat the input as an old boolean style
    :type old_bool: bool
    :return: List of dictionaries with the comments, parameter, and setting values.
    :rtype: list
    """
    global _p_display_d, _std_font, _align_wrap_r, _border_thin, _lookup_d

    comment_l, font = alert_eval(obj, col_d)
    key, span_l = col_d['key'], span_to_list(col_d.get('span'))

    try:
        value = _p_display_d[key]['p']
    except KeyError:
        value = obj.r_get(key)
        if old_bool:
            value = bool(value)
        if isinstance(value, bool):
            value = str(value)
    if isinstance(value, list):
        value = '\n'.join(value)
    try:
        parameter = _p_display_d[key]['b']
    except KeyError:
        try:
            parameter = _lookup_d[brcddb_class_util.get_simple_class_type(obj)].get(key)
        except TypeError:
            parameter = 'Invalid object: ' + str(type(obj))
            brcdapi_log.exception(parameter, echo=True)
        except KeyError:
            parameter = 'Invalid object type: ' + brcddb_class_util.get_simple_class_type(obj)
            brcdapi_log.exception(
                [parameter, ' Supported types are:', '\n'.join([str(k) for k in _lookup_d.keys()])],
                echo=True
            )

    return add_content_defaults(
        [
            [
                dict(buf=None if len(comment_l) == 0 else '\n'.join(comment_l), span=span_l[0]),
                dict(buf=parameter, span=span_l[1]),
                dict(buf=value, span=span_l[2]),
            ]
        ],
        dict(font=font, align=col_d.get('align'), border=col_d.get('border'))
    )[0]


def alert_summary(obj, col_d):
    """Inserts each alert as a row.

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param col_d: Dictionary defining the column. Search for _contents: in brcddb.report.switch for details.
    :type col_d: dict
    :return: List of lists containing dictionaries with MAPS alerts associated with the object.
    :rtype: list
    """
    rl, sheet_d = list(), obj.r_get('report_app/worksheet')
    if isinstance(sheet_d, dict):
        for alert_obj in obj.r_alert_objects():
            rl.append([dict(buf=alert_obj.fmt_msg())])
    if len(rl) == 0:
        rl.append([dict(buf='None')])

    return add_content_defaults(rl, dict(font=_std_font, align=_align_wrap, span=sheet_d['num_columns']))


def alert_eval(obj, col_d):
    """Determines if there are alerts associated with an object.

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param col_d: Dictionary in the sub-list of _contents
    :type col_d: dict
    :return comment_l: List of alerts for the comments field
    :rtype: list
    :return font: Font type for the cell
    :rtype: tuple
    """
    global _alert_d, alert_font_d

    comment_l, alert_level = list(), alert_class.ALERT_SEV.GENERAL
    for alert_obj in obj.r_alert_objects():
        if alert_obj.alert_num() in _alert_d.get(col_d['key'], list()):
            comment_l.append(alert_obj.fmt_msg())
            alert_level = max(alert_level, alert_obj.sev())

    return comment_l, alert_font_d.get(alert_level, col_d.get('font'))


def seconds_to_days(obj, col_d):
    """Converts a value in seconds to days. Returns the comments, parameter, and value in a list for column insertion

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param col_d: Dictionary in the sub-list of _contents
    :type col_d: dict
    :return: List of dictionaries with the comments, parameter, and setting values.
    :rtype: list
    """
    global _p_display_d

    key, align, border, span_l = col_d['key'], col_d.get('align'), col_d.get('border'), span_to_list(col_d.get('span'))
    comment_l, font = alert_eval(obj, col_d)
    try:
        parameter = _p_display_d[key]['b']
    except KeyError:
        parameter = rt.Switch.switch_display_tbl.get(key)
    x = obj.r_get(key)
    value = int(x / 86400 + 0.5) if isinstance(x, int) else None
    return [
        dict(buf=None if len(comment_l) == 0 else '\n'.join(comment_l),
             font=font, align=align, border=border, span=span_l[0]),
        dict(buf=parameter, font=font, align=align, border=border, span=span_l[1]),
        dict(buf=value, font=font, align=align, border=border, span=span_l[2]),
    ]


def list_to_str(obj, col_d):
    """Converts a list to a \n seperated list. Returns the column list.

    :param obj: Switch or chassis object
    :type obj: brcddb.classes.switch.SwitchObj, brcddb.classes.chassis.ChassisObj
    :param col_d: Dictionary in the sub-list of _contents
    :type col_d: dict
    :return: List of dictionaries with the comments, parameter, and setting values.
    :rtype: list
    """
    global _align_wrap

    comment_l, font = alert_eval(obj, col_d)
    key, align, border, span_l = col_d['key'], col_d.get('align'), col_d.get('border'), span_to_list(col_d.get('span'))
    try:
        parameter = _p_display_d[key]['b']
    except KeyError:
        parameter = rt.Switch.switch_display_tbl.get(key)
    sl = obj.r_get(key)
    value = None if sl is None else '\n'.join(sl)
    return [
        dict(buf=None if len(comment_l) == 0 else '\n'.join(comment_l),
             font=font, align=align, border=border, span=span_l[0]),
        dict(buf=parameter, font=font, align=_align_wrap, border=border, span=span_l[1]),
        dict(buf=value, font=font, align=align, border=border, span=span_l[2]),
    ]


def add_contents(obj, contents_l):
    """Adds contents to the worksheet

    :param obj: brcddb object.
    :type obj: ProjectObj, FabricObj, SwitchObj, ChassisObj
    :param contents_l: List of dictionaries defining the contents. See brcddb.report.switch._contents for definition
    :type contents_l: list
    """
    global _yellow_fill

    col = 1
    sheet_d = obj.r_get('report_app/worksheet')
    if sheet_d is None:
        brcdapi_log.exception(str(type(obj)) + ' missing "report_app/worksheet"', echo=True)
        return
    sheet = sheet_d['sheet']

    # Add the contents
    for row_item in contents_l:
        for row_l in row_item(obj, None) if callable(row_item) else [row_item]:
            for col_item_d in row_l:
                action = col_item_d.get('a')
                col_l = action(obj, col_item_d) if callable(action) else [col_item_d]
                for col_d in col_l:
                    key, buf, buf_item = col_d.get('key', None), None, col_d.get('buf', None)
                    if callable(buf_item):
                        buf = buf_item(obj, col_d)
                    elif key is not None:
                        if isinstance(buf_item, dict):
                            buf = buf_item.get(key, None)
                        else:
                            buf = obj.r_get(key)
                    elif isinstance(buf_item, (str, int, float)):
                        buf = buf_item
                    elif buf_item is not None:
                        brcdapi_log.exception('Unexpected type, ' + str(type(buf_item)), echo=False)
                        try:
                            buf = str(type(buf_item))
                        except BaseException as e:
                            brcdapi_log.exception(['Could not convert buf_item to string.', e], echo=True)
                            buf = None  # IDK why it wouldn't be None. This is just to be certain.

                    # Is there any conditional formatting?
                    rule, rule_key = None, None
                    class_type = brcddb_class_util.get_simple_class_type(obj)
                    try:
                        rule_key = col_d.get(_obj_type_to_key_d[class_type])
                    except KeyError:
                        brcdapi_log.exception('Unexpected object type, ' + class_type, echo=True)
                    if isinstance(rule_key, str):
                        rule = Rule(type="expression",
                                    dxf=DifferentialStyle(fill=_yellow_fill, font=_bold_font),
                                    stopIfTrue=True)
                        rule.formula = [rule_key + '>0']

                    # Add the cell contents
                    excel_util.cell_update(
                        sheet,
                        sheet_d['row'],
                        col,
                        buf,
                        font=col_d.get('font'),
                        fill=col_d.get('fill'),
                        align=col_d.get('align'),
                        border=col_d.get('border'),
                        comments=col_d.get('comments'),
                        comment_height=len(col_d.get('comments', list())) + 20,
                        link=col_d.get('link'),
                        cf=rule,
                        dv=col_d.get('dv')
                    )

                    # Merge cells, if necessary
                    span = col_d.get('span', 1)
                    if span > 1:
                        sheet.merge_cells(
                            start_row=sheet_d['row'],
                            start_column=col,
                            end_row=sheet_d['row'],
                            end_column=col + span - 1)
                    col += span

            sheet_d['row'], col = sheet_d['row'] + 1, 1


def about_page(wb, sheet_i, sheet_name, file_name, version, description, tc=None):
    """Inserts a standard "About" page in a workbook

    :param wb: Workbook object
    :type wb: class
    :param sheet_i: Sheet index where page is to be placed. Typically, 0. Default is 0
    :type sheet_i: int, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param description: Workbook description. May be a string or list of strings
    :type description: str,list,tuple
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    """
    global _std_font

    # Add an "about" sheet
    sheet = wb.create_sheet(index=sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.column_dimensions['A'].width = 80

    row = col = 1
    buf_xlate_d = dict(
        script_name=file_name,
        version=version,
        description=description,
        default='',
    )

    # Add link to Table of Contents
    if isinstance(tc, str):
        excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link=tc)
        row += 1

    # First portion of the About page
    for d in _about_sheet_l:
        if d is not None:
            buf_l = gen_util.convert_to_list(buf_xlate_d.get(d.get('s', 'default')))
            for buf in buf_l:
                excel_util.cell_update(sheet, row, col, d['t'] + buf, font=d['f'], align=_align_wrap, link=d.get('l'))
                row += 1
        else:
            row += 1
