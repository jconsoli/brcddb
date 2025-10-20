"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack_consoli@yahoo.com for
details.

**Description**

Adds MAPS policy, group, and rules pages to a workbook

**Public Methods**

+-----------------------+---------------------------------------------------------------------------------------+
| Method                | Description                                                                           |
+=======================+=======================================================================================+
| maps_policy           | Creates a MAPS policy page                                                            |
+-----------------------+---------------------------------------------------------------------------------------+
| maps_groups           | Creates a MAPS groups page                                                            |
+-----------------------+---------------------------------------------------------------------------------------+
| maps_rules            | Creates a MAPS groups page                                                            |
+-----------------------+---------------------------------------------------------------------------------------+
| maps_rules            | Creates a MAPS rules page                                                             |
+-----------------------+---------------------------------------------------------------------------------------+
| read_maps             | Reads a MAPS workbook (same format as output from maps_report.py) and adds            |
|                       |'brocade-maps' to each LS with minimal switch attributes (FID and switch name)         |
+-----------------------+---------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | xx xxx 2023   | Initial launch                                                                        |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 25 Aug 2025   | Updated email address in __email__ only.                                              |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 19 Oct 2025   | Updated comments only.                                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025 Consoli Solutions, LLC'
__date__ = '19 Oct 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.3'

import collections
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.gen_util as gen_util
import brcddb.brcddb_switch as brcddb_switch
import brcddb.classes.util as class_util

_std_font = excel_fonts.font_type('std')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')

# Used in xxx. Key is what to find in the header. Value is the API key for the dictionary in brocade-maps
_rule_key_d = {'MAPS Policies': 'maps-policy', 'MAPS Groups': 'group', 'MAPS Rules': 'rule'}


class Skip(Exception):
    pass


# Used in _group_hdr_d and _rules_hdr_d to determine values
def _simple_val_act(switch_obj, d, key):
    val = d.get(key)
    try:
        return int(val)
    except (ValueError, TypeError):
        return val


def _tf_to_check(switch_obj, d, key):
    return '\u221A' if d.get(key, False) else ''


def _csv_list(d, key, sub_key):
    try:
        return ', '.join(gen_util.convert_to_list(d[key][sub_key]))
    except (ValueError, KeyError):
        return ''


def _member_csv_list(switch_obj, d, key):
    return _csv_list(d, key, 'member')


def _action_csv_list(switch_obj, d, key):
    return _csv_list(d, key, 'action')


def _rule_csv_list(switch_obj, d, key):
    return _csv_list(d, key, 'rule')


def _action_def_policy(switch_obj, d, key):
    return ', '.join(gen_util.convert_to_list(switch_obj.r_get('maps_report/rule_d/' + d['name'] + '/default')))


def _action_cust_policy(switch_obj, d, key):
    return ', '.join(gen_util.convert_to_list(switch_obj.r_get('maps_report/rule_d/' + d['name'] + '/custom')))


"""For _policy_hdr_d, _group_hdr_d, and _rules_hdr_d: Key is the key in brocade-maps/group and value is a dictionary as
follows:
+-------+-------+---------------------------------------------------------------------------------------------------+
| Key   | Type  | Description                                                                                       |
+=======+=======+===================================================================================================+
| bt    | bool  | Used by maps_config. If True, the check mark read in the report is to be interpreted as a boolean |
+-------+-------+---------------------------------------------------------------------------------------------------+
| w     | int   | Column width                                                                                      |
+-------+-------+---------------------------------------------------------------------------------------------------+
| align | class | Alignment                                                                                         |
+-------+-------+---------------------------------------------------------------------------------------------------+
| t     | str   | Header text                                                                                       |
+-------+-------+---------------------------------------------------------------------------------------------------+
| v     | func  | Method to call to determine wheat to put in the cell for each group value                         |
+-------+-------+---------------------------------------------------------------------------------------------------+
| va    | class | Alignment for the value                                                                           |
+-------+-------+---------------------------------------------------------------------------------------------------+

_r_maps is used in read_maps() to convert the table back to API format
"""
_policy_hdr_d = collections.OrderedDict()
_policy_hdr_d['name'] = dict(w=32, align=_align_wrap, t='Policy Name', v=_simple_val_act, va=_align_wrap)
_policy_hdr_d['is-predefined-policy'] = dict(w=5, align=_align_wrap_vc, t='Predefined', v=_tf_to_check,
                                             va=_align_wrap_c, bt=True)
_policy_hdr_d['is-active-policy'] = dict(w=5, align=_align_wrap_vc, t='Active', v=_tf_to_check, va=_align_wrap_c,
                                         bt=True)
_policy_hdr_d['rule-list'] = dict(w=120, align=_align_wrap, t='Members', v=_rule_csv_list, va=_align_wrap)

_group_hdr_d = collections.OrderedDict()
_group_hdr_d['name'] = dict(w=32, align=_align_wrap, t='Group Name', v=_simple_val_act, va=_align_wrap)
_group_hdr_d['group-type'] = dict(w=12, align=_align_wrap, t='Group Type', v=_simple_val_act, va=_align_wrap)
_group_hdr_d['is-modifiable'] = dict(w=5, align=_align_wrap_vc, t='Modifiable', v=_tf_to_check, va=_align_wrap_c,
                                     bt=True)
_group_hdr_d['is-predefined'] = dict(w=5, align=_align_wrap_vc, t='Predefined', v=_tf_to_check, va=_align_wrap_c,
                                     bt=True)
_group_hdr_d['members'] = dict(w=80, align=_align_wrap, t='Members', v=_member_csv_list, va=_align_wrap)

_rules_hdr_d = collections.OrderedDict()
_rules_hdr_d['name'] = dict(w=40, align=_align_wrap, t='Rule Name', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['group-name'] = dict(w=32, align=_align_wrap, t='Group Name', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['is-predefined'] = dict(w=5, align=_align_wrap_vc, t='Predefined', v=_tf_to_check, va=_align_wrap_c,
                                     bt=True)
_rules_hdr_d['is-rule-on-rule'] = dict(w=5, align=_align_wrap_vc, t='Rule-on-Rule', v=_tf_to_check, va=_align_wrap_c,
                                       bt=True)
_rules_hdr_d['monitoring-system'] = dict(w=20, align=_align_wrap, t='Monitoring System', v=_simple_val_act,
                                         va=_align_wrap)
_rules_hdr_d['logical-operator'] = dict(w=5, align=_align_wrap_vc, t='Operator', v=_simple_val_act, va=_align_wrap_c)
_rules_hdr_d['threshold-value'] = dict(w=18, align=_align_wrap, t='Threshold', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['quiet-time'] = dict(w=8, align=_align_wrap_vc, t='Quiet Time', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['time-base'] = dict(w=12, align=_align_wrap, t='Time Base', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['event-severity'] = dict(w=8, align=_align_wrap_vc, t='Severity', v=_simple_val_act, va=_align_wrap)
_rules_hdr_d['actions'] = dict(w=32, align=_align_wrap, t='Actions', v=_action_csv_list, va=_align_wrap)
_rules_hdr_d['def_policy'] = dict(w=32, align=_align_wrap, t='Default Policies', v=_action_def_policy, va=_align_wrap)
_rules_hdr_d['cust_policy'] = dict(w=32, align=_align_wrap, t='Custom Policies', v=_action_cust_policy, va=_align_wrap)

_r_maps = {'maps-policy': dict(), 'group': dict(), 'rule': dict()}
for _k, _d in _policy_hdr_d.items():
    _r_maps['maps-policy'].update({_d['t']: _k})
for _k, _d in _group_hdr_d.items():
    _r_maps['group'].update({_d['t']: _k})
for _k, _d in _rules_hdr_d.items():
    _r_maps['rule'].update({_d['t']: _k})
_hdr_d = {'maps-policy': _policy_hdr_d, 'group': _group_hdr_d, 'rule': _rules_hdr_d}


def maps_policy(wb, content_d, switch_obj, i, sheet_i=None):
    """Creates a MAPS policy page

    :param wb: Workbook object
    :type wb: class
    :param content_d: Dictionary for table of contents
    :type content_d: dict
    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param i: Unique sheet tab identifier
    :type i: int
    :param sheet_i: Location in workbook where sheet is to be inserted
    :type sheet_i: int, None
    :rtype: None
    """
    global _policy_hdr_d, _hdr1_font, _hdr2_font, _std_font, _align_wrap, _border_thin

    # Create the worksheet and add to table of contents
    sheet_name = excel_util.valid_sheet_name.sub('_', brcddb_switch.best_switch_name(switch_obj))
    sheet_name = sheet_name[0:21] + '_policy_' + str(i) if len(sheet_name) > 21 else sheet_name + '_policy_' + str(i)
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    content_d[switch_obj.r_obj_key()].update(policy=sheet_name)

    # Set the column widths
    row = col = 1
    for d in _policy_hdr_d.values():
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['w']
        col += 1

    # Switch name and column headers
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_policy_hdr_d))
    excel_util.cell_update(sheet,
                           row,
                           col,
                           'MAPS Policies for: ' + brcddb_switch.best_switch_name(switch_obj, wwn=True, fid=True),
                           font=_hdr1_font,
                           align=_align_wrap,
                           border=_border_thin)
    row, col = row+1, 1
    for d in _policy_hdr_d.values():
        excel_util.cell_update(sheet, row, col, d['t'], font=_hdr2_font, align=d['align'], border=_border_thin)
        col += 1
    sheet.freeze_panes = sheet['A3']
    
    # Add the content and track which rule is used in which policy
    for d in gen_util.convert_to_list(switch_obj.r_get(brcdapi_util.maps_policy)):
        row, col = row+1, 1
        for rule in gen_util.convert_to_list(gen_util.get_key_val(d, 'rule-list/rule')):
            k = 'maps_report/rule_d/' + rule + '/'
            policy_l = class_util.get_or_add(switch_obj, k + 'default', list()) if d['is-predefined-policy'] else \
                class_util.get_or_add(switch_obj, k + 'custom', list())
            policy_l.append(d['name'])
        for k, cd in _policy_hdr_d.items():
            excel_util.cell_update(sheet,
                                   row,
                                   col,
                                   cd['v'](switch_obj, d, k),
                                   font=_std_font,
                                   align=cd['va'],
                                   border=_border_thin)
            col += 1


def maps_groups(wb, content_d, switch_obj, i, sheet_i=None):
    """Creates a MAPS groups page

    :param wb: Workbook object
    :type wb: class
    :param content_d: Dictionary for table of contents
    :type content_d: dict
    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param i: Unique sheet tab identifier
    :type i: int
    :param sheet_i: Location in workbook where sheet is to be inserted
    :type sheet_i: int, None
    :rtype: None
    """
    global _group_hdr_d, _hdr1_font, _hdr2_font, _std_font, _align_wrap, _border_thin

    # Create the worksheet
    sheet_name = excel_util.valid_sheet_name.sub('_', brcddb_switch.best_switch_name(switch_obj))
    sheet_name = sheet_name[0:21] + '_group_' + str(i) if len(sheet_name) > 21 else sheet_name + '_group_' + str(i)
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    content_d[switch_obj.r_obj_key()].update(group=sheet_name)

    # Set the column widths
    row = col = 1
    for d in _group_hdr_d.values():
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['w']
        col += 1

    # Switch name and column headers
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_group_hdr_d))
    excel_util.cell_update(sheet,
                           row,
                           col,
                           'MAPS Groups for: ' + brcddb_switch.best_switch_name(switch_obj, wwn=True, fid=True),
                           font=_hdr1_font,
                           align=_align_wrap,
                           border=_border_thin)
    row, col = row+1, 1
    for d in _group_hdr_d.values():
        excel_util.cell_update(sheet, row, col, d['t'], font=_hdr2_font, align=d['align'], border=_border_thin)
        col += 1
    sheet.freeze_panes = sheet['A3']

    # Add the content
    for d in gen_util.convert_to_list(switch_obj.r_get(brcdapi_util.maps_group)):
        row, col = row+1, 1
        for k, cd in _group_hdr_d.items():
            excel_util.cell_update(sheet,
                                   row,
                                   col,
                                   cd['v'](switch_obj, d, k),
                                   font=_std_font,
                                   align=cd['va'],
                                   border=_border_thin)
            col += 1


def maps_rules(wb, content_d, switch_obj, i, sheet_i=None):
    """Creates a MAPS rules page

    :param wb: Workbook object
    :type wb: class
    :param content_d: Dictionary for table of contents
    :type content_d: dict
    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param i: Unique sheet tab identifier
    :type i: int
    :param sheet_i: Location in workbook where sheet is to be inserted
    :type sheet_i: int, None
    :rtype: None
    """
    global _rules_hdr_d, _hdr1_font, _hdr2_font, _std_font, _align_wrap, _border_thin

    # Create the worksheet
    sheet_name = excel_util.valid_sheet_name.sub('_', brcddb_switch.best_switch_name(switch_obj))
    sheet_name = sheet_name[0:21] + '_rule_' + str(i) if len(sheet_name) > 21 else sheet_name + '_rule_' + str(i)
    sheet = wb.create_sheet(index=0 if sheet_i is None else sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    content_d[switch_obj.r_obj_key()].update(rule=sheet_name)

    # Set the column widths
    row = col = 1
    for d in _rules_hdr_d.values():
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['w']
        col += 1

    # Switch name and column headers
    col = 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(_rules_hdr_d))
    excel_util.cell_update(sheet,
                           row,
                           col,
                           'MAPS Rules for: ' + brcddb_switch.best_switch_name(switch_obj, wwn=True, fid=True),
                           font=_hdr1_font,
                           align=_align_wrap,
                           border=_border_thin)
    row, col = row+1, 1
    for d in _rules_hdr_d.values():
        excel_util.cell_update(sheet, row, col, d['t'], font=_hdr2_font, align=d['align'], border=_border_thin)
        col += 1
    sheet.freeze_panes = sheet['A3']

    # Add the content
    for d in gen_util.convert_to_list(switch_obj.r_get(brcdapi_util.maps_rule)):
        row, col = row+1, 1
        for k, cd in _rules_hdr_d.items():
            excel_util.cell_update(sheet,
                                   row,
                                   col,
                                   cd['v'](switch_obj, d, k),
                                   font=_std_font,
                                   align=cd['va'],
                                   border=_border_thin)
            col += 1


def read_maps(proj_obj, maps_file, non_def_only=False, echo=False):
    """Reads a MAPS workbook (same format as output from maps_report.py) and adds 'brocade-maps' to each LS

    :param proj_obj: Project object
    :type proj_obj: brcddb.classes.project.ProjectObj
    :param maps_file: Name of workbook to read
    :type maps_file: str
    :param non_def_only: If True, filter out all the default policies
    :type non_def_only: bool
    :param echo: If True, echo activity to STD_OUT
    :type echo: bool
    :return error_l: List of error encountered.
    :rtype error_l: list
    """
    global _r_maps, _rule_key_d, _hdr_d

    # Read the workbook
    brcdapi_log.log('Reading maps workbook: ' + maps_file, echo=echo)
    error_l, sheet_l = excel_util.read_workbook(maps_file, dm=3, order='col', skip_sheets='contents')
    if len(error_l) > 0:
        return error_l

    # Add brocade-maps to each switch object
    for sheet_d in sheet_l:

        # Figure out where the headers are, if it's a policy, group, or rule sheet, and add the switch object
        al = sheet_d['al']
        if len(al) < 3:
            continue
        temp_l = al[0][0].split(' ')
        wwn, fid, name = temp_l[6].replace('(', '').replace(')', ''), int(temp_l[5]), temp_l[3]
        switch_obj = proj_obj.s_add_switch(wwn)
        switch_obj.s_new_key('brocade-fabric',
                             {'fabric-switch': {'name': wwn, 'switch-user-friendly-name': name}})
        switch_obj.s_new_key('brocade-fibrechannel-switch',
                             {'fibrechannel-switch': {'name': wwn, 'user-friendly-name': name, 'vf-id': fid}})
        switch_obj.s_new_key('brocade-fibrechannel-logical-switch',
                             {'fibrechannel-logical-switch': {'switch-wwn': wwn, 'fabric-id': fid}})
        header_d = excel_util.find_headers(al[1])  # Column number for each header
        policy_key = _rule_key_d.get(' '.join(al[0][0].split(' ')[0:2]))
        if policy_key is None:
            error_l.append('Invalid sheet: ' + sheet_d['sheet'])
            continue

        # Add the content to the switch object
        reverse_d = _r_maps[policy_key]
        policy_l = class_util.get_or_add(switch_obj, 'brocade-maps/' + policy_key, list())
        for row_l in al[2:]:
            try:
                d = dict()
                for col_hdr, item_key in reverse_d.items():
                    v = row_l[header_d[col_hdr]]
                    if _hdr_d[policy_key][item_key].get('bt', False):
                        v = True if isinstance(v, str) and len(v) > 0 else False
                        if non_def_only and v and 'is-predefined' in item_key:
                            raise Skip
                    d.update({item_key: v})
                policy_l.append(d)
            except Skip:
                pass

    return error_l
