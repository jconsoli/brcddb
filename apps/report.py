"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

:mod:`brcddb.apps.report` - Creates a report in Excel Workbook format from a brcddb project

**Public Methods**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| report                | Creates an Excel report. Sort of a SAN Health like report.                                |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 20 Oct 2024   | Updated call to switch_page() to match new inputs.                                    |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 06 Dec 2024   | Added "About" page.                                                                   |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 26 Dec 2024   | Removed unused parameter in call to chassis_hidden_port_page()                        |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 01 Mar 2025   | Added zone configuration cleanup page.                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024, 2025 Consoli Solutions, LLC'
__date__ = '01 Mar 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.5'

import os
import collections
import copy
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcddb.util.util as brcddb_util
import brcddb.app_data.report_tables as rt
import brcddb.brcddb_chassis as brcddb_chassis
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_project as brcddb_project
import brcddb.brcddb_switch as brcddb_switch
import brcddb.report.bp as report_bp
import brcddb.report.chassis as report_chassis
import brcddb.report.fabric as report_fabric
import brcddb.report.iocp as report_iocp
import brcddb.report.login as report_login
import brcddb.report.port as report_port
import brcddb.report.switch as report_switch
import brcddb.report.zone as report_zone
import brcddb.report.utils as report_utils
import brcddb.util.search as brcddb_search
import brcddb.classes.util as brcddb_class_util
import brcddb.classes.alert as alert_class
import brcddb.util.obj_convert as obj_convert

"""
References are added to most objects so that when creating the Workbook so that hyperlinks can be added. For example,
from the fabric summary page the reader of the report can simply click on the switch name and be taken directly to the
switch detail. For this reason, sheet names for each major object are created in advance. The sheets for subordinate
objects are then created. For example, for each fabric the port configuration sheets are created before the switch
sheet. Since the switch sheet name was already defined, links back to the parent switch can be added to the port
configuration sheet. Then when creating the switch detail sheets, the links for the individual ports can be added to the
switch detail sheets.

An object with key 'report_app' is added to each object in the report. The report_app object is a dictionary of
dictionaries defined as follows:

+-----------+-------+-----------+-----------------------------------------------------------------------------------+
| Key       | Key   | Object(s) | Description                                                                       |
+===========+=======+===========+===================================================================================+
| control   |       |           | Used for determining naming for determining names used in 'link'. Also used to    |
|           |       |           | determine how information is displayed. Search for _xxx_control_d for details.    |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | bp    | Project,  | Best practice                                                                     |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | dup   | Project   | Duplicate WWNs                                                                    |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | fab   | Fabric    | Fabric summary sheet                                                              |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | db    | Project   | Dashboard.                                                                        |
|           |       | Fabric    |                                                                                   |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | sd    | Fabric    | Switch detail                                                                     |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pc    | Fabric    | Port configuration                                                                |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | ps    | Fabric    | Port statistics                                                                   |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pz    | Fabric    | Ports by Zone and Login                                                           |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pr    | Fabric    | Port RNID Data                                                                    |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | sfp   | Fabric    | SFP report                                                                        |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | za    | Fabric    | Zone Analysis                                                                     |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | zt    | Fabric    | Zone by Target                                                                    |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | znt   | Fabric    | Zone by Non-Targets                                                               |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | ali   | Fabric    | Alias Detail                                                                      |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | log   | Fabric    | Logins                                                                            |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
| link      |       |           | Cell references. These are used to create hyperlinks in the Workbook.             |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | tc    | All       | A link to the table of context.                                                   |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | bp    | Project,  | Best practice                                                                     |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | dup   | Project   | Duplicate WWNs                                                                    |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | fab   | Fabric    | Fabric summary sheet                                                              |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | db    | Project   | Dashboard.                                                                        |
|           |       | Fabric    |                                                                                   |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | sw    | Switch    | Sheet reference: switch detail sheet. Cell reference: A1                          |
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pc    | Port      | Link to where this port appears on the Port Configurations worksheet.             |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pl    | Port      | Link to where this port appears on the Port Links worksheet.                      |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | ps    | Port      | Link to where this port appears on the Port Statistics worksheet.                 |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pz    | Port      | Link to where this port appears on the Ports by Zone and Login worksheet.         |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | pr    | Port      | Link to where this port appears on the Port RNID worksheet.                       |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
|           | sfp   | Port      | Link to where this port appears on the SFP Report worksheet.                      |
|           |       | Switch    | Column link is 'A'. Row is for the first port in the switch.                      |  
+-----------+-------+-----------+-----------------------------------------------------------------------------------+
"""
_MAX_DB_SIZE = 10  # Set the top xx dashboard size
_unique_index = 0  # The openpyxl library appends a number if necessary to make worksheet names unique; however, this
# module creates all worksheet names in advance so that links to them can be added before the worksheet has been
# created. _unique_index therefore is used to ensure all worksheet names are unique before they are created.
_dup_login_tbl = (
    '_LOGIN_COMMENTS',
    brcdapi_util.bns_port_name,
    '_ALIAS',
    '_FABRIC_NAME',
    '_SWITCH_NAME',
    '_PORT_NUMBER',
    brcdapi_util.bns_port_id,
    '_ZONES_DEF',
    brcdapi_util.bns_fc4_features,
    brcdapi_util.bns_link_speed,
    brcdapi_util.bns_ns_dev_type,
    brcdapi_util.bns_node_name,
    brcdapi_util.bns_node_symbol,
    brcdapi_util.bns_port_symbol,
)
_port_links_tbl = (
    '_PORT_COMMENTS',
    '_SWITCH_LINK',
    '_PORT_NUMBER',
    '_ALIAS',
    '_BEST_DESC',
    '_CONFIG_LINK',
    '_STATS_LINK',
    '_ZONE_LINK',
    '_SFP_LINK',
    '_RNID_LINK',
)

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_link_font = excel_fonts.font_type('link')
_hdr2_font = excel_fonts.font_type('hdr_2')
_hdr1_font = excel_fonts.font_type('hdr_1')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')

_about_buf_l = (
    'The contents of this Workbook were generated from a script, report.py, that reads a data collection file and '
    'generates a general SAN report.',
    '',
    'The intended uses are:',
    '  *   Determine where connections are made',
    '  *   Zone analysis',
    '  *   Checking for errors and best practice violations',
    '  *   Analyzing enclosure usage (typically storage arrays)',
)

_zc_comment = 'Zone cleanup for $fab. Copy this zone cleanup worksheet to a zone configuration workbook. See '\
              '"zone_config_sample.xlsx" and zone_config.py.'


def _dashboard(obj, wb, sheet_index):
    """Adds a port performance dashboard. See _add_fabric_summary() for parameter definitions
    """
    global _MAX_DB_SIZE

    control_d = obj.r_get('report_app/control/db')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)

    dashboard_item = collections.OrderedDict()
    dashboard_item['bad-eofs-received'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' Bad EOF', port_list=list())
    dashboard_item['class-3-discards'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' C3 Discards', port_list=list())
    dashboard_item['in-crc-errors'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' CRC With Good EOF', port_list=list())
    dashboard_item['crc-errors'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' CRC', port_list=list())
    dashboard_item['loss-of-signal'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' Loss of Signal', port_list=list())
    dashboard_item['bb-credit-zero'] = dict(title='Top ' + str(_MAX_DB_SIZE) + ' BB Credit Zero', port_list=list())
    port_list = obj.r_port_objects()
    for k, db in dashboard_item.items():
        db_list = brcddb_search.test_threshold(port_list, 'fibrechannel-statistics/' + k, '>', 0)
        db_list = gen_util.sort_obj_num(db_list, 'fibrechannel-statistics/' + k, True)
        if len(db_list) > _MAX_DB_SIZE:
            del db_list[_MAX_DB_SIZE:]
        db['port_list'].extend(db_list)
    report_port.performance_dashboard(wb,
                                      obj.r_get('report_app/hyperlink/tc'),
                                      control_d['sn'],
                                      sheet_index,
                                      control_d['t'],
                                      dashboard_item)
    return 1


# Actions in _xxx_control_d
# Fabrics
def _add_fabric_summary(fab_obj, wb, sheet_index):
    """ Adds the fabric summary sheet

    :param fab_obj: Fabric object
    :type fab_obj: brcddb.classes.fabric.FabricObj
    :param wb: Workbook object
    :type wb: openpyxl.Workbook
    :param sheet_index: Location for this sheet.
    :type sheet_index: int
    :param sheet_title: Title to add to the top of the worksheet
    :type sheet_title: str
    :return: Number of sheets added
    :rtype: int
    """
    control_d = fab_obj.r_get('report_app/control/fab')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_fabric.fabric_page(wb,
                              fab_obj.r_get('report_app/hyperlink/tc'),
                              sheet_index,
                              control_d['sn'],
                              control_d['t'],
                              fab_obj)
    return 1


def _add_switch_page(switch_obj, wb, sheet_index):
    """Adds the switch page. See _add_fabric_summary() for parameter definitions"""
    control_d = switch_obj.r_get('report_app/control/switch')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_switch.switch_page(wb,
                              switch_obj.r_get('report_app/hyperlink/tc'),
                              control_d['sn'],
                              sheet_index,
                              control_d['t'],
                              switch_obj,
                              in_display=rt.Switch.switch_display_tbl)
    return 1


def _add_switch_ports(switch_obj, wb, sheet_index):
    """Adds the port map. See _add_fabric_summary() for parameter definitions"""
    report_switch.add_port_map(switch_obj)

    return 0


def _add_port_page(fab_obj, wb, sheet_index, control_d, config_tbl, login_flag, link_type):
    """Adds the port detail pages. See _add_fabric_summary()

    :param control_d: Control data structure. Search for _xxx_control_d for details.
    :type control_d: dict
    :param config_tbl: Table to use to determine what items to add to the report
    :type config_tbl: list, tuple, None
    :param login_flag: If True, add the NPIV logins to the report
    :type login_flag: bool
    :param link_type: Type of link, see brcddb.apps.report. Valid types are: pc, pl, ps, pz, pr, and sfp
    :type link_type: str, None
    """
    control_d = fab_obj.r_get('report_app/control/'+link_type)
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_port.port_page(wb,
                          fab_obj.r_get('report_app/hyperlink/tc'),
                          control_d['sn'],
                          sheet_index,
                          control_d['t'],
                          brcddb_util.sort_ports(fab_obj.r_all_port_objects()),
                          in_display=config_tbl,
                          in_port_display_tbl=rt.Port.port_display_tbl,
                          login_flag=login_flag,
                          link_type=link_type)
    return 1


def _add_port_config(fab_obj, wb, sheet_index):
    """Adds the port configuration page. See _add_fabric_summary() for parameter definitions"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/pc'), rt.Port.port_config_tbl, False,
                   'pc')
    return 1


def _add_port_links(fab_obj, wb, sheet_index):
    """Adds the port configuration page. See _add_fabric_summary() for parameter definitions"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/pl'), _port_links_tbl, False, 'pl')
    return 1


def _add_port_stats(fab_obj, wb, sheet_index):
    """Adds the port statistics page. See _add_fabric_summary() for parameter definitions"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/ps'), rt.Port.port_stats_tbl, False,
                   'ps')
    return 1


def _add_port_login(fab_obj, wb, sheet_index):
    """Adds the port by zone and login page. See _add_fabric_summary() for parameter definitions"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/pz'), rt.Port.port_zone_tbl, True, 'pz')
    return 1


def _add_port_rnid(fab_obj, wb, sheet_index):
    """Adds the port RNID page. See _add_fabric_summary() for parameter definitions
    $ToDo: only add RNID sheet if RNID data is present"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/pr'), rt.Port.port_rnid_tbl, False, 'pr')
    return 1


def _add_port_sfp(fab_obj, wb, sheet_index):
    """Adds the port SFP report page. See _add_fabric_summary() for parameter definitions"""
    _add_port_page(fab_obj, wb, sheet_index, fab_obj.r_get('report_app/control/sfp'), rt.Port.port_sfp_tbl, False,
                   'sfp')
    return 1


def _add_zone_analysis(fab_obj, wb, sheet_index):
    """Adds the Zone Analysis page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/za')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.zone_page(
        fab_obj,
        fab_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])
    return 1


def _add_zone_by_target(fab_obj, wb, sheet_index):
    """Adds the Zone by Target page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/zt')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.target_zone_page(
        fab_obj,
        fab_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])
    return 1


def _add_zone_by_non_target(fab_obj, wb, sheet_index):
    """Adds the Zone by Non-Target page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/znt')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.non_target_zone_page(
        fab_obj,
        fab_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])
    return 1


def _add_zone_clean(fab_obj, wb, sheet_index):
    """Adds the zone cleanup page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/zc')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.zone_clean_page(
        fab_obj,
        fab_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])

    return 1


def _add_alias_detail(fab_obj, wb, sheet_index):
    """Adds the Alias Detail page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/ali')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.alias_page(
        fab_obj,
        fab_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])
    return 1


def _add_login_detail(fab_obj, wb, sheet_index):
    """Adds the Logins page. See _add_fabric_summary() for parameter definitions"""
    control_d = fab_obj.r_get('report_app/control/log')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_login.login_page(
        wb,
        fab_obj.r_get('report_app/hyperlink/tc'),
        control_d['sn'],
        sheet_index,
        control_d['t'],
        fab_obj.r_login_objects())
    return 1


# Chassis
def _add_chassis(chassis_obj, wb, sheet_index):
    """Adds the chassis page. See _add_fabric_summary() for parameter definitions"""
    control_d = chassis_obj.r_get('report_app/control/chassis')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_chassis.chassis_page(
        wb,
        chassis_obj.r_get('report_app/hyperlink/tc'),
        control_d['sn'],
        sheet_index,
        chassis_obj)
    return 1


# Chassis
def _add_hidden_chassis_port_page(chassis_obj, wb, sheet_index):
    """Adds the hidden chassis port page. See _add_fabric_summary() for parameter definitions"""
    control_d = chassis_obj.r_get('report_app/control/chassis_p')
    report_chassis.chassis_hidden_port_page(wb, sheet_index, chassis_obj)
    return 1


def _add_chassis_ports(chassis_obj, wb, sheet_index):
    """Adds the port map. See _add_fabric_summary() for parameter definitions"""
    report_chassis.add_port_map(chassis_obj)

    return 0


# Project
def _add_project_bp(proj_obj, wb, sheet_index):
    """Adds the best practice page. See _add_fabric_summary() for parameter definitions"""
    control_d = proj_obj.r_get('report_app/control/bp')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_bp.bp_page(
        wb,
        proj_obj.r_get('report_app/hyperlink/tc'),
        control_d['sn'],
        sheet_index,
        control_d['t'],
        proj_obj,
        rt.BestPractice.bp_tbl,
        rt.BestPractice.bp_display_tbl)
    return 1


def _add_project_dup(proj_obj, wb, sheet_index):
    """Checks for duplicate WWNs and adds the page if necessary. See _add_fabric_summary() for parameter definitions"""
    global _dup_login_tbl

    control_d = proj_obj.r_get('report_app/control/dup')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    wl = brcddb_project.dup_wwn(proj_obj)
    if len(wl) > 0:
        report_login.login_page(
            wb,
            proj_obj.r_get('report_app/hyperlink/tc'),
            control_d['sn'],
            sheet_index,
            control_d['t'],
            wl,
            _dup_login_tbl,
            rt.Login.login_display_tbl,
            False)
        return 1
    return 0


def _add_zone_by_group(proj_obj, wb, sheet_index):
    """Adds the zone by groups page. See _add_fabric_summary() for parameter definitions"""
    if len(proj_obj.r_get('report_app/group_d', dict())) == 0:
        return 0

    control_d = proj_obj.r_get('report_app/control/zg')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_zone.group_zone_page(
        proj_obj,
        proj_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])

    return 1


def _add_about(proj_obj, wb, sheet_index):
    """Adds the zone by groups page. See _add_fabric_summary() for parameter definitions"""
    global __version__, _about_buf_l

    control_d = proj_obj.r_get('report_app/control/ab')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_utils.about_page(
        wb,
        sheet_index,
        control_d['sn'],
        os.path.basename(__file__),
        __version__,
        _about_buf_l,
        tc=proj_obj.r_get('report_app/hyperlink/tc'))

    return 1


def _add_project_tc(proj_obj, wb, sheet_index):
    """Adds the "Table of Contents" page. See _add_fabric_summary() for parameter definitions"""
    global _hdr1_font, _std_font, _align_wrap, _link_font, _hdr2_font

    # Add and set up the worksheet
    control_d = proj_obj.r_get('report_app/control')
    hyper_d = proj_obj.r_get('report_app/hyperlink')
    sheet = wb.create_sheet(index=sheet_index, title=control_d['tc']['sn'])
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    row = col = 1
    for i in (4, 20, 62):
        sheet.column_dimensions[xl.get_column_letter(col)].width = i
        col += 1

    # Add the title
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col-1)
    col = 1
    excel_util.cell_update(sheet, row, col, control_d['tc']['t'], font=_hdr1_font, align=_align_wrap,
                           fill=excel_fonts.fill_type('lightblue'))

    # Add the project description and date
    row += 2
    for d in (dict(t='Description', c=proj_obj.c_description(), f=excel_fonts.font_type('cli')),
              dict(t='Data collected', c=proj_obj.r_date(), f=_std_font)):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        excel_util.cell_update(sheet, row, 1, d['t'], font=_std_font, align=_align_wrap)
        excel_util.cell_update(sheet, row, 3, d['c'], font=d['f'], align=_align_wrap)
        row += 1

    # Figure out what to add to the table of contents. Start with basic project stuff & chassis
    # control_d and hyper_d are set to the 'report_app/control' and 'report_app/hyperlink' of the project object.
    # contents_l is a list of dictionaries. 't' is the section title for the table of contents and cl is a list of
    # dictionaries that describes the subsections. In cl, 't' is the subsection title and 'l' is the link
    dup_d = dict(t=control_d['dup']['tc'], l=hyper_d['dup']) if control_d['dup']['sn'] in wb.sheetnames \
        else dict(t='No Duplicate WWNs Found')

    contents_l = [
        dict(t='About', cl=[dict(t=control_d['ab']['tc'], l=hyper_d['ab'])]),
        dict(t='Zone Groups', zg=True, cl=[dict(t=control_d['zg']['tc'], l=hyper_d['zg'])]),
        dict(t='Best Practice Violations', cl=[dict(t=control_d['bp']['tc'], l=hyper_d['bp'])]),
        dict(t='Duplicate WWNs', cl=[dup_d]),
        dict(t=control_d['db']['tc'], cl=[dict(t=control_d['db']['tc'], l=hyper_d['db'])]),
        dict(t='Chassis',
             cl=[dict(t=brcddb_chassis.best_chassis_name(obj), l=obj.r_get('report_app/hyperlink/chassis'))
                 for obj in proj_obj.r_chassis_objects()]),
    ]

    # Add the fabrics. See notes above with how table of contents was built for the project
    for fab_obj in proj_obj.r_fabric_objects():
        control_d = fab_obj.r_get('report_app/control')
        hyper_d = fab_obj.r_get('report_app/hyperlink')
        sub_content_l = [dict(t=control_d[key]['tc'], l=hyper_d[key]) for key in ('fab', 'db')]
        sub_content_l.extend(
            [dict(t=control_d[key]['tc'], l=hyper_d[key]) for key in ('pl', 'za', 'zt', 'znt', 'zc', 'ali', 'log')]
        )
        contents_l.append(dict(t=brcddb_fabric.best_fab_name(fab_obj, wwn=True, fid=True), cl=sub_content_l))

    # Add the IOCPs
    temp_l = proj_obj.r_iocp_objects()
    if len(temp_l) > 0:
        cl = [dict(t=obj.r_get('report_app/control/iocp/tc'), l=obj.r_get('report_app/hyperlink/iocp')) for
              obj in proj_obj.r_iocp_objects()]
        contents_l.append(dict(t='IOCPs',
                               cl=[dict(t=obj.r_get('report_app/control/iocp/t'),
                                        l=obj.r_get('report_app/hyperlink/iocp'))
                                   for obj in proj_obj.r_iocp_objects()]))

    # Add all the Table of Contents items
    for d in contents_l:
        row += 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        excel_util.cell_update(sheet, row, 1, d['t'], font=_hdr2_font, align=_align_wrap)
        row += 1
        for obj in d['cl']:
            if bool(d.get('zg')) and len(proj_obj.r_get('report_app/group_d')) == 0:
                continue
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            excel_util.cell_update(sheet, row, 2, obj['t'], font=_link_font if 'l' in obj else _std_font,
                                   align=_align_wrap, link=obj.get('l'))
            row += 1

    # Add chassis not polled
    temp_l = [obj for obj in proj_obj.r_chassis_objects() if obj.r_get('brocade-chassis') is None]
    if len(temp_l) > 0:
        row += 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        excel_util.cell_update(sheet, row, 1, 'Missing chassis (discovered in fabrics but not polled)',
                               font=_hdr2_font, align=_align_wrap)
        row += 1
    for obj in temp_l:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        buf = brcddb_chassis.best_chassis_name(obj, wwn=True) + '. Known logical switches:'
        excel_util.cell_update(sheet, row, 2, buf, font=_std_font, align=_align_wrap)
        row += 1
        for switch_obj in obj.r_switch_objects():
            excel_util.cell_update(sheet, row, 3, brcddb_switch.best_switch_name(switch_obj, wwn=True),
                                   font=_std_font, align=_align_wrap)
            row += 1

    return 1


# IOCP
def _add_iocp(iocp_obj, wb, sheet_index):
    """Adds the IOCP pages, if any. See _add_fabric_summary() for parameter definitions"""
    control_d = iocp_obj.r_get('report_app/control/iocp')
    brcdapi_log.log('    Adding ' + control_d['sn'], echo=True)
    report_iocp.iocp_page(
        iocp_obj,
        iocp_obj.r_get('report_app/hyperlink/tc'),
        wb,
        control_d['sn'],
        sheet_index,
        control_d['t'])
    return 1


"""" _xxx_control_d dictionaries are used to build the content for report_app. Some of the content is filled in by the
action methods.

+-------+-----------+-----------------------------------------------------------------------------------------------+
| Key   | Type      | Description                                                                                   |
+=======+===========+===============================================================================================+
| a     | method    | Pointer to action method that creates the worksheet                                           |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| p     | str       | Sheet name prefix. A prefix to the value specified in sn.                                     |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| rs    | bool      | If True, reset to the sheet_index to 0. Default is False                                      |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| s     | bool      | If True, add the object name to the sheet name. Used in _add_sheet_names() to determine 'sn'  |
|       |           | Default is False.                                                                             |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| sn    | str       | Sheet name. Filled in by _add_sheet_names(). If None, no sheet is added. None is useful for   |
|       |           | wrap up processing not specific to adding sheets.                                             |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| sf    | bool      | True: Use starting index. False: Use current index                                            |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| st    | bool      | If True, the object name is added to the sheet title, 't', in _add_sheet_names(). Default is  |
|       |           | False.                                                                                        |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| t     | str, None | Sheet title. The object name is typically appended. See 'st'. Not used if sn is None.         |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| tc    | str       | Name for this page to put in the table of contents. if None, sheet is not added to the table  |
|       |           | of contents.                                                                                  |
+-------+-----------+-----------------------------------------------------------------------------------------------+
| u     | bool      | If True, add a unique number. Used to guarantee unique sheet names. Default is False          |
+-------+-----------+-----------------------------------------------------------------------------------------------+
"""
_tc = dict(p='table_of_contents', t='Table of Contents', u=False, s=False, sf=True, a=_add_project_tc)
_proj_control_d = dict(
    tc=_tc,
    ab=dict(p='About', tc='About', t=None, sf=True, a=_add_about),
    db=dict(p='proj_dashboard', tc='Project Dashboard', t='Project Dashboard', a=_dashboard),
    bp=dict(p='Best_Practice', tc='Best Practice Violations', t='Best Practice Violations', a=_add_project_bp),
    dup=dict(p='dup_wwns', tc='Duplicate WWNs', t='Duplicate WWNs', a=_add_project_dup),
    zg=dict(p='Zone_Groups', tc='Zone Groups', t='Zone Groups', a=_add_zone_by_group),
)
_fab_control_d = dict(
    tc=_tc,
    fab=dict(p='fab_', tc='Fabric Summary', t='Fabric Summary: ', u=True, s=True, sf=True, a=_add_fabric_summary),
    db=dict(p='db_', tc='Fabric Dashboard', t='Fabric Dashboard: ', u=True, s=True, sf=True, a=_dashboard),
    pc=dict(p='pc_', tc='Port Configurations', t='Port Configurations For Fabric: ', u=True, s=True,
            a=_add_port_config),
    pl=dict(p='pl_', tc='Port Page Links', t='Port Links For Fabric: ', u=True, s=True, sf=True, a=_add_port_links),
    ps=dict(p='ps_', tc='Port Statistics', t='Port Statistics For Fabric: ', u=True, s=True, a=_add_port_stats),
    pz=dict(p='pz_', tc='Ports by Zone and Login', t='Ports by Zone and Login For Fabric: ', u=True, s=True,
            a=_add_port_login),
    pr=dict(p='pr_', tc='Port RNID Data', t='Port RNID Data For Fabric: ', u=True, s=True, a=_add_port_rnid),
    sfp=dict(p='sfp_', tc='SFP Report', t='SFP Report For Fabric: ', u=True, s=True, a=_add_port_sfp),
    za=dict(p='za_', tc='Zone Analysis', t='Zone Analysis For Fabric: ', u=True, s=True, a=_add_zone_analysis),
    zt=dict(p='zt_', tc='Zone by Target', t='Zone by Target For Fabric: ', u=True, s=True,  a=_add_zone_by_target),
    znt=dict(p='znt_', tc='Zone by Non-Targets', t='Zone by Non-Targets For Fabric: ', u=True, s=True,
             a=_add_zone_by_non_target),
    zc=dict(p='zc_', tc='Zone Clean', t='', u=True, s=True, a=_add_zone_clean),
    ali=dict(p='ali_', tc='Alias Detail', t='Alias Detail For Fabric: ', u=True, s=True, a=_add_alias_detail),
    log=dict(p='log_', tc='Logins', t='Logins For Fabric: ', u=True, s=True, a=_add_login_detail)
)
_chassis_control_d = dict(
    tc=_tc,
    chassis=dict(tc='Chassis', t='Chassis Detail For: ', u=True, s=True, a=_add_chassis),
    chassis_p=dict(tc=None, t='Chassis Hidden Port Page: ', u=True, s=True, a=_add_hidden_chassis_port_page),
)
_chassis_ports_control_d = dict(
    chassis=dict(a=_add_chassis_ports)
)
_iocp_control_d = dict(
    tc=_tc,
    iocp=dict(p='iocp_', tc='IOCP', t='IOCP Detail For: ', s=True, a=_add_iocp),
)
_switch_control_d = dict(
    tc=_tc,
    switch=dict(t='Switch Detail: ', u=True, s=True, a=_add_switch_page)
)
_switch_ports_control_d = dict(
    switch=dict(a=_add_switch_ports)
)


def _add_sheet_names(proj_obj):
    """Adds sheet names to the major objects so that pages can be created with links to pages not yet created.

    :param proj_obj: Project object
    :type proj_obj: brcddb.classes.project.ProjectObj
    :rtype: None
    """
    global _proj_control_d, _chassis_control_d, _fab_control_d, _switch_control_d, _unique_index

    # Set up the control data structures.
    add_l = (
        dict(obj_l=[proj_obj],  # List of objects to create worksheets for.
             control_d=_proj_control_d,  # Data structure that controls how to set up the worksheet.
             name_m=''),  # Sheet title and sheet name prefix or pointer to function to supply the same.
        dict(obj_l=proj_obj.r_chassis_objects(),
             control_d=_chassis_control_d,
             name_m=brcddb_chassis.best_chassis_name),
        dict(obj_l=proj_obj.r_switch_objects(),
             control_d=_switch_control_d,
             name_m=brcddb_switch.best_switch_name),
        dict(obj_l=proj_obj.r_fabric_objects(),
             control_d=_fab_control_d,
             name_m=brcddb_fabric.best_fab_name),
        dict(obj_l=proj_obj.r_iocp_objects(),
             control_d=_iocp_control_d,
             name_m=_iocp_name),
    )

    # Create unique sheet names and add to the associated objects
    for obj_d in add_l:
        for obj in obj_d['obj_l']:
            control_d = copy.deepcopy(obj_d['control_d'])
            for d in control_d.values():
                d.pop('a', None)  # Remove the pointer to the action method
            if callable(obj_d['name_m']):
                sname_base = obj_d['name_m'](obj).replace(':', '')[0:25].replace(' ', '_')
                sname_base = excel_util.valid_sheet_name.sub('', sname_base)
            else:
                sname_base = obj_d['name_m']
            sname_d = dict()
            for k, d in control_d.items():
                prefix = d.get('p', '')
                sname = excel_util.valid_sheet_name.sub('_', prefix if k == 'tc' else prefix + sname_base)
                if d.get('u', False):
                    if len(sname) > 26:
                        sname = sname[0:27]
                    _unique_index += 1
                    sname += '_' + str(_unique_index)
                sname = gen_util.remove_duplicate_char(sname, '_')
                temp_d = dict(sn=sname, t=obj_d['name_m']) if d.get('st', False) else dict(sn=sname)
                d.update(temp_d)
                sname_d.update({k: '#' + sname + '!A1'})
            brcddb_util.add_to_obj(obj, 'report_app/control', control_d)
            brcddb_util.add_to_obj(obj, 'report_app/hyperlink', sname_d)

    return


# Object names. See obj_name in report_l in module report()
def _fabric_name(obj):
    return brcddb_fabric.best_fab_name(obj, wwn=True, fid=True)


def _chassis_name(obj):
    return brcddb_chassis.best_chassis_name(obj, wwn=True)


def _project_name(obj):
    try:
        for buf in obj.r_description().split('\n'):
            if 'In file, -i:' in buf:
                return buf.split(':')[1].strip()
    except IndexError:
        pass
    return obj.r_description()


def _iocp_name(obj):
    return obj.r_obj_key()


def _switch_name(obj):
    return brcddb_switch.best_switch_name(obj, wwn=True, did=True, fid=True)


def report(proj_obj, outf, group_d=None):
    """Creates an Excel report. Sort of a SAN Health like report.

    :param proj_obj: Project object
    :type proj_obj: brcddb.classes.project.ProjectObj
    :param outf: Output file name
    :type outf: str
    :param group_d: Zone groups. Key: Group name. Value: list of port objects
    :type group_d: None, dict
    """
    global _fab_control_d

    # Set up the workbook and give all the major objects (Project, Chassis, Fabric, and Switch) sheet names
    sheet_index, wb, working_group_d = 0, excel_util.new_report(), dict() if group_d is None else group_d
    _add_sheet_names(proj_obj)
    brcddb_util.add_to_obj(proj_obj, 'report_app/group_d', working_group_d)

    """report_l is a list of dictionaries in the order they are to be processed. The dictionaries control sheet creation
    as follows:
    
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | Key       | Type      | Description                                                                           |
    +===========+=======+===========================================================================================+
    | obj_l     | list      | Objects for which sheets are to be added                                              |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | add_name  | tuple     | Key into _xxx_control_d tables for sheet titles the object name should be added to.   |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | order     | tuple     | Key into _xxx_control_d tables in the order sheets are to be added for this object    |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | feedback  | str       | Prefix to add to object name for user message to display on STD_OUT when processing   |
    |           |           | this object                                                                           |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | control   | dict      | Control dictionary. Search for _xxx_control_d for an additional description.          |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | obj_name  | method    | Pointer to method that retrieves the object name.                                     |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    | rs        | bool      | If True, reset the sheet_index. Default is False.                                     |
    +-----------+-----------+---------------------------------------------------------------------------------------+
    """
    report_l = (
        dict(obj_l=proj_obj.r_fabric_objects(),
             add_name=('pc', 'ps', 'pz', 'pr', 'sfp', 'pl', 'za', 'zt', 'znt', 'zc', 'ali', 'log', 'db', 'fab'),
             order=('pc', 'ps', 'pz', 'pr', 'sfp', 'pl', 'za', 'zt', 'znt', 'zc', 'ali', 'log', 'db', 'fab'),
             feedback='Processing fabric: ',
             control=_fab_control_d,
             obj_name=_fabric_name),
        dict(obj_l=proj_obj.r_iocp_objects(),
             add_name=('iocp',),
             order=('iocp',),
             feedback='Processing IOCP: ',
             control=_iocp_control_d,
             obj_name=_iocp_name),
        dict(obj_l=proj_obj.r_switch_objects(),
             add_name=('switch',),
             order=('switch',),
             feedback='Processing switch: ',
             control=_switch_control_d,
             obj_name=_switch_name,
             rs=True),
        dict(obj_l=proj_obj.r_chassis_objects(),
             add_name=('chassis', 'chassis_p'),
             order=('chassis', 'chassis_p'),
             feedback='Processing chassis: ',
             control=_chassis_control_d,
             obj_name=_chassis_name,
             rs=True),
        dict(obj_l=proj_obj.r_chassis_objects(),
             add_name=('chassis',),
             order=('chassis',),
             feedback='Adding port map to chassis: ',
             control=_chassis_ports_control_d,
             obj_name=_chassis_name,
             rs=True),
        dict(obj_l=proj_obj.r_switch_objects(),
             add_name=('switch',),
             order=('switch',),
             feedback='Adding port map to switch: ',
             control=_switch_ports_control_d,
             obj_name=_switch_name,
             rs=True),
        dict(obj_l=[proj_obj],
             add_name=(),
             order=('zg', 'dup', 'bp', 'db', 'ab', 'tc'),
             feedback='Processing: ',
             control=_proj_control_d,
             obj_name=_project_name,
             rs=True),
    )

    # Add the pages to the report
    for report_d in report_l:
        if report_d.get('rs', False):
            sheet_index = 0
        for obj in report_d['obj_l']:
            start_i = sheet_index
            obj_name = report_d['obj_name'](obj)
            control_d = obj.r_get('report_app/control')
            if isinstance(control_d, dict):  # It should always be a dict. This is just belt and suspenders
                for ctl_d in [control_d[k] for k in report_d['add_name']]:  # Add the object name to the titles
                    ctl_d['t'] += obj_name
            brcdapi_log.log(report_d['feedback'] + obj_name, echo=True)
            for ctl_d in [report_d['control'][k] for k in report_d['order']]:
                sheet_index += ctl_d['a'](obj,
                                          wb,
                                          start_i if 'sf' in ctl_d and ctl_d['sf'] else sheet_index)

    # Save the report.
    brcdapi_log.log('Saving ' + outf, echo=True)
    try:
        excel_util.save_report(wb, outf)
    except PermissionError:
        brcdapi_log.log(['', 'Permission error writing ' + outf + '. File may be open in another application.'],
                        echo=True)
        proj_obj.s_error_flag()
    except FileNotFoundError:
        brcdapi_log.log(['', 'Write report failed. Folder in ' + outf + ' does not exist.'], echo=True)
        proj_obj.s_error_flag()

    return
