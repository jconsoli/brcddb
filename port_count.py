#!/usr/bin/python
# -*- coding: utf-8 -*-
# Copyright 2022, 2023 Jack Consoli.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
:mod:`port_count` - Creates a port count & type report in Excel Workbook format from a brcddb project

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 1.0.0     | 23 Jun 2022   | Initial Launch                                                                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 1.0.1     | 25 Jul 2022   | Added UCS counter                                                                 |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 1.0.2     | 01 Jan 2023   | Added additional summaries and totals.                                            |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 1.0.3     | 11 Feb 2023   | Added login summaries.                                                            |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""

__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2022, 2023 Jack Consoli'
__date__ = '11 Feb 2023'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack.consoli@broadcom.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '1.0.3'

import argparse
import copy
import openpyxl as xl
import openpyxl.utils.cell as xl_util
import collections
import brcddb.brcddb_project as brcddb_project
import brcdapi.log as brcdapi_log
import brcdapi.excel_fonts as excel_fonts
import brcdapi.excel_util as excel_util
import brcddb.util.search as brcddb_search
import brcddb.brcddb_common as brcddb_common
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_switch as brcddb_switch
import brcddb.util.file as brcddb_file
import brcddb.util.util as brcddb_util

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
_DEBUG = False   # When True, use _DEBUG_xxx below instead of parameters passed from the command line.
_DEBUG_i = 'test/180604'
_DEBUG_o = 'test/test_pc_out'
_DEBUG_log = '_logs'
_DEBUG_nl = False

# _switch_d is a template copied for each switch to track the number of login types, speed, and SFP type. _speed and
# 0 in _speed is used for unknown login speeds.
_speed = {0: 0, 1: 0, 2: 0, 4: 0, 8: 0, 16: 0, 32: 0, 64: 0}
_speed_keys = [int(i) for i in _speed.keys()]
_speed_keys.sort()
# _fc4_types are templates copied in _switch_d.
_fc4_types = collections.OrderedDict()  # Ordered so that all reports are consistent
_fc4_types['isl'] = dict(other=dict(h='Other', s=()))
_fc4_types['targets'] = dict(
    ibm=dict(h='IBM', s=('ibm',)),
    hv=dict(h='Hitachi-Vantara', s=('hitachi', 'hv')),
    dell=dict(h='Dell-EMC', s=('dell', 'emc')),
    pure=dict(h='Pure', s=('pure',)),
    netapp=dict(h='NetApp', s=('netapp',)),
    hpe=dict(h='HPE', s=('hp', '3par')),
    other=dict(h='Other', s=tuple())
)
_fc4_types['initiators'] = dict(
    e=dict(h='Emulex', s=('emulex', 'lpe')),
    q=dict(h='QLogic', s=('qlog', 'qle')),
    ucs=dict(h='UCS', s=('cisco', 'fcoe')),
    other=dict(h='Other', s=tuple())
)
_fc4_types['unknown'] = dict(other=dict(h='Other', s=tuple()))
_node_symb_key = 'brocade-name-server/fibrechannel-name-server/node-symbolic-name'

_switch_d = dict()
for _fc4_k, _fc4_d in _fc4_types.items():
    _d0 = dict()
    _switch_d.update({_fc4_k: _d0})
    for _type in ('lwl', 'swl'):
        _d1 = dict()
        _d0.update({_type: _d1})
        for _key in _fc4_d.keys():
            _d1.update({_key: copy.copy(_speed)})

_border = excel_fonts.border_type('thin')
_wrap = excel_fonts.align_type('wrap')
_wrap_center = excel_fonts.align_type('wrap_center')
_text_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_hdr_font = excel_fonts.font_type('hdr_1')

def _sheet_for_report(wb, sheet_index, sname, stitle):
    """Creates a work sheet with margins and header

    :param wb: Workbook object
    :type wb: class
    :param sheet_index: Where to insert the sheet in the workbook
    :type sheet_index: int
    :param sname: Sheet tab name
    :type sname: str
    :param stitle: Sheet title
    :type stitle: str
    """
    global _speed_keys, _border, _wrap, _wrap_center, _hdr_font

    # Create the worksheet
    sheet = wb.create_sheet(index=sheet_index, title=sname)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.freeze_panes = sheet['A2']

    # Set up the column widths and add the headers
    row = col = 1
    for col_width in (5, 30):
        sheet.column_dimensions[xl_util.get_column_letter(col)].width = col_width
        excel_util.cell_update(sheet, row, col, None, font=_hdr_font, align=_wrap, border=_border)
        col += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet['A1'] = stitle
    for i in _speed_keys:
        sheet.column_dimensions[xl_util.get_column_letter(col)].width = 6
        excel_util.cell_update(sheet, row, col, i, font=_hdr_font, align=_wrap, border=_border)
        col += 1
    sheet.column_dimensions[xl_util.get_column_letter(col)].width = 8
    excel_util.cell_update(sheet, row, col, 'Total', font=_hdr_font, align=_wrap_center, border=_border)

    return sheet


def _summary_sheet(proj_obj, wb, summary_l):
    """Creates a summary sheet

    :param proj_obj: Project object
    :type proj_obj: brcddb.classes.project.ProjectObj
    :param wb: openpyxl workbook object
    :type wb: Workbook
    :param summary_l: List of fabric dictionaries. See fab_sum_l in _create_report()
    :type summary_l: list()
    """
    global _speed, _speed_keys, _fc4_types, _border, _wrap, _text_font, _bold_font

    # Initialize local variables
    sheet = _sheet_for_report(wb, 0, 'Summary', 'Summary')
    total_sum = list()
    row = 2

    # Get a summary of all the brands for each FC4 type. Note that there may be different brands on different switches.
    sum_d = dict()
    for fc4_type in _fc4_types.keys():
        sum_d.update({fc4_type: dict()})

    for fc4_type in _fc4_types.keys():
        brand_l = list()
        for fab_d in summary_l:
            for switch_d in fab_d['switch_l']:
                brand_l.extend([str(b) for b in switch_d[fc4_type].keys()])
        for brand in brcddb_util.remove_duplicates(brand_l):
            sum_d[fc4_type].update({brand: dict()})
            for speed in _speed_keys:
                sum_d[fc4_type][brand].update({speed: list()})

    # Now get all the cell references
    for fab_d in summary_l:
        for switch_d in fab_d['switch_l']:
            for fc4_type in _fc4_types.keys():
                fc4_type_d = switch_d[fc4_type]
                for brand in sum_d[fc4_type].keys():
                    for speed in _speed_keys:
                        try:
                            sum_d[fc4_type][brand][speed].append(fab_d['sheet'] + '!' + fc4_type_d[brand][speed])
                        except KeyError:
                            pass  # Not all fabrics have the same list of brands

    for fc4_type in _fc4_types:

        # Add the FC4 type header
        row, col = row + 1, 1
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        excel_util.cell_update(sheet, row, col, fc4_type, font=_bold_font, align=_wrap, border=_border)
        row += 1
        type_start_row = row

        # Add the brands
        fc4_type_d = sum_d[fc4_type]
        for brand in fc4_type_d.keys():
            brand_d = fc4_type_d.get(brand)
            if brand_d is None:
                continue

            # Add the brand name to the worksheet
            col = 2
            excel_util.cell_update(sheet, row, col, str(brand), font=_text_font, align=_wrap, border=_border)

            # Add each speed within the brand
            for speed in _speed_keys:
                col += 1
                val = 0 if len(brand_d[speed]) == 0 else '=' + '+'.join(brand_d[speed])
                excel_util.cell_update(sheet, row, col, val, font=_text_font, align=_wrap, border=_border)
            buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
            excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)
            row += 1

        # Add the sub-total for each FC4 type
        type_end_row, col = row, 2
        row += 1
        total_sum.append(row)
        excel_util.cell_update(sheet, row, col, 'sub-total', font=_text_font, align=_wrap, border=_border)
        for k3 in range(0, len(_speed_keys)):
            col += 1
            col_letter = xl_util.get_column_letter(col)
            buf = '=sum(' + col_letter + str(type_start_row) + ':' + col_letter + str(type_end_row) + ')'
            excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
        buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
        excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)
        row += 1

    # Add a project summary
    row, col = row + 1, 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    excel_util.cell_update(sheet, row, col, 'Project total', font=_bold_font, align=_wrap)
    buf = '=$' + str(total_sum.pop(0))
    for i in total_sum:
        buf += '+$' + str(i)
    col += 2
    for k3 in range(0, len(_speed_keys)):
        excel_util.cell_update(sheet, row, col, buf.replace('$', xl_util.get_column_letter(col)), font=_bold_font,
                               align=_wrap, border=_border)
        col += 1
    buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col-1) + str(row) + ')'
    excel_util.cell_update(sheet, row, col, buf, font=_bold_font, align=_wrap, border=_border)

    # Add a login summary
    row, col = row + 2, 1
    excel_util.cell_update(sheet, row, col, 'Total Logins', font=_bold_font, align=_wrap, border=_border)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    col += 2
    excel_util.cell_update(sheet, row, col, len(proj_obj.r_login_keys()), font=_bold_font, align=_wrap, border=_border)


def _update_switch_in_fab_sheet(row, sheet, sub_hdr, switch_d):
    """Adds the port counts for a switch in a fabric worksheet

    :param row: Starting row
    :type row: int
    :param sheet: Worksheet object
    :type sheet: class
    :param sub_hdr: Section header
    :type sub_hdr: str
    :param switch_d: Dictionary of port counts by port and login type
    :type switch_d: dict
    :return row: Last row used on worksheet
    :rtype row: int
    :return rd: Fabric level dictionary. See fab_sum_l in _create_report()
    :rtype rd: dict
    """
    global _speed_keys, _fc4_types, _border, _wrap, _text_font, _bold_font

    switch_sum = list()  # A list of row numbers for each type summary
    rd = collections.OrderedDict()
    col = 1

    # Add the sub-header. Typically the switch name
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=2+len(_speed_keys))
    excel_util.cell_update(sheet, row, col, sub_hdr, font=excel_fonts.font_type('hdr_2'), align=_wrap)
    row += 1

    # The report is a list of ports based on the FC4 type (ISL, target, initiator, and unknown) associated with the
    # login for each switch. Within each FC4 type the sub-categories are the SFP type (SWL or LWL) then brand. Note that
    # SFP type and brand are combined into a single cell. The next category is the speed which is displayed in the
    # columns.
    for k0 in _fc4_types.keys():

        d0 = switch_d.get(k0)
        if d0 is None:
            continue  # When I wrote this, d0 should never be None. This is just future proofing.
        col, row = 1, row + 1
        rd.update({k0: dict()})
        excel_util.cell_update(sheet, row, col, str(k0), font=_bold_font, align=_wrap)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        row += 1
        type_start_row = row  # Used for the formula in the type sub-total

        for k1, d1 in d0.items():  # Wavelength (LWL & SWL)

            for k2, d2 in d1.items():  # Brand
                col = 2
                buf = str(k1).upper() + ', ' + _fc4_types[k0][k2]['h']
                excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
                brand_d = dict()
                rd[k0].update({buf: brand_d})

                for k3 in _speed_keys:  # Login speed
                    col += 1
                    excel_util.cell_update(sheet, row, col, d2[k3], font=_text_font, align=_wrap, border=_border)
                    brand_d.update({k3: xl_util.get_column_letter(col) + str(row)})
                buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
                excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)
                row += 1

        # Add the sub-total for each FC4 type
        type_end_row, row, col = row, row + 1, 2
        switch_sum.append(row)
        excel_util.cell_update(sheet, row, col, 'sub-total', font=_text_font, align=_wrap, border=_border)
        for k3 in range(0, len(_speed_keys)):
            col += 1
            col_letter = xl_util.get_column_letter(col)
            buf = '=sum(' + col_letter + str(type_start_row) + ':' + col_letter + str(type_end_row) + ')'
            excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
        buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
        excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)
        row += 1

    # Add the totals for the switch
    row, col = row + 1, 1
    excel_util.cell_update(sheet, row, col, 'Switch total', font=_bold_font, align=_wrap)
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    buf = '=$' + str(switch_sum.pop(0))
    for i in switch_sum:
        buf += '+$' + str(i)
    col += 2
    for col in range(col, len(_speed_keys)+col):
        col_letter = xl_util.get_column_letter(col)
        excel_util.cell_update(sheet, row, col, buf.replace('$', col_letter), font=_bold_font, align=_wrap,
                               border=_border)
    buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col-1) + str(row) + ')'
    excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)

    return row, rd


def _create_report(proj_obj, file_name):
    """ Create an Excel workbook with the port counts by fabric

    :param proj_obj: Project object
    :type proj_obj: brcddb.classes.project.ProjectObj
    :param file_name: Name of Excel workbook
    :type file_name: str
    :return: Exit code
    :rtype: int
    """
    global _speed_keys, _fc4_types, _border, _wrap, _bold_font, _text_font

    wb = xl.Workbook()
    sheet_index = 0
    ec = brcddb_common.EXIT_STATUS_OK
    fab_sum_l = list()  # Used to reference cells for summaries and totals. A list of dictionaries as follows:
    """
    +---------------+-------+-----------------------------------------------------------+
    | sheet         | str   | Sheet name                                                |
    +---------------+-------+-----------------------------------------------------------+
    | switch_l      | list  | List of switch dictionaries.                              |
    +---------------+-------+-----------------------------------------------------------+
    
    Switch dictionaries:
    
    +---------------+-------+-----------------------------------------------------------+
    | isl           | dict  | Key is the brand. The value is a dictionary as follows:   |
    |               |       | Key   Value                                               |
    |               |       | 0     Cell reference for speed 0 (unknown speed)          |
    |               |       | 1     Cell reference for speed 1 Gbps                     |
    |               |       | etc. - one entry for each speed in _speed                 |
    +---------------+-------+-----------------------------------------------------------+
    | target        | dict  | Same as isl                                               |
    +---------------+-------+-----------------------------------------------------------+
    | initiators    | dict  | Same as isl                                               |
    +---------------+-------+-----------------------------------------------------------+
    | unknown       | dict  | Same as isl                                               |
    +---------------+-------+-----------------------------------------------------------+
    """
    for fab_obj in proj_obj.r_fabric_objects():  # Add a sheet for each fabric

        # Create the fabric sheet
        sname = excel_util.valid_sheet_name.sub('', brcddb_fabric.best_fab_name(fab_obj).replace(' ', '_'))[:22] + \
                '_' + str(sheet_index)
        fab_name = brcddb_fabric.best_fab_name(fab_obj, wwn=False)
        sheet = _sheet_for_report(wb, sheet_index, sname, fab_name)
        row = 3

        # Set up the tracker for the summary sheet
        fab_sum = list()  # A list of row numbers for each type summary

        # Add the switches
        switch_l, fab_switch_ref_d = list(), collections.OrderedDict()
        fab_d = dict(sheet=sname, switch_l=switch_l, fab_switch_ref_d=fab_switch_ref_d)
        for switch_obj in fab_obj.r_switch_objects():
            row, cell_d = _update_switch_in_fab_sheet(row,
                                                      sheet,
                                                      brcddb_switch.best_switch_name(switch_obj, wwn=True),
                                                      switch_obj.r_get('switch_d'))
            fab_sum.append(row)
            switch_l.append(cell_d)
            row += 2
        fab_sum_l.append(fab_d)

        # Add the header for the fabric summary at the bottom of the page.
        col = 1
        excel_util.cell_update(sheet, row, col, 'Fabric total', font=_bold_font, align=_wrap)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

        # Adding the same level of detail to the the fabric section as is with switches was an after thought. Creating a
        # dict that matches the switch dict with a list of switch references kept it simple.
        cell_d_l = fab_d['switch_l']
        if len(cell_d_l) > 0:
            cell_d = cell_d_l[0]
            for k, switch_type_d in cell_d.items():
                fab_type_d = dict()
                fab_switch_ref_d.update({k: fab_type_d})
                for k0, switch_ref_d in switch_type_d.items():
                    fab_ref_d = dict()
                    fab_type_d.update({k0: fab_ref_d})
                    for k1, cell in switch_ref_d.items():
                        fab_ref_d.update({k1: '=' + cell})

        # Get the remaining switch cell references for the fabric
        if len(cell_d_l) > 1:
            for cell_d in cell_d_l[1:]:
                for k, switch_type_d in cell_d.items():
                    fab_type_d = fab_switch_ref_d[k]
                    for k0, switch_ref_d in switch_type_d.items():
                        fab_ref_d = fab_type_d[k0]
                        for k1, cell in switch_ref_d.items():
                            fab_ref_d[k1] += '+' + cell

        # Add the fabric summaries to the worksheet

        # fab_total_d is a list of cell references to be added together for the fabric total
        fab_total_d = collections.OrderedDict()
        for k in _speed_keys:
            fab_total_d.update({k: list()})

        # Fill in the fabric section
        for k, fab_type_d in fab_switch_ref_d.items():
            row, col = row+2, 1
            excel_util.cell_update(sheet, row, col, str(k), font=_bold_font, align=_wrap)
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            row, col = row+1, 2
            start_row = row
            for k0, cell_d in fab_type_d.items():
                excel_util.cell_update(sheet, row, col, str(k0), font=_text_font, align=_wrap, border=_border)
                for buf in [cell_d[k1] for k1 in _speed_keys]:
                    col += 1
                    excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
                buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
                excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)
                row, col = row+1, 2

            # Add the sub-total
            row, col = row+1, 2
            excel_util.cell_update(sheet, row, col, 'fabric sub-total', font=_text_font, align=_wrap, border=_border)
            for speed_k in _speed_keys:
                col += 1
                col_letter = xl_util.get_column_letter(col)
                buf = '=sum(' + col_letter + str(start_row) + ':' + col_letter + str(row-1) + ')'
                excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
                fab_total_d[speed_k].append(col_letter + str(row))
            buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
            excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)

        # Add the fabric total
        row, col = row+2, 1
        excel_util.cell_update(sheet, row, col, 'Fabric total', font=_bold_font, align=_wrap)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        col = 2
        for k, fab_ref_d in fab_total_d.items():
            col += 1
            buf = '=' + fab_ref_d[0] + '+' + '+'.join(fab_ref_d[1:])
            excel_util.cell_update(sheet, row, col, buf, font=_text_font, align=_wrap, border=_border)
        buf = '=sum(C' + str(row) + ':' + xl_util.get_column_letter(col) + str(row) + ')'
        excel_util.cell_update(sheet, row, col+1, buf, font=_text_font, align=_wrap, border=_border)

        sheet_index += 1

    # Add a summary sheet
    _summary_sheet(proj_obj, wb, fab_sum_l)

    try:
        wb.save(file_name)
    except PermissionError:
        buf = 'Write access permission for ' + file_name + ' denied. This typically occurs when the file is open'
        brcdapi_log.log(['', buf], echo=True)
        ec = brcddb_common.EXIT_STATUS_USER_ERROR
    except FileExistsError:
        brcdapi_log.log(['', 'The folder in the path for ' + file_name + ' does not exist.'], echo=True)

    return ec


def _get_input():
    """Parses the module load command line
    
    :return: file
    :rtype: str
    """
    global _DEBUG, _DEBUG_i, _DEBUG_o, _DEBUG_log, _DEBUG_nl

    if _DEBUG:
        args_i, args_o, args_log, args_nl = _DEBUG_i, _DEBUG_o, _DEBUG_log, _DEBUG_nl
    else:
        parser = argparse.ArgumentParser(description='Create a general report in Excel.')
        buf = 'Required. Name of input file generated by capture.py, combine.py, or multi_capture.py'
        parser.add_argument('-i', help=buf, required=True)
        parser.add_argument('-o', help='Required. Name of output file. ".xlsx" is automatically appended.',
                            required=True)
        buf = '(Optional) Directory where log file is to be created. Default is to use the current directory. The ' \
              'log file name will always be "Log_xxxx" where xxxx is a time and date stamp.'
        parser.add_argument('-log', help=buf, required=False, )
        buf = '(Optional) No parameters. When set, a log file is not created. The default is to create a log file.'
        parser.add_argument('-nl', help=buf, action='store_true', required=False)
        args = parser.parse_args()
        args_i, args_o, args_log, args_nl = args.i, args.o, args.log, args.nl

    # Set up the log file
    if not args_nl:
        brcdapi_log.open_log(args_log)

    return brcddb_file.full_file_name(args_i, '.json'), brcddb_file.full_file_name(args_o, '.xlsx')


def psuedo_main():
    """Basically the main(). Did it this way so it can easily be used as a standalone module or called from another.

    :return: Exit code. See exist codes in brcddb.brcddb_common
    :rtype: int
    """
    global _DEBUG, __version__, _speed, _switch_d, _node_symb_key

    ec = brcddb_common.EXIT_STATUS_OK

    # Get and validate user input
    inf, outf = _get_input()
    ml = ['WARNING!!! Debug is enabled'] if _DEBUG else list()
    ml.append('Version:  ' + __version__)
    ml.append('In file:  ' + inf)
    ml.append('Out file: ' + outf)
    brcdapi_log.log(ml, echo=True)

    # Get a project object
    try:
        proj_obj = brcddb_project.read_from(inf)
    except FileNotFoundError:
        brcdapi_log.log(['', inf + ' not found.'], echo=True)
        return brcddb_common.EXIT_STATUS_USER_ERROR
    if proj_obj is None:
        brcdapi_log.log(['', 'Unknown error reading: ' + inf], echo=True)
        return brcddb_common.EXIT_STATUS_ERROR
    brcddb_project.build_xref(proj_obj)
    brcddb_project.add_custom_search_terms(proj_obj)

    # Determine the number of logins for each media type and speed
    for fab_obj in proj_obj.r_fabric_objects():

        # Set up the speed tracking for each switch in each fabric
        for switch_obj in fab_obj.r_switch_objects():
            switch_d = copy.deepcopy(_switch_d)
            switch_obj.s_new_key('switch_d', switch_d)

            # Now figure out which port dictionary to use
            for port_obj in switch_obj.r_port_objects():
                # Default keys
                k0 = 'unknown'  # Basic FC4 type: isl, targets, initiators, or unknown
                wave_length = port_obj.r_get('media-rdp/wavelength')
                k1 = 'lwl' if isinstance(wave_length, int) and wave_length > 1300 else 'swl'
                k2 = 'other'  # brand
                k3 = 0  # Speed

                # k0 - Login type
                buf = brcddb_port.port_type(port_obj)
                if 'E-Port' in buf:
                    k0 = 'isl'
                elif 'F-Port' in buf or 'N-Port' in buf:

                    # Is it a target or initiator?
                    all_logins_l = port_obj.r_login_objects()
                    if len(brcddb_search.match_test(all_logins_l, brcddb_search.target)) > 0:
                        k0 = 'targets'
                    elif len(brcddb_search.match_test(all_logins_l, brcddb_search.initiator)) > 0:
                        k0 = 'initiators'

                    # Determine the brand, k2.
                    fc4_d = _fc4_types.get(k0)
                    if isinstance(fc4_d, dict):
                        for login_obj in all_logins_l:
                            if k2 != 'other':
                                break

                            buf = login_obj.r_get(_node_symb_key)
                            if isinstance(buf, str):
                                buf = buf.lower()
                                for k, d in fc4_d.items():
                                    if k2 != 'other':
                                        break
                                    for search_term in d['s']:
                                        if search_term in buf:
                                            k2 = k
                                            break

                # k3 - Speed
                login_speed = port_obj.r_get('_search/speed')
                if isinstance(login_speed, int) and login_speed in _speed:
                    k3 = login_speed

                # Increment the port counter
                switch_d[k0][k1][k2][k3] += 1

    # Generate the report and return
    return _create_report(proj_obj, outf)


##################################################################
#
#                    Main Entry Point
#
###################################################################
if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(0)

_ec = psuedo_main()
brcdapi_log.close_log(['', 'All processing complete. Exit code: ' + str(_ec)], echo=True)
exit(_ec)
